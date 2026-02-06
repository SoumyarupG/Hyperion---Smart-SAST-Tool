import os
import psycopg2
from datetime import datetime, timedelta, timezone, date
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
import logging

# === Logging Setup ===
os.makedirs("logs", exist_ok=True)
LOG_FILE = "logs/forecast_file.log"
logging.basicConfig(filename=LOG_FILE, level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# === DB Config ===
db_params = {
    'dbname': 'demand_forecast_states',
    'user': 'prasadbabu',
    'password': 'BabuPrasad#123',
    'host': '10.0.100.79',
    'port': '5432'
}

# === Excel Templates ===
shared_drive_path = r"D:\scripts\DFD\RAW FILES"
file_templates = {
    "day": os.path.join(shared_drive_path, "day_ahead_consolidated.xlsx"),
    "week": os.path.join(shared_drive_path, "week_ahead_consolidated.xlsx"),
    "month": os.path.join(shared_drive_path, "month_ahead_consolidated.xlsx"),
    "year": os.path.join(shared_drive_path, "year_ahead_consolidated.xlsx"),
}

state_map = {1: 'KAR', 2: 'TN', 3: 'TG', 4: 'AP', 5: 'KER', 7: 'PONDY'}

# ---------------------------------------------------------------
# DB UTILITIES
# ---------------------------------------------------------------
def fetch_data(forecast_type, from_date, to_date):
    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    if forecast_type == "day":
        query = """
            SELECT t1.state_id, t1.file_data
            FROM public.file_uploads t1
            INNER JOIN (
                SELECT state_id, MAX(revision_no) AS max_rev
                FROM public.file_uploads
                WHERE upload_date = %s
                GROUP BY state_id
            ) t2 ON t1.state_id = t2.state_id AND t1.revision_no = t2.max_rev
            WHERE t1.upload_date = %s;
        """
        cur.execute(query, (from_date, from_date))
    else:
        query = f"""
            SELECT t1.state_id, t1.file_data
            FROM public.{forecast_type}_ahead_file_uploads t1
            INNER JOIN (
                SELECT state_id, from_date, to_date, MAX(revision_no) AS max_rev
                FROM public.{forecast_type}_ahead_file_uploads
                WHERE from_date = %s AND to_date = %s
                GROUP BY state_id, from_date, to_date
            ) t2 ON t1.state_id = t2.state_id
                 AND t1.from_date = t2.from_date
                 AND t1.to_date = t2.to_date
                 AND t1.revision_no = t2.max_rev;
        """
        cur.execute(query, (from_date, to_date))

    result = cur.fetchall()
    cur.close()
    conn.close()
    return result

# ---------------------------------------------------------------
# EXCEL UTILITIES
# ---------------------------------------------------------------
def replace_null_with_zero(file_data):
    return [[0 if v is None else v for v in row] for row in file_data]

def safe_write(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                break
    cell.value = value

def add_footer(sheet, file_data, start_row, is_week_like):
    total_rows = len(file_data)
    footer_row = start_row + total_rows
    label_col = 3 if is_week_like else 2

    safe_write(sheet, footer_row, label_col, "Total MUs")
    sheet.cell(row=footer_row, column=label_col).font = Font(bold=True)

    for col in range(label_col + 1, len(file_data[0]) + 1):
        total = sum((sheet.cell(row=r, column=col).value or 0)
                    for r in range(start_row, start_row + total_rows))
        sheet.cell(row=footer_row, column=col).value = round(total / 4000, 2)

def combine_state_data(sr_data, is_week_like):
    combined = []
    for rowset in zip(*sr_data):
        if is_week_like:
            base = [rowset[0][0], rowset[0][1], rowset[0][2]]
            sums = [sum(float(v or 0) for v in col) for col in zip(*[r[3:] for r in rowset])]
        else:
            base = [rowset[0][0], rowset[0][1]]
            sums = [sum(float(v or 0) for v in col) for col in zip(*[r[2:] for r in rowset])]
        combined.append(base + sums)
    return combined

def write_excel(forecast_type, db_data, from_date, to_date):
    template_path = file_templates[forecast_type]
    wb = load_workbook(template_path)
    is_week_like = forecast_type != "day"
    sr_data = []

    for state_id, file_data in db_data:
        state = state_map[state_id]
        ws = wb[state]

        ws["C2"].value = state
        ws["C3"].value = f"{from_date} to {to_date}" if is_week_like else from_date

        file_data = replace_null_with_zero(file_data)
        start_row = 8
        for i, row in enumerate(file_data):
            for j, val in enumerate(row):
                safe_write(ws, start_row + i, j + 1, val)

        add_footer(ws, file_data, start_row, is_week_like)
        sr_data.append(file_data)

    sr = wb["SR"]
    #  Write SR name and date header (previously missing)
    sr["C2"].value = "SR"
    sr["C3"].value = f"{from_date} to {to_date}" if is_week_like else from_date
    combined = combine_state_data(sr_data, is_week_like)
    for i, row in enumerate(combined):
        for j, val in enumerate(row):
            safe_write(sr, 8 + i, j + 1, val)

    add_footer(sr, combined, 8, is_week_like)

    #  Proper formatted file name
    fd = from_date.strftime("%d-%m-%Y")
    td = to_date.strftime("%d-%m-%Y")
    suffix = f"{fd}_to_{td}" if is_week_like else fd
    save_path = template_path.replace(".xlsx", f"_consolidated_{suffix}.xlsx")

    wb.save(save_path)
    return save_path

# ---------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------
def generate_consolidated_forecast_file(forecast_type, from_date, to_date):
    try:
        # Convert ISO → IST Date
        def to_ist(d):
            if isinstance(d, date): return d
            dt = datetime.fromisoformat(d.replace("Z","+00:00"))
            return dt.astimezone(timezone(timedelta(hours=5,minutes=30))).date()

        from_date, to_date = to_ist(from_date), to_ist(to_date)

        db_data = fetch_data(forecast_type, from_date, to_date)
        if not db_data:
            return {"status":"failure","message":"No data found for given date(s)"}

        required = {1,2,3,4,5,7}
        uploaded = {state for state,_ in db_data}
        if not required.issubset(uploaded):
            missing = ", ".join(state_map[s] for s in (required - uploaded))
            return {"status":"failure","message":f"Missing uploads from: {missing}"}

        return {"status":"success","file_path":write_excel(forecast_type, db_data, from_date, to_date)}

    except Exception as e:
        logger.error(str(e), exc_info=True)
        return {"status":"failure","message":str(e)}
