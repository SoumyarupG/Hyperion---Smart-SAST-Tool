import os
import psycopg2
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import logging

LOG_FILE = "mail_status_w_m_y.log"
logging.basicConfig(
    filename=LOG_FILE,
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger()

db_params = {
    'dbname': 'demand_forecast_states',
    'user': 'prasadbabu',
    'password': 'BabuPrasad#123',
    'host': '10.0.100.79',
    'port': '5432'
}

# send_mails_list_db = ['nldcforecasting@grid-india.in', 'amangautam@grid-india.in']
# cc_mails_list_db = ['srldcos@grid-india.in']

send_mails_list_db = ['mprashad@grid-india.in']
cc_mails_list_db = ['srldcos@grid-india.in']

file_paths = {
    'day': r'D:\scripts\DFD\RAW FILES\day_ahead_consolidated.xlsx',
    'week': r'D:\scripts\DFD\RAW FILES\week_ahead_consolidated.xlsx',
    'month': r'D:\scripts\DFD\RAW FILES\month_ahead_consolidated.xlsx',
    'year': r'D:\scripts\DFD\RAW FILES\year_ahead_consolidated.xlsx'
}


state_sheet_names = ['KAR', 'TN', 'TG', 'AP', 'KER', 'PONDY', 'SR']

def sendMail(send_mails_list_db, message):
    mail_id = 'srldcos'
    mail_pwd = 'OS@Opra#$1234'
    port = 587
    smtp_server = "mail.grid-india.in"
    mail_server = smtplib.SMTP(smtp_server, port)
    mail_server.starttls()
    mail_server.login(mail_id, mail_pwd)
    mail_server.sendmail('srldcos@grid-india.in', send_mails_list_db, message.as_string())

def send_mail_with_attachment(forecast_type,send_mails_list_db,cc_mails_list_db,file_path,custom_from_date,custom_to_date=None):
    # ---------- SUBJECT ----------
    if forecast_type == "day":
        subject = (
            f"Southern Region Day Ahead Demand Forecast "
            f"for {custom_from_date.strftime('%d-%m-%Y')}"
        )
    elif forecast_type == "week":
        subject = (
            f"Southern Region Week Ahead Demand Forecast "
            f"from {custom_from_date.strftime('%d-%m-%Y')} "
            f"to {custom_to_date.strftime('%d-%m-%Y')}"
        )
    elif forecast_type == "month":
        subject = (
            f"Southern Region Month Ahead Demand Forecast "
            f"from {custom_from_date.strftime('%d-%m-%Y')} "
            f"to {custom_to_date.strftime('%d-%m-%Y')}"
        )
    else:  # year
        subject = (
            f"Southern Region Year Ahead Demand Forecast "
            f"from {custom_from_date.strftime('%d-%m-%Y')} "
            f"to {custom_to_date.strftime('%d-%m-%Y')}"
        )

    # ---------- BODY ----------
    if forecast_type == "day":
        msg_body = f"""
        Sir/Madam,<br><br>
        Please find the Southern Region <b>Day Ahead Demand Forecast</b>
        for <b>{custom_from_date.strftime('%d-%m-%Y')}</b>,
        as received from the states.<br><br>
        """
    else:
        msg_body = f"""
        Sir/Madam,<br><br>
        Please find the Southern Region <b>{forecast_type.capitalize()} Ahead Demand Forecast</b>
        for the period from <b>{custom_from_date.strftime('%d-%m-%Y')}</b>
        to <b>{custom_to_date.strftime('%d-%m-%Y')}</b>,
        as received from the states.<br><br>
        """

    msg_body += """
    धन्यवाद एवं आभार / Thanks and Regards,<br>
    --<br>
    SO-II / एस.ओ-II<br>
    SRLDC, Bengaluru<br>
    Grid Controller of India Limited
    """

    message = MIMEMultipart()
    message["Subject"] = subject
    message["From"] = 'srldcos@grid-india.in'
    message["To"] = ', '.join(send_mails_list_db)
    message["Cc"] = ', '.join(cc_mails_list_db)
    message.attach(MIMEText(msg_body, 'html'))

    with open(file_path, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename={os.path.basename(file_path)}'
        )
        message.attach(part)

    sendMail(send_mails_list_db + cc_mails_list_db, message)

def get_next_ranges():
    today = datetime.today()
    next_monday = today + timedelta(days=(7 - today.weekday()))
    next_sunday = next_monday + timedelta(days=6)
    next_month_start = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
    next_month_end = (next_month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    fy_start_year = today.year + 1 if today.month >= 4 else today.year
    fy_start = datetime(fy_start_year, 4, 1)
    fy_end = datetime(fy_start_year + 1, 3, 31)
    return {
        'week': (next_monday, next_sunday),
        'month': (next_month_start, next_month_end),
        'year': (fy_start, fy_end)
    }


# === Utility: Create empty Excel template if not present ===
def ensure_excel_template_exists(file_path):
    if not os.path.exists(file_path):
        logger.info(f"Template file not found at {file_path}. Creating a new one...")
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        for sheet_name in state_sheet_names:
            ws = wb.create_sheet(title=sheet_name)
            ws['C2'] = ''
            ws['C3'] = ''
            ws.append(['Date', 'Block', 'Period'] + [f'Col{i}' for i in range(1, 97)])
        wb.save(file_path)
        logger.info(f"Created new Excel template at {file_path}")

# === DB Mail Tracking ===
def check_mail_sent(forecast_type, from_date, to_date):
    logger.info(f"Checking if mail already sent for {forecast_type} from {from_date} to {to_date}...")
    if isinstance(from_date, datetime):
        from_date = from_date.date()
    if isinstance(to_date, datetime):
        to_date = to_date.date()
    query = """
        SELECT 1 FROM public.forecast_mail_status
        WHERE forecast_type = %s AND from_date = %s AND to_date = %s
        LIMIT 1
    """
    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()
    cur.execute(query, (forecast_type, from_date, to_date))
    result = cur.fetchone()
    print("Result", result)
    cur.close()
    conn.close()
    return result is not None

def mark_mail_sent(forecast_type, from_date, to_date):
    logger.info(f"Marking mail as sent for {forecast_type} from {from_date} to {to_date}...")
    query = """
        INSERT INTO public.forecast_mail_status (forecast_type, from_date, to_date)
        VALUES (%s, %s, %s)
        ON CONFLICT DO NOTHING
    """
    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()
    cur.execute(query, (forecast_type, from_date, to_date))
    conn.commit()
    cur.close()
    conn.close()

# === Excel Utils ===
def replace_null_with_zero(file_data):
    return [[0 if value is None else value for value in row] for row in file_data]

def get_state_name(state_id):
    state_map = {1: 'KAR', 2: 'TN', 3: 'TG', 4: 'AP', 5: 'KER', 6: 'SR', 7: 'PONDY'}
    return state_map.get(state_id, '')

def add_footer(sheet, file_data, start_row):
    if not file_data:
        logger.info("No data to add footer.")
        return
    last_row = start_row + len(file_data) - 1
    footer_row = last_row + 1
    cell = sheet.cell(row=footer_row, column=3)
    if type(cell).__name__ == 'MergedCell':
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    cell.value = "Total MUs"
    cell.font = Font(bold=True)
    for col in range(4, len(file_data[0]) + 1):
        sum_value = sum(sheet.cell(row=row, column=col).value or 0 for row in range(start_row, last_row + 1))
        sheet.cell(row=footer_row, column=col).value = round(sum_value / 4000, 2)

def combine_state_data(sr_data, forecast_type):
    """
    Properly sum state data into SR.
    Handles DAY differently from WEEK/MONTH/YEAR.
    """
    if not sr_data:
        return []

    combined = []

    for rows in zip(*sr_data):
        # ---- DAY AHEAD ----
        if forecast_type == "day":
            # Columns: [Block, Period, MW, MW, MW, ...]
            base = rows[0][:2]  # Block, Period
            summed = [
                sum(float(r[i] or 0) for r in rows)
                for i in range(2, len(rows[0]))
            ]
            combined.append(base + summed)

        # ---- WEEK / MONTH / YEAR ----
        else:
            base = rows[0][:3]  # Date, Block, Period
            summed = [
                sum(float(r[i] or 0) for r in rows)
                for i in range(3, len(rows[0]))
            ]
            combined.append(base + summed)

    return combined


def safe_write(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                sheet.unmerge_cells(str(merged_range))
                break
        cell = sheet.cell(row=row, column=col)
    cell.value = value

def pad_file_data(file_data, target_length):
    current_length = len(file_data)
    if current_length < target_length:
        width = len(file_data[0]) if file_data else 0
        pad_row = ['' for _ in range(width)]
        for _ in range(target_length - current_length):
            file_data.append(pad_row.copy())
    return file_data

def write_to_excel_generic(file_path, db_data, from_date, to_date, forecast_type):
    ensure_excel_template_exists(file_path)

    if isinstance(from_date, datetime):
        from_date = from_date.date()
    if isinstance(to_date, datetime):
        to_date = to_date.date()

    logger.info(f"Generating Excel for {forecast_type.upper()} from {from_date} to {to_date}...")
    wb = load_workbook(file_path)

    sr_data = []
    max_rows = max((len(fd) for _, fd in db_data), default=0)

    expected_state_ids = {1, 2, 3, 4, 5, 7}
    uploaded_state_ids = {sid for sid, _ in db_data}

    # ---- Missing state check ----
    if not expected_state_ids.issubset(uploaded_state_ids):
        missing = expected_state_ids - uploaded_state_ids
        logger.info(f"[{forecast_type.upper()}] Not all states have uploaded data.")
        for sid in sorted(missing):
            logger.info(f"  - {get_state_name(sid)}")
        return None

    # =========================
    # STATE SHEETS
    # =========================
    for state_id, file_data in db_data:
        state_name = get_state_name(state_id)

        if state_name and state_name in wb.sheetnames:
            logger.info(f"Writing data for {state_name}")
            sheet = wb[state_name]

            safe_write(sheet, 2, 3, state_name)

            # ---- Date header ----
            if forecast_type == "day":
                safe_write(sheet, 3, 3, from_date.strftime("%Y-%m-%d"))
            else:
                safe_write(
                    sheet,
                    3,
                    3,
                    f"{from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')}"
                )

            file_data = replace_null_with_zero(file_data)
            file_data = pad_file_data(file_data, max_rows)

            start_row = 8
            for i, row_data in enumerate(file_data):
                for j, value in enumerate(row_data):
                    safe_write(sheet, start_row + i, j + 1, value)

            sr_data.append(file_data)
            add_footer(sheet, file_data, start_row)

    # =========================
    # SR CONSOLIDATED SHEET
    # =========================
    sr_sheet = wb['SR']
    logger.info("Writing SR consolidated data")

    safe_write(sr_sheet, 2, 3, "SR")

    if forecast_type == "day":
        safe_write(sr_sheet, 3, 3, from_date.strftime("%Y-%m-%d"))
    else:
        safe_write(
            sr_sheet,
            3,
            3,
            f"{from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')}"
        )

    combined_sr_data = combine_state_data(sr_data, forecast_type)

    for i, row_data in enumerate(combined_sr_data):
        for j, value in enumerate(row_data):
            safe_write(sr_sheet, 8 + i, j + 1, value)

    add_footer(sr_sheet, combined_sr_data, 8)

    # =========================
    # FINAL FILE NAME
    # =========================
    if forecast_type == "day":
        final_path = file_path.replace(
            "day_ahead_consolidated.xlsx",
            f"day_ahead_consolidated_for {from_date.strftime('%Y-%m-%d')}.xlsx"
        )
    else:
        final_path = file_path.replace(
            f"{forecast_type}_ahead_consolidated.xlsx",
            f"{forecast_type}_ahead_consolidated_for "
            f"{from_date.strftime('%Y-%m-%d')}-{to_date.strftime('%Y-%m-%d')}.xlsx"
        )

    wb.save(final_path)
    logger.info(f"Saved Excel file to {final_path}")
    return final_path

def fetch_max_revision_data(forecast_type, from_date, to_date):
    logger.info(f"Fetching max revision data from DB for {forecast_type.upper()}...")

    # Normalize to date
    if isinstance(from_date, datetime):
        from_date = from_date.date()
    if isinstance(to_date, datetime):
        to_date = to_date.date()

    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    # =========================
    # DAY AHEAD
    # =========================
    if forecast_type == "day":
        query = """
            SELECT t1.state_id, t1.file_data
            FROM public.file_uploads t1
            INNER JOIN (
                SELECT state_id, upload_date, MAX(revision_no) AS max_rev
                FROM public.file_uploads
                WHERE upload_date = %s
                GROUP BY state_id, upload_date
            ) t2
            ON t1.state_id = t2.state_id
           AND t1.upload_date = t2.upload_date
           AND t1.revision_no = t2.max_rev;
        """
        cur.execute(query, (from_date,))

    # =========================
    # WEEK / MONTH / YEAR
    # =========================
    else:
        table = f"{forecast_type}_ahead_file_uploads"
        query = f"""
            SELECT t1.state_id, t1.file_data
            FROM public.{table} t1
            INNER JOIN (
                SELECT state_id, from_date, to_date, MAX(revision_no) AS max_rev
                FROM public.{table}
                WHERE from_date = %s AND to_date = %s
                GROUP BY state_id, from_date, to_date
            ) t2
            ON t1.state_id = t2.state_id
           AND t1.from_date = t2.from_date
           AND t1.to_date = t2.to_date
           AND t1.revision_no = t2.max_rev;
        """
        cur.execute(query, (from_date, to_date))

    result = cur.fetchall()
    cur.close()
    conn.close()

    logger.info(f"Fetched {len(result)} rows of data for {forecast_type.upper()}")
    return result


def check_and_send(forecast_type, from_date, to_date):
    if check_mail_sent(forecast_type, from_date, to_date):
        logger.info(
            f"[{forecast_type.upper()}] Mail already sent for range: "
            f"{from_date} to {to_date}"
        )
        return f"Mail already sent for {forecast_type} from {from_date} to {to_date}"

    # 1️⃣ Fetch DB data
    db_data = fetch_max_revision_data(forecast_type, from_date, to_date)

    # 2️⃣ Generate Excel
    file_path = write_to_excel_generic(
        file_paths[forecast_type],
        db_data,
        from_date,
        to_date,
        forecast_type
    )

    # 3️⃣ Stop if Excel not created
    if file_path is None:
        logger.info(
            f"[{forecast_type.upper()}] Skipping mail — "
            f"not all states uploaded or Excel not created."
        )
        return (
            f"Skipping mail for {forecast_type} — "
            f"not all states uploaded or Excel not created."
        )

    # 4️⃣  SEND MAIL (THIS WAS MISSING)
    send_mail_with_attachment(
        forecast_type=forecast_type,
        send_mails_list_db=send_mails_list_db,
        cc_mails_list_db=cc_mails_list_db,
        file_path=file_path,
        custom_from_date=from_date,
        custom_to_date=None if forecast_type == "day" else to_date
    )

    # 5️⃣ Mark mail sent
    mark_mail_sent(forecast_type, from_date, to_date)

    logger.info(
        f"[{forecast_type.upper()}] Mail sent successfully "
        f"for {from_date} to {to_date}"
    )

    return f"Mail sent for {forecast_type} from {from_date} to {to_date}"


def forecast_mail_status(forecast_type, from_date, to_date):
    from datetime import datetime, timedelta
    try:
        # Parse dates if they are strings
        if isinstance(from_date, str):
            from_date = datetime.strptime(from_date, "%Y-%m-%d")
        if isinstance(to_date, str):
            to_date = datetime.strptime(to_date, "%Y-%m-%d")
        if forecast_type not in ['day', 'week', 'month', 'year']:
            return {'status': 'failure', 'message': 'Invalid forecast type'}

        # Validate perfect ranges
        if forecast_type == 'day':
            to_date = from_date

        if forecast_type == 'week':
            # Week must start on Monday and end on Sunday
            if from_date.weekday() != 0 or to_date.weekday() != 6 or (to_date - from_date).days != 6:
                return {
                    'status': 'failure',
                    'message': 'Week must start on Monday and end on Sunday (7 days).'
                }
        elif forecast_type == 'month':
            # Month must start on 1st and end on last day
            if from_date.day != 1 or (to_date.month != from_date.month) or (to_date != (from_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)):
                return {
                    'status': 'failure',
                    'message': 'Month must start on 1st and end on last day of the same month.'
                }
        elif forecast_type == 'year':
            # Financial year: April 1 to March 31 next year
            fy_start = datetime(from_date.year if from_date.month >= 4 else from_date.year - 1, 4, 1)
            fy_end = datetime(fy_start.year + 1, 3, 31)
            if from_date != fy_start or to_date != fy_end:
                return {
                    'status': 'failure',
                    'message': 'Year must be financial year: April 1 to March 31.'
                }

        result = check_and_send(forecast_type, from_date, to_date)
        return {'status': 'success', 'result': result}
    except Exception as e:
        logger.error(f"Error in forecast_mail_status: {e}")
        return {'status': 'failure', 'message': str(e)}
