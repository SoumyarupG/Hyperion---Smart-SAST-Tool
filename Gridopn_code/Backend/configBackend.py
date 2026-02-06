
from base64 import decode
from datetime import time, timedelta

from flask import Flask, request
from flask import jsonify
from flask_cors import CORS
# import ldap
# import requests

from flask_jwt_extended import create_access_token
from flask_jwt_extended import create_refresh_token
from flask_jwt_extended import get_jwt_identity
from flask_jwt_extended import jwt_required
from flask_jwt_extended import JWTManager
from flask_jwt_extended import decode_token
from openpyxl import load_workbook
from openpyxl.styles import Font
from configBackend import *

import numpy as np
import pandas as pd
import psycopg2

from collections import defaultdict

import os, sys

import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
# import time
from psycopg2 import pool



from ldap3 import Server, Connection, SIMPLE, SYNC, ALL

server = Server('10.10.100.222')





from flask_bcrypt import Bcrypt

import datetime
from datetime import timedelta, datetime

# from ldap3 import Server, Connection, SIMPLE, SYNC, ALL

# server = Server('10.0.100.218')





import logging

# Ensure the logs directory exists
if not os.path.exists('logs'):
    os.makedirs('logs')

# Set the log filename
log_filename = os.path.join('logs', 'error_logs.log')

# Configure logging
logging.basicConfig(
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s [in %(pathname)s:%(lineno)d]',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)  # This will log to the console as well
    ]
)

logger = logging.getLogger(__name__)


app = Flask(__name__)

app.config["JWT_SECRET_KEY"] = "super-secret"  # Change this!
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(days=1)
app.config['RECAPTCHA_SECRET_KEY'] = '6LftdjQqAAAAAM9YSjjc39SA562278xDqjQDnAbI'
app.config['JWT_TOKEN_LOCATION'] = ["cookies"]  #  Accept JWT from cookies
app.config['JWT_COOKIE_CSRF_PROTECT'] = False  #  Disable CSRF for API
app.config['JWT_COOKIE_SECURE'] = False  #  Allow in non-HTTPS environments
app.config['JWT_ACCESS_COOKIE_NAME'] = "access_token"
# app.config["JWT_REFRESH_TOKEN_EXPIRES"] = timedelta(minutes=5)
app.config["ALGORITHM"] = "HS256"
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500 MB
jwt = JWTManager(app)

# CORS(app)
CORS(app, supports_credentials=True)
# CORS(app, resources={r"/*": {"origins": "http://192.168.57.50:4200"}}, supports_credentials=True)
# CORS(app, supports_credentials=True, origins="http://192.168.57.50:4200")



# Database connection details
db_params = {
    'dbname': 'demand_forecast_states',
    'user': 'prasadbabu',
    'password': 'BabuPrasad#123',
    'host': '10.0.100.79',
    'port': '5432'
}

mdp_db_params = {
    'dbname': 'MDPNew',
    'user': 'prasadbabu',
    'password': 'BabuPrasad#123',
    'host': '10.0.100.219',
    'port': '5432'
}


scada_db_params = {
    'dbname': 'COMMON_DATABASE',
    'user': 'prasadbabu',
    'password': 'BabuPrasad#123',
    'host': '10.0.100.79',
    'port': '5432'
}


keys_dict = {
    3: {1: "SUBSTN.SYSC1_PG.SYSTEM.KA_THERM_GEN.ADD.MW", 2:"SUBSTN.SYSC1_PG.SYSTEM.TN_THERM_GEN.ADD.MW", 3: "SUBSTN.SYSC1_PG.SYSTEM.TS_THERM_GEN.ADD.MW",4: "SUBSTN.SYSC1_PG.SYSTEM.AP_THERM_GEN.ADD.MW", 5:"SUBSTN.SYSC1_PG.SYSTEM.KL_THERM_GEN.MES1.MW", 6: "SUBSTN.SYSC1_PG.SYSTEM.RG_THERM_TOT.ADD.MW"},
    7: {1: "SUBSTN.SYSCA_KA.SYSTEM.SOLAR_GEN_MW.MES1.MW", 2:"SUBSTN.SYSCA_TN.SYSTEM.SOLAR_TOT_MW.MEAS.MW", 3: "SUBSTN.SYSCA_TS.STTN.SOLAR_TOT.MES1.MW",4: "SUBSTN.SYSCA_AT.SYSTEM.GEN_SOLAR.MES.MW", 5:"SUBSTN.SYSCA_KL.SYSTEM.SOLR_TOTL.MEAS.MW", 6: "SUBSTN.SYSC2_PG.SYSTEM.SR_SOLAR_TOT.ADD.MW"},
    8: {1: "SUBSTN.SYSCA_KA.SYSTEM.WIND_TOTAL.MES1.MW", 2:"SUBSTN.SYSCA_TN.SYSTEM.TN_WIND_TOT.MEAS.MW", 3: "SUBSTN.SYSCA_TS.STTN.WIND_TOT.MEAS.MW",4: "SUBSTN.SYSCA_TS.STTN.WIND_TOT.MEAS.MW", 5:"SUBSTN.SYSCA_KL.SYSTEM.WIND_TTL.MEAS.MW", 6: "SUBSTN.SYSC2_PG.SYSTEM.SR_WIND_TOT.ADD.MW"},
    "DEMAND": {1: "SUBSTN.SYSC1_PG.SYSTEM.KA_DEMAND_GEN.ADD.MW", 2:"SUBSTN.SYSC1_PG.SYSTEM.TN_DEMAND_GEN.ADD.MW", 3: "SUBSTN.SYSC1_PG.SYSTEM.TS_DEMAND_GEN.ADD.MW",4: "SUBSTN.SYSC1_PG.SYSTEM.AP_DEMAND_GEN.ADD.MW", 5:"SUBSTN.SYSC1_PG.SYSTEM.KL_DEMAND_GEN.ADD.MW", 6: "SUBSTN.SYSC1_PG.SYSTEM.RG_DEMAND_TOT.ADD.MW"},
    5: {1: "SUBSTN.SYSC1_PG.SYSTEM.KA_HYDRO_GEN.ADD.MW", 2:"SUBSTN.SYSC1_PG.SYSTEM.TN_HYDRO_GEN.ADD.MW", 3: "SUBSTN.SYSC1_PG.SYSTEM.TS_HYDRO_GEN.ADD.MW",4: "SUBSTN.SYSC1_PG.SYSTEM.AP_HYDRO_GEN.ADD.MW", 5:"SUBSTN.SYSC1_PG.SYSTEM.KL_HYDRO_GEN.ADD.MW", 6: "SUBSTN.SYSC1_PG.SYSTEM.RG_HYDRO_TOT.ADD.MW"},
    4: {1: "SUBSTN.SYSC1_PG.SYSTEM.KA_GAS_DSL_GEN.ADD.MW", 2:"SUBSTN.SYSC1_PG.SYSTEM.TN_GAS_DSL_GEN.ADD.MW", 3: "SUBSTN.SYSC1_PG.SYSTEM.TL_GAS_DIES.MES1.MW",4: "SUBSTN.SYSC1_PG.SYSTEM.AP_GAS_DSL_GEN.ADD.MW", 5:"SUBSTN.SYSC1_PG.SYSTEM.KL_GAS_DSL_GEN.ADD.MW", 6: "SUBSTN.SYSC1_PG.SYSTEM.RG_GAS_DSL_TOT.ADD.MW"}
}


# app = Flask(__name__)
bcrypt = Bcrypt(app)

try:  
    conn = psycopg2.connect(
        database="demand_forecast_states", user='prasadbabu', 
    password='BabuPrasad#123', host='10.0.100.79', port='5432'
    )
    conn.autocommit = True
except Exception as e:
   pass

try:
    conn2 = psycopg2.connect(
        database="MDPNew", user = 'prasadbabu',
        password = 'BabuPrasad#123', host = '10.0.100.219', port = '5432'
    )
except Exception as e:
   pass

# shared_drive_path = r"E:\GridOpn\Excel_files"
shared_drive_path = r"D:\Demand Forecasting Page\custom scripts\Excel_files"

host_url = "http://192.168.57.50:5000/api"

# security details 
# username: security1
# password: security@srldc

#username: security2
# password: secure@srldc


# code to create security login 

# from flask import Flask, request

from flask_bcrypt import Bcrypt
# app = Flask(__name__)
bcrypt = Bcrypt(app)

# emp_pwd_hash=bcrypt.generate_password_hash(str('pgcilsr2@12345')).decode('utf-8')



# # print(emp_pwd_hash)

# cursor.execute("insert into security_details (username,enc_password, role) values ('{0}', '{1}', '{2}')".format('security2', emp_pwd_hash, 'security'))

# print("Security details inserted!")

# cursor = conn.cursor()

# state_details = []

# password_list = ["kar@12345", "tn#12345", "tg$12345", "ap%12345", "ker^12345", "admin&12345", "pondy*12345", 'pgcilsr1!12345', 'pgcilsr2@12345']

# password_hash_list = [bcrypt.generate_password_hash(i).decode('utf-8') for i in password_list]

# state_details.append({"state_id": 1, "state_name": "Karnataka","username": "kar_state", "password": password_hash_list[0], "acronym": "KAR", "role": "user" })
# state_details.append({"state_id": 2, "state_name": "Tamilnadu","username": "tn_state", "password": password_hash_list[1], "acronym": "TN", "role": "user" })
# state_details.append({"state_id": 3, "state_name": "Telangana","username": "tg_state", "password": password_hash_list[2], "acronym": "TLG", "role": "user" })
# state_details.append({"state_id": 4, "state_name": "Andhra Pradesh","username": "ap_state", "password": password_hash_list[3], "acronym": "AP", "role": "user" })
# state_details.append({"state_id": 5, "state_name": "Kerala","username": "ker_state", "password": password_hash_list[4], "acronym": "KER", "role": "user" })
# state_details.append({"state_id": 6, "state_name": "","username": "admin", "password": password_hash_list[5], "acronym": "admin", "role": "admin" })
# state_details.append({"state_id": 7, "state_name": "Pondicherry", "username": "pondy_state", "password": emp_pwd_hash, "acronym": "PDY", "role": "user"})
# state_details.append({"state_id": 8, "state_name": "PGCIL SR-1", "username": "pgcil_sr_1", "password": emp_pwd_hash, "acronym": "PGCIL SR-1", "role": "user"})
# state_details.append({"state_id": 9, "state_name": "PGCIL SR-2", "username": "pgcil_sr_2", "password": emp_pwd_hash, "acronym": "PGCIL SR-2", "role": "user"})


# For timing entry
# state_details = [
#     {"state_id": 11, "state_name": "Karnataka", "username": "t_kar_state", "password": "ka@#$12345", "acronym": "KA", "role": "user"},
#     {"state_id": 12, "state_name": "Tamilnadu", "username": "t_tn_state", "password": "tu#$%12345", "acronym": "TN", "role": "user"},
#     {"state_id": 13, "state_name": "Telangana", "username": "t_tg_state", "password": "ta$%^12345", "acronym": "TG", "role": "user"},
#     {"state_id": 14, "state_name": "Andhra Pradesh", "username": "t_ap_state", "password": "ah%^&12345", "acronym": "AP", "role": "user"},
#     {"state_id": 15, "state_name": "Kerala", "username": "t_ker_state", "password": "ka^&*12345", "acronym": "KL", "role": "user"},
#     {"state_id": 17, "state_name": "Pondicherry", "username": "t_pondy_state", "password": "py&*!12345", "acronym": "PD", "role": "user"},
        # {"state_id": 21, "state_name": "NLDC", "username": "nldc", "password": "nldc*12345", "acronym": "NLDC", "role": "user"}
# ]



# for i in state_details:
#     print(i)
#     cursor.execute("insert into states (state_id, state_name, username, password_hash, acronym, user_role) values ({0}, '{1}', '{2}', '{3}', '{4}', '{5}')".format(i["state_id"], i["state_name"], i["username"], i["password"], i["acronym"], i["role"]))

# conn.commit()

# print("Insertion successful!")



def calculate_mape(actual_values, forecast_values):
    if len(actual_values) != len(forecast_values):
        return -50  # Length mismatch error
    
    actual_values = np.array(actual_values)
    forecast_values = np.array(forecast_values)
    
    if np.any(actual_values == 0):
        return "Actual Values contain 0 due to which MAPE Calculation is not possible"
    
    # Vectorized calculation for MAPE
    mape = np.mean(np.abs((actual_values - forecast_values) / actual_values)) * 100
    return mape


def sendMail(send_mails_list_db, message):
  mail_id='srldcos'

  mail_pwd='OS@Oct#$2024'

  port = 587
  smtp_server="mail.grid-india.in"
  mail_server = smtplib.SMTP(smtp_server,port)
  mail_server.starttls()
  mail_server.login(mail_id,mail_pwd)
  mail_server.sendmail('srldcos@grid-india.in', (send_mails_list_db), message.as_string())


def generate_time_slots(start_date, end_date):
    # Convert string dates to datetime objects
    # print("Part 1")
    start = datetime.strptime(start_date, "%d/%m/%Y")
    end = datetime.strptime(end_date, "%d/%m/%Y")
    
    # Calculate the number of days in the range
    num_days = (end - start).days + 1
    
    time_slots = []
    time_format = "%H:%M"

    # print("Part2")
    
    # Generate time slots for each day
    for day in range(num_days):
        current_date = start + timedelta(days=day)
        date_str = current_date.strftime("%d/%m/%Y")
        
        # 96 time slots per day (15 minutes each)
        for slot in range(1, 97):
            start_time = datetime.combine(current_date, time(hour=(slot-1)*15//60, minute=(slot-1)*15%60))
            end_time = start_time + timedelta(minutes=15)
            
            # Append the date, slot number, and time range to the list
            time_slots.append([
                date_str,
                slot,
                f"{start_time.strftime(time_format)} - {end_time.strftime(time_format)}"
            ])
    
    # print("Part3")
    return time_slots

def generate_15min_time_labels():
    """
    Returns 96 time labels:
    ['00:00 - 00:15', '00:15 - 00:30', ..., '23:45 - 24:00']
    """
    today = datetime.today().strftime("%d/%m/%Y")

    slots = generate_time_slots(today, today)

    # Extract only the time-range column
    return [slot[2] for slot in slots]


def generate_15min_time_blocks():
    blocks = []
    for i in range(96):
        start_min = i * 15
        end_min = start_min + 15

        sh, sm = divmod(start_min, 60)
        eh, em = divmod(end_min, 60)

        blocks.append(f"{sh:02d}:{sm:02d} - {eh:02d}:{em:02d}")
    return blocks


################## DATA STATUS 

BASE_FORECAST_COLUMNS = [
    'Forecasted Demand',
    'THERMAL',
    'GAS',
    'HYDRO',
    'TOTAL',
    'SOLAR',
    'WIND',
    'Other RES (biomass)',
    'Renewable Sources_TOTAL',
    'ISGS & Other LTA & MTOA',
    'Bilateral Transaction (Advance+ FCFS)',
    'Total Availability',
    'Gap between Demand & Availability',
    'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency)',
    'Proposed Procurement_Through Power Exchange',
    'Shortages after day ahead procurement from market',
    'Relief through planned restrictions/ rostering/ power cuts',
    'Additional Load shedding proposed',
    'Reactive Power Forecast_MVar'
]

# Extra columns only for Day Ahead
DAY_AHEAD_EXTRA_COLUMNS = [
    'SRAS UP',
    'SRAS DOWN',
    'TRAS UP',
    'TRAS DOWN'
]


def validate_forecast_data(file_data, start_index, forecast_columns):
    """
    Validate whether each forecast column has exactly 96 non-null values for the day.
    :param file_data: JSON array (list of lists) from DB
    :param start_index: index where forecast columns start
    :param forecast_columns: list of column names for this dataType
    :return: dict {column_name: "tick"/"wrong"}
    """
    result = {}

    for i, col_name in enumerate(forecast_columns):
        col_index = start_index + i
        try:
            values = [
                float(row[col_index]) if row[col_index] not in [None, '', ' '] else None
                for row in file_data
            ]
        except Exception:
            values = []

        if len(values) == 96 and all(v is not None for v in values):
            result[col_name] = "tick"
        else:
            result[col_name] = "wrong"

    return result





############### Day ahead consolidated Mail sending


# Function to fetch data with the max revision_no for each state on a specific date
def fetch_max_revision_data(for_week_ahead, custom_date=None, custom_from_date=None, custom_to_date=None):
    if for_week_ahead == 'week':
        # print("entered max_revision_data", custom_from_date, custom_to_date)
        query = """
            SELECT t1.state_id, t1.file_data
            FROM public.week_ahead_file_uploads t1
            INNER JOIN (
                SELECT state_id, from_date, to_date, MAX(revision_no) as max_revision
                FROM public.week_ahead_file_uploads
                WHERE from_date = to_date(%s, 'DD/MM/YYYY') AND to_date = to_date(%s,'DD/MM/YYYY')
                GROUP BY state_id, from_date, to_date
            ) t2 ON t1.state_id = t2.state_id AND t1.from_date = t2.from_date AND t1.to_date = t2.to_date
            AND t1.revision_no = t2.max_revision;
        """
        conn = psycopg2.connect(**db_params)
        cur = conn.cursor()
        cur.execute(query, (custom_from_date, custom_to_date))
        data = cur.fetchall()
    elif for_week_ahead == 'month':
        query = """
            SELECT t1.state_id, t1.file_data
            FROM public.month_ahead_file_uploads t1
            INNER JOIN (
                SELECT state_id, from_date, to_date, MAX(revision_no) as max_revision
                FROM public.month_ahead_file_uploads
                WHERE from_date = to_date(%s, 'DD/MM/YYYY') AND to_date = to_date(%s,'DD/MM/YYYY')
                GROUP BY state_id, from_date, to_date
            ) t2 ON t1.state_id = t2.state_id AND t1.from_date = t2.from_date AND t1.to_date = t2.to_date
            AND t1.revision_no = t2.max_revision;
        """
        conn = psycopg2.connect(**db_params)
        cur = conn.cursor()
        cur.execute(query, (custom_from_date, custom_to_date))
        data = cur.fetchall()
    else:
        query = """
            SELECT t1.state_id, t1.file_data
            FROM public.file_uploads t1
            INNER JOIN (
                SELECT state_id, MAX(revision_no) as max_revision
                FROM public.file_uploads
                WHERE upload_date = %s
                GROUP BY state_id
            ) t2
            ON t1.state_id = t2.state_id AND t1.revision_no = t2.max_revision
            WHERE t1.upload_date = %s;
        """
        conn = psycopg2.connect(**db_params)
        cur = conn.cursor()
        cur.execute(query, (custom_date, custom_date))
        data = cur.fetchall()

    cur.close()
    conn.close()
    return data

# Function to replace None (null) values with 0 in a 2D list
def replace_null_with_zero(file_data):
    return [[0 if value is None else value for value in row] for row in file_data]

# Function to write data into Excel file and add footers
def write_to_excel(file_path, db_data, for_week_ahead, custom_date=None, custom_from_date=None, custom_to_date=None):
    # Load the workbook and initialize variables
    print(file_path)
    wb = load_workbook(file_path)

    
    sr_data = []  # To store consolidated data for SR
    
    for state_id, file_data in db_data:
        # Map state_id to the corresponding sheet name
        state_name = get_state_name(state_id)
        if state_name and state_name in wb.sheetnames:
            sheet = wb[state_name]

            # Set C2 with the state name and set C3 with the date or date range
            sheet['C2'] = state_name
            if for_week_ahead in ['week', 'month']:
                sheet['C3'] = f'{custom_from_date.strftime("%Y-%m-%d")} to {custom_to_date.strftime("%Y-%m-%d")}'
            else:
                sheet['C3'] = custom_date

            # Replace any None values with 0
            file_data = replace_null_with_zero(file_data)

            # Start writing the data from the 8th row
            start_row = 8
            for i, row_data in enumerate(file_data):
                for j, value in enumerate(row_data):
                    sheet.cell(row=i + start_row, column=j + 1).value = value

            # Append the state's data to the SR consolidated list
            sr_data.append(file_data)

            # Add footer with 'Total MUs' and sums
            add_footer(sheet, file_data, start_row, for_week_ahead)

    # Consolidate SR data (by summing or appending as per your requirement)
    sr_sheet = wb['SR']
    combined_sr_data = combine_state_data(sr_data, for_week_ahead)  # This function combines all state data
    for i, row_data in enumerate(combined_sr_data):
        for j, value in enumerate(row_data):
            sr_sheet.cell(row=i + start_row, column=j + 1).value = value

    # Add footer to SR sheet as well
    
    add_footer(sr_sheet, combined_sr_data, start_row, for_week_ahead)

    # Save the updated Excel file with a unique name based on the forecast type
    if for_week_ahead == 'week':
        save_file_path = file_path.replace("week_ahead_consolidated.xlsx", f"week_ahead_consolidated_for_{custom_from_date.strftime('%Y-%m-%d')}_to_{custom_to_date.strftime('%Y-%m-%d')}.xlsx")
    elif for_week_ahead == 'day':
        save_file_path = file_path.replace("day_ahead_consolidated.xlsx", f"day_ahead_consolidated_for_{custom_date}.xlsx")
    elif for_week_ahead == 'month':
        save_file_path = file_path.replace("month_ahead_consolidated.xlsx", f"month_ahead_consolidated_for_{custom_from_date.strftime('%Y-%m-%d')}_to_{custom_to_date.strftime('%Y-%m-%d')}.xlsx")

    wb.save(save_file_path)
    print(f"Excel file saved as {save_file_path}")
    
    return save_file_path  # Return the saved file path to attach in the email

# Function to combine data from all states for SR sheet
def combine_state_data(sr_data, for_week_ahead):
    combined_data = []
    
    for rows in zip(*sr_data):
        combined_row = []
        
        # For day ahead, keep the first two columns (block, period)
        # For week ahead, keep the first three columns (date, block, period)
        if for_week_ahead in ['week', 'month']:
            combined_row.append(rows[0][0])  # Date (first column)
            combined_row.append(rows[0][1])  # Block (second column)
            combined_row.append(rows[0][2])  # Period (third column)
            
            # Sum the remaining columns (starting from the fourth column onwards)
            for col in zip(*[row[3:] for row in rows]):
                try:
                    # Convert all values to float and sum them
                    combined_row.append(sum(float(value) for value in col))
                except ValueError:
                    # If conversion fails (e.g., non-numeric values), handle it gracefully by ignoring or treating as 0
                    combined_row.append(0)
        else:
            combined_row.append(rows[0][0])  # Block (first column)
            combined_row.append(rows[0][1])  # Period (second column)
            
            # Sum the remaining columns (starting from the third column onwards)
            for col in zip(*[row[2:] for row in rows]):
                try:
                    # Convert all values to float and sum them
                    combined_row.append(sum(float(value) for value in col))
                except ValueError:
                    # If conversion fails (e.g., non-numeric values), handle it gracefully by ignoring or treating as 0
                    combined_row.append(0)
        
        combined_data.append(combined_row)
    
    return combined_data


# Helper function to get the top-left cell in a merged range
def get_top_left_cell(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Return the top-left cell of the merged range
            return sheet.cell(merged_range.min_row, merged_range.min_col)
    # If not merged, return the cell itself
    return cell

# Function to safely write to cells, considering merged cells
def safe_write_to_cell(sheet, row, column, value, is_bold=False):
    try:
        cell = sheet.cell(row=row, column=column)
        print(f"Attempting to write to cell {cell.coordinate}")

        # Get the top-left cell of the merged range or the cell itself
        top_left_cell = get_top_left_cell(sheet, row, column)
        print(f"Writing to top-left cell: {top_left_cell.coordinate}")

        # Write value only to the top-left cell
        top_left_cell.value = value
        if is_bold:
            top_left_cell.font = Font(bold=True)

    except Exception as e:
        print(f"Error writing to cell {cell.coordinate}: {e}")

# Function to add footer to the Excel sheet
def add_footer(sheet, file_data, start_row, for_week_ahead):
    total_rows = len(file_data)
    last_row = start_row + total_rows - 1
    footer_row = last_row + 1  # Row directly below the last non-empty row

    # Handle based on file type: 'month', 'week', 'day'
    if for_week_ahead == 'month' or for_week_ahead == 'week':
        first_col = 3  # Handle week/month ahead, starting from 4th column
        safe_write_to_cell(sheet, footer_row, 1, "")  # First column empty
        safe_write_to_cell(sheet, footer_row, 2, "")  # Second column empty
        safe_write_to_cell(sheet, footer_row, 3, "Total MUs", is_bold=True)
    else:
        first_col = 2  # Handle day ahead, starting from 3rd column
        safe_write_to_cell(sheet, footer_row, 1, "")  # First column empty
        safe_write_to_cell(sheet, footer_row, 2, "Total MUs", is_bold=True)

    # Add sum for remaining columns divided by 4000, rounded to 2 decimals
    num_columns = len(file_data[0])  # Get the number of columns in the data
    for col in range(first_col + 1, num_columns + 1):
        sum_value = 0
        for row in range(start_row, last_row + 1):
            cell_value = sheet.cell(row=row, column=col).value
            # Check if the value is numeric (integer or float) before summing
            if isinstance(cell_value, (int, float)):
                sum_value += cell_value

        # Write the sum to the footer row
        safe_write_to_cell(sheet, footer_row, col, round(sum_value / 4000, 2))
# Function to map state_id to state names (sheet names)
def get_state_name(state_id):
    state_map = {
        1: 'KAR',  
        2: 'TN',   
        3: 'TG',   
        4: 'AP',   
        5: 'KER',  
        6: 'SR',   
        7: 'PONDY'  
    }
    return state_map.get(state_id, '')

def sendMail(send_mails_list_db, message):
  mail_id='srldcos'

  mail_pwd='OS@Sep#$2024'

  port = 587
  smtp_server="mail.grid-india.in"
  mail_server = smtplib.SMTP(smtp_server,port)
  mail_server.starttls()
  mail_server.login(mail_id,mail_pwd)
  mail_server.sendmail('srldcos@grid-india.in', (send_mails_list_db), message.as_string())

# Function to send an email with an attachment
def send_mail_with_attachment(send_mails_list_db, cc_mails_list_db, file_path, for_week_ahead, custom_date=None, custom_from_date=None, custom_to_date=None):
    subject = ''
    msg_body = ''
    if for_week_ahead == 'week':
        subject = f'Southern Region Week Ahead Demand forecast from {custom_from_date.strftime("%d-%m-%Y")} to {custom_to_date.strftime("%d-%m-%Y")}'
        msg_body = f"""
        महोदय/महोदया,<br>
        Sir/Madam,<br><br>
        Please find the Southern Region Week Ahead Demand forecast from {custom_from_date.strftime('%d-%m-%Y')} to {custom_to_date.strftime('%d-%m-%Y')} as received from the states.<br><br>
        धन्यवाद एवं आभार / Thanks and Regards,<br>
        --<br>
        SO-II / एस.ओ-II<br>
        SRLDC, Bengaluru / एस आर एल डी सी, बेंगलुरु<br>
        Grid Controller of India Limited / ग्रिड कंट्रोलर ऑफ इंडिया लिमिटेड<br>
        Erstwhile POWER SYSTEM OPERATION CORPORATION LIMITED / भूतपूर्व पावर सिस्टम ऑपरेशन कारपोरेशन लिमिटेड
        """
    elif for_week_ahead == 'day':
        subject = f'Southern Region Daily Demand forecast for {datetime.strptime(custom_date, "%Y-%m-%d").strftime("%d-%m-%Y")}'
        msg_body = f"""
        महोदय/महोदया,<br>
        Sir/Madam,<br><br>
        Please find the Southern Region Daily Demand forecast for {datetime.strptime(custom_date, '%Y-%m-%d').strftime('%d-%m-%Y')} as received from the states.<br><br>
        धन्यवाद एवं आभार / Thanks and Regards,<br>
        --<br>
        SO-II / एस.ओ-II<br>
        SRLDC, Bengaluru / एस आर एल डी सी, बेंगलुरु<br>
        Grid Controller of India Limited / ग्रिड कंट्रोलर ऑफ इंडिया लिमिटेड<br>
        Erstwhile POWER SYSTEM OPERATION CORPORATION LIMITED / भूतपूर्व पावर सिस्टम ऑपरेशन कारपोरेशन लिमिटेड
        """
    elif for_week_ahead == 'month':
        subject = f'Southern Region Month Ahead Demand forecast from {custom_from_date.strftime("%d-%m-%Y")} to {custom_to_date.strftime("%d-%m-%Y")}'
        msg_body = f"""
        महोदय/महोदया,<br>
        Sir/Madam,<br><br>
        Please find the Southern Region Month Ahead Demand forecast from {custom_from_date.strftime('%d-%m-%Y')} to {custom_to_date.strftime('%d-%m-%Y')} as received from the states.<br><br>
        धन्यवाद एवं आभार / Thanks and Regards,<br>
        --<br>
        SO-II / एस.ओ-II<br>
        SRLDC, Bengaluru / एस आर एल डी सी, बेंगलुरु<br>
        Grid Controller of India Limited / ग्रिड कंट्रोलर ऑफ इंडिया लिमिटेड<br>
        Erstwhile POWER SYSTEM OPERATION CORPORATION LIMITED / भूतपूर्व पावर सिस्टम ऑपरेशन कारपोरेशन लिमिटेड
        """

    message = MIMEMultipart()
    message["Subject"] = subject
    message["From"] = 'srldcos@grid-india.in'
    message["To"] = ', '.join(send_mails_list_db)
    message["Cc"] = ', '.join(cc_mails_list_db)
    
    message.attach(MIMEText(msg_body, 'html'))

    # Attach the Excel file
    with open(file_path, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
        message.attach(part)

    # Send the email
    sendMail(send_mails_list_db+cc_mails_list_db, message)




##### Week Ahead, Month ahead and Year ahead


def fmt_date(dt):
    if pd.isna(dt):
        return ""   # or return None
    return dt.strftime("%Y-%m-%d")

def safe_round(val):
    try:
        return round(float(val))
    except:
        return 0
