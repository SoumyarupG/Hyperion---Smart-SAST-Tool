from base64 import decode
from datetime import date, timedelta, datetime
from functools import wraps
import hashlib
from operator import itemgetter
import subprocess
import sys
import re, io
# import time
import traceback

from flask import Flask, Response, request, send_file
from http import HTTPStatus
from flask import jsonify
from flask_cors import CORS
from auth_decorators import role_required, rbac_required, permission_required
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ...existing code...


# import ldap
# import requests

from flask_jwt_extended import create_access_token, get_jwt
from flask_jwt_extended import create_refresh_token
from flask_jwt_extended import get_jwt_identity
from flask_jwt_extended import jwt_required
from flask_jwt_extended import JWTManager
from flask_jwt_extended import decode_token
# from grpc import Status
import openpyxl
import requests
from configBackend import *
# from configBackend import generate_time_slots
# from configBackend import generate_time_slots
# print("generate_time_slots type:", type(generate_time_slots))

import numpy as np
import pandas as pd
import json
import jwt
from flask_cors import CORS
import psycopg2
import calendar
from calendar import monthrange
from dateutil.relativedelta import relativedelta




import datetime
from datetime import timedelta, datetime
import warnings
warnings.filterwarnings('ignore')

# APIs

from flask import Flask, request, jsonify, redirect, make_response
from werkzeug.utils import secure_filename
import os

import secrets



# cursor = conn.cursor()


# cur = conn2.cursor()

# @app.before_request
# def force_https():
#     if not request.is_secure:
#         return redirect(request.url.replace('http://', 'https://'))

def fetch_thermal_outages_data(url):
    """
    Fetches thermal outages data from the given API URL.

    Args:
        url (str): The API URL.

    Returns:
        dict or None: The JSON response as a dictionary, or None if an error occurs.
    """
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise HTTPError for bad responses (4xx or 5xx)
        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON: {e}")
    return None





# def validate_captcha(captcha_token):
#     """Validate CAPTCHA token with the CAPTCHA provider."""
#     secret_key = app.config.get('RECAPTCHA_SECRET_KEY')  # Your CAPTCHA secret key
#     verify_url = "https://www.google.com/recaptcha/api/siteverify"

#     payload = {
#         "secret": secret_key,
#         "response": captcha_token
#     }

#     try:
#         response = requests.post(verify_url, data=payload)
#         result = response.json()
#         return result.get("success", False)  # CAPTCHA is valid if "success" is True
#     except Exception as e:
#         log_error("captcha_validation", e)
#         return False

# import requests

def validate_captcha(captcha_token, remote_ip=None):
    """Validate CAPTCHA token with Google reCAPTCHA siteverify API."""
    secret_key = app.config.get('RECAPTCHA_SECRET_KEY')  # Your secret key
    verify_url = "https://www.google.com/recaptcha/api/siteverify"

    payload = {
        "secret": secret_key,
        "response": captcha_token
    }
    if remote_ip:
        payload["remoteip"] = remote_ip

    try:
        response = requests.post(
            verify_url,
            data=payload,
            timeout=5  # prevent hanging (in seconds)
        )
        response.raise_for_status()  # raise for HTTP errors
        result = response.json()

        # print("CAPTCHA validation response:", result)  # Debugging line

        # success = True means valid captcha
        return result.get("success", False)

    except requests.exceptions.Timeout:
        log_error("captcha_validation", "Google reCAPTCHA request timed out")
        return False  # fail closed

    except Exception as e:
        log_error("captcha_validation", e)
        return False  # fail closed



# Rate limiting parameters
RATE_LIMIT = 5  # Max login attempts
RATE_LIMIT_WINDOW = timedelta(minutes=5)  # Time window for rate limiting

def is_rate_limited(ip_address, username, device_id):
    """Check if the IP address, username, and device ID are rate-limited."""
    current_time = datetime.utcnow()
    window_start = current_time - RATE_LIMIT_WINDOW

    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Count login attempts for the device ID within the time window
        cursor.execute(
            """
            SELECT COUNT(*) 
            FROM login_attempts 
            WHERE device_id = %s AND username = %s AND attempt_time > %s
            """,
            (device_id, username, window_start)
        )
        attempt_count = cursor.fetchone()[0]

        # Check if the rate limit is exceeded
        if attempt_count >= RATE_LIMIT:
            return True

        # Log the current login attempt
        cursor.execute(
            """
            INSERT INTO login_attempts (ip_address, username, device_id, attempt_time) 
            VALUES (%s, %s, %s, %s)
            """,
            (ip_address, username, device_id, current_time)
        )
        conn.commit()
    finally:
        cursor.close()
        conn.close()

    return False

# app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def log_error(api_name, error):
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    line_number = exc_tb.tb_lineno
    logger.error("Error in %s: %s at %s:%d", api_name, error, fname, line_number)


def subtract_months(source_date, months):
    month = source_date.month - 1 - months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, [31, 29 if year % 4 == 0 and not year % 400 == 0 else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month-1])
    return datetime(year, month, day)

# Function to generate the date range with each date repeating 96 times
def generate_date_range(from_date, to_date):
    num_days = (to_date - from_date).days + 1
    date_list = []
    for single_date in (from_date + timedelta(n) for n in range(num_days)):
        date_list.extend([single_date] * 96)  # Repeat each date 96 times
    return date_list


# Helper function to generate empty columns
def _column(df, col_name):
    df[col_name] = '--'
    return df

# Function to calculate column statistics
def calculate_column_stats(column):
    pos_sum = ( column[column > 0].sum() ) /1000
    neg_sum = ( column[column < 0].sum() ) /1000
    pos_max = ( column[column > 0].max() ) *4 if len(column[column > 0]) > 0 else 0
    neg_max = ( column[column < 0].min() )*4  if len(column[column < 0]) > 0 else 0

    return pd.Series([pos_sum, neg_sum, pos_max, neg_max], index=['Pos_Sum', 'Neg_Sum', 'Pos_Max', 'Neg_Max'])

def split_string(string):
    return re.findall(r'\(([A-Za-z0-9-]+)\)', string)

def generate_dataframe_forrange(start_date,end_date):
      data = []
      current_date = start_date
      # while current_date <= end_date:
      while current_date<=end_date:
            current_time = datetime.combine(current_date, datetime.min.time())
            end_time = datetime.combine(current_date, datetime.min.time()) + timedelta(days=1)
            while current_time < end_time:
                  data.append({
                  'DATE': current_date,
                  'TIME': current_time.time()
                  })
                  current_time += timedelta(minutes=15)
            
            current_date += timedelta(days=1)
      
      df = pd.DataFrame(data)
      return df


# Function to check start and end date for column
def check_start_enddate_col(full_data_df , start_date_time , end_date_time,col_name):
      try:
            full_data_df['DATETIME'] = pd.to_datetime(full_data_df['DATE'].astype(str) + ' ' + full_data_df['TIME'].astype(str))
                                               
            temp_endtimestamp=pd.to_datetime(end_date_time,errors='coerce')
            
            if pd.isnull(temp_endtimestamp):
                  default_endtimestamp = pd.to_datetime('2050-01-01')
            else: default_endtimestamp = temp_endtimestamp

            record_dt_time_condition=full_data_df['DATETIME'].apply(lambda dt: dt >= start_date_time and dt <= default_endtimestamp )
            
            full_data_df.loc[~(record_dt_time_condition), col_name] = '--'
            full_data_df.drop(columns=['DATETIME'],inplace=True)

            return full_data_df
      except Exception as e:
            return full_data_df


def calculate_formula(row,short_locations,formula):
    try:
        for loc in short_locations:
                formula=formula.replace(loc , str(row[loc]))
        return eval(formula)
    except Exception as e:
        # extractdb_errormsg(e)
        return '--'




from functools import wraps
from flask import request, jsonify
from datetime import datetime
import psycopg2

def session_token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        try:
            session_token = request.headers.get('X-Session-Token')
            if not session_token:
                return jsonify({
                    "status": "failure",
                    "message": "Session token is missing"
                }), 401

            device_id = request.headers.get('Device-ID', 'unknown_device')

            conn = psycopg2.connect(**db_params)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT 
                    us.expires_at,
                    us.device_id,
                    us.is_active,
                    s.username,
                    s.user_role,
                    s.state_id,
                    (
                        SELECT json_agg(pm.permission_key)
                        FROM role_permissions rp 
                        JOIN permissions_master pm ON pm.permission_id = rp.permission_id
                        JOIN roles r ON r.role_id = rp.role_id
                        WHERE LOWER(r.role_name) = LOWER(s.user_role)
                    ) AS permissions
                FROM user_sessions us
                JOIN states s ON s.username = us.username
                WHERE us.session_token = %s
            """, (session_token,))

            row = cursor.fetchone()
            cursor.close()
            conn.close()

            if not row:
                return jsonify({
                    "status": "failure",
                    "message": "Invalid session token"
                }), 401

            expires_at, stored_device, is_active, username, role, state_id, permissions = row

            role = (role or "").lower().strip()
            permissions = permissions or []

            # -----------------------------
            #  Normalized role mapping
            # -----------------------------
            if role.startswith("t_") and role.endswith("_state"):
                normalized_role = "t_user"
            elif role.startswith("f_") and role.endswith("_state"):
                normalized_role = "f_user"
            else:
                normalized_role = role  # admin / nldc unchanged

            # -----------------------------
            # Session validations
            # -----------------------------
            if not is_active:
                return jsonify({"status": "failure", "message": "Session invalidated"}), 401

            if datetime.utcnow() > expires_at:
                return jsonify({"status": "failure", "message": "Session expired"}), 401

            if stored_device != device_id:
                return jsonify({"status": "failure", "message": "Device mismatch"}), 401

            # -----------------------------
            # Final injected user object
            # -----------------------------
            request.user = {
                "username": username,
                "role": normalized_role,
                "state_id": state_id,
                "permissions": permissions
            }

            # print("Validated user:", request.user)

            return f(*args, **kwargs)

        except Exception as e:
            log_error("session_token_validation", e)
            return jsonify({
                "status": "failure",
                "message": "Session validation failed"
            }), 500

    return decorated

# Custom token_required decorator for script access
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        jwt_data = get_jwt()

        # Special script-only access
        if jwt_data.get("script_access"):

            # Only allow these endpoints
            if request.endpoint not in ["get_data", "get_forecast_data"]:
                return jsonify({
                    "status": "failure",
                    "message": "Script token not authorized for this endpoint"
                }), 403

        return f(*args, **kwargs)

    return decorated

@app.route("/api/me", methods=["GET"])
@jwt_required(locations=["cookies"])
@session_token_required
def me():
    try:
        u = request.user   # normalized role: admin, nldc, f_user, t_user

        return jsonify({
            "status": "success",

            # Angular expects both
            "user": u["username"],
            "username": u["username"],

            # critical: Angular stores this in storage for RBAC
            "role": u["role"],           
            "state_id": u["state_id"],

            # Always include permissions (even if empty)
            "permissions": u.get("permissions", []),

            # optional but consistent
            "state_name": ""
        })

    except Exception as e:
        log_error("me_api", e)
        return jsonify(status="failure", message="Unable to fetch user details"), 500


@app.route("/api/logout", methods=["POST"])
def logout():
    response = make_response(jsonify({"message": "Logged out successfully"}))
    response.set_cookie("access_token", "", httponly=True, secure=True, samesite="None", expires=0)
    return response

# @app.route("/api/login", methods=["POST"])
# def login():
#     try:
#         ip_address = request.remote_addr
#         device_id = request.headers.get('Device-ID', 'unknown_device')

#         data = request.get_json()
#         username = data.get('username')
#         password = data.get('password')
#         captcha_token = data.get('recaptcha')

#         # ---------------- CAPTCHA VALIDATION ----------------
#         if not validate_captcha(captcha_token, remote_ip=ip_address):
#             return jsonify({"error": "Invalid CAPTCHA.", "status": "failure"})

#         # ---------------- RATE LIMIT CHECK ----------------
#         if is_rate_limited(ip_address, username, device_id):
#             return jsonify({
#                 "error": "Too many login attempts. Try again later.",
#                 "retry_after": RATE_LIMIT_WINDOW.seconds // 60,
#                 "status": "failure"
#             })

#         if not username or not password:
#             return jsonify({"error": "Username & password required", "status": "failure"})

#         # ---------------- DATABASE CONNECTION ----------------
#         conn = psycopg2.connect(**db_params)
#         cursor = conn.cursor()

#         # ---------------- FETCH USER ----------------
#         cursor.execute("""
#             SELECT username, password_hash, user_role, state_name, state_id 
#             FROM states 
#             WHERE username = %s
#         """, (username,))
        
#         login_data = cursor.fetchone()
#         if not login_data:
#             return jsonify({"error": "Invalid credentials", "status": "failure"})

#         db_username, enc_password, role_name, state_name, state_id = login_data

#         # ---------------- PASSWORD CHECK ----------------
#         if not bcrypt.check_password_hash(enc_password, password):
#             return jsonify({"error": "Invalid credentials", "status": "failure"})

#         # ---------------- FETCH ROLE ID ----------------
#         cursor.execute("""
#             SELECT role_id 
#             FROM roles 
#             WHERE role_name = %s
#         """, (role_name,))
#         role_row = cursor.fetchone()
#         if not role_row:
#             return jsonify({"error": "Role not mapped!", "status": "failure"})
        
#         role_id = role_row[0]

#         # ---------------- FETCH PERMISSIONS BY ROLE ----------------
#         cursor.execute("""
#             SELECT pm.permission_key
#             FROM role_permissions rp
#             JOIN permissions_master pm
#               ON rp.permission_id = pm.permission_id
#             WHERE rp.role_id = %s
#         """, (role_id,))

#         permissions_rows = cursor.fetchall()
#         permissions_list = [row[0] for row in permissions_rows]

#         # ---------------- CREATE SESSION TOKEN ----------------
#         session_token = secrets.token_urlsafe(32)
#         session_expiration = datetime.utcnow() + timedelta(minutes=60)

#         cursor.execute("""
#             INSERT INTO user_sessions 
#             (username, session_token, ip_address, device_id, user_agent, created_at, expires_at)
#             VALUES (%s, %s, %s, %s, %s, NOW(), %s)
#         """, (
#             username,
#             session_token,
#             ip_address,
#             device_id,
#             request.headers.get('User-Agent'),
#             session_expiration
#         ))
#         conn.commit()

#         # ---------------- CREATE JWT ----------------
#         jwt_data = {
#             "sub": username,
#             "role": role_name,
#             "exp": datetime.utcnow() + app.config.get('JWT_ACCESS_TOKEN_EXPIRES')
#         }

#         access_token = jwt.encode(
#             payload=jwt_data,
#             key=app.config.get('JWT_SECRET_KEY'),
#             algorithm=app.config.get('ALGORITHM')
#         )

#         # ---------------- RESPONSE ----------------
#         response = make_response(jsonify({
#             "session_token": session_token,
#             "device_id": device_id,
#             "user": username,
#             "username": state_name,
#             "role": role_name,
#             "state_id": state_id,
#             "permissions": permissions_list,   # <── FINAL RBAC FIX
#             "expires_at": session_expiration.strftime("%Y-%m-%d %H:%M:%S"),
#             "status": "success"
#         }))

#         # Set cookie for JWT
#         response.set_cookie(
#             "access_token",
#             access_token,
#             httponly=True,
#             secure=False,
#             samesite="Lax",
#             path="/"
#         )

#         return response

#     except Exception as e:
#         log_error("login", e)
#         return jsonify({"error": "Login failed!", "status": "failure"})

# Backend: simplified login (Flask)
@app.route("/api/login", methods=["POST"])
def login():
    try:
        ip_address = request.remote_addr
        device_id = request.headers.get('Device-ID', 'unknown_device')

        data = request.get_json() or {}
        username = data.get('username')
        password = data.get('password')
        captcha_token = data.get('recaptcha')

        # CAPTCHA
        if not validate_captcha(captcha_token, remote_ip=ip_address):
            return jsonify({"status": "failure", "error": "Invalid CAPTCHA"})

        # Rate limit
        if is_rate_limited(ip_address, username, device_id):
            return jsonify({"status": "failure", "error": "Too many login attempts", "retry_after": RATE_LIMIT_WINDOW.seconds // 60})

        if not username or not password:
            return jsonify({"status": "failure", "error": "Username & password required"})

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # FETCH USER using user_role
        cursor.execute("""
            SELECT username, password_hash, user_role, state_name, state_id
            FROM states
            WHERE username = %s
        """, (username,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"status": "failure", "error": "Invalid credentials"})

        db_username, db_pass_hash, role_name, state_name, state_id = row

        # password check
        if not bcrypt.check_password_hash(db_pass_hash, password):
            return jsonify({"status": "failure", "error": "Invalid credentials"})

        # normalize role and verify in roles table
        role_name = (role_name or "").lower().strip()

        cursor.execute("""
            SELECT role_id
            FROM roles
            WHERE LOWER(role_name) = %s
        """, (role_name,))
        role_row = cursor.fetchone()
        if not role_row:
            return jsonify({"status": "failure", "error": "Role not mapped!"})
        role_id = role_row[0]

        # fetch permissions for role (lowercased)
        cursor.execute("""
            SELECT pm.permission_key
            FROM role_permissions rp
            JOIN permissions_master pm ON rp.permission_id = pm.permission_id
            WHERE rp.role_id = %s
        """, (role_id,))
        permissions_rows = cursor.fetchall()
        permissions_list = [r[0].lower() for r in permissions_rows]

        # session token + store
        session_token = secrets.token_urlsafe(32)
        expiry = datetime.utcnow() + timedelta(minutes=60)
        cursor.execute("""
            INSERT INTO user_sessions (username, session_token, ip_address, device_id, user_agent, created_at, expires_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), %s)
        """, (username, session_token, ip_address, device_id, request.headers.get('User-Agent'), expiry))
        conn.commit()

        # jwt (optional)
        jwt_data = {"sub": username, "role": role_name, "exp": datetime.utcnow() + app.config.get('JWT_ACCESS_TOKEN_EXPIRES')}
        access_token = jwt.encode(jwt_data, app.config.get('JWT_SECRET_KEY'), algorithm=app.config.get('ALGORITHM'))

        resp = make_response(jsonify({
            "status": "success",
            "user": username,
            "username": state_name or username,
            "role": role_name,
            "state_id": state_id,
            "state_name": state_name,
            "permissions": permissions_list,
            "session_token": session_token,
            "device_id": device_id,
            "expires_at": expiry.strftime("%Y-%m-%d %H:%M:%S")
        }))
        resp.set_cookie("access_token", access_token, httponly=True, secure=False, samesite="Lax", path="/")
        return resp

    except Exception as e:
        log_error("login", e)
        return jsonify({"status": "failure", "error": "Login failed!"})


@app.route('/api/fetchrevisions', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_day_revisions")
def fetchRevisions():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        date = params["date"]
        params["date"] = date
        state = params["state"]
        cursor.execute("select revision_no from file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY')".format(params["state"], params["date"]))
        revisions_data = cursor.fetchall()

        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for {0}".format(date), revisions=revision_list)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for {0} state for date {1}".format(state_name, date))
    except Exception as e:
        log_error("fetchrevisions", e)
        return jsonify(message="There is Some Problem, Please contact SRLDC IT")


@app.route('/api/fetchintradayrevisions', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_intraday_revisions", enforce_state=True)
def fetchIntradayRevisions():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        date = params["date"]
        params["date"] = date
        state = params["state"]
        cursor.execute("select revision_no, file_type from intraday_file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY')".format(params["state"], params["date"]))
        revisions_data = cursor.fetchall()

        revision_list = [(i[0],i[1]) for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]

        # 3) Fetch state_name for messaging
        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        state_name = cursor.fetchone()[0]
        cursor.close()

        # Prepare a list of dicts -> [{ "revision": ..., "fetch_from": ... }, ...]
        final_revisions = []

        if revision_list:
            # For example, if your first intraday revision is 0, it means intraday(0).
            final_revisions = [
                {
                    "revision": rev[0],
                    "fetch_from": rev[1]
                } 
                for rev in revision_list
            ]
            
        else:
            # No day-ahead and no intraday
            return jsonify(
                status="failure",
                message=f"No day-ahead data or intraday revisions found for {state_name} on {date}"
            )
    
        return jsonify(
            status="success",
            message=f"Fetched Revisions Successfully for {state_name} on {date}",
            revisions=final_revisions
        )

    except Exception as e:
        log_error("fetchintradayrevisions", e)
        return jsonify(message="There is Some Problem, Please contact SRLDC IT")



@app.route('/api/fetchweekrevisions', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_week_revisions")
def fetchWeekRevisions():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        from_date = params["from_date"]
        to_date = params["to_date"]
        state = params["state"]
        cursor.execute("select revision_no from week_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY')".format(params["state"], from_date, to_date))
        revisions_data = cursor.fetchall()
        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for '{0}'-'{1}'".format(from_date, to_date), revisions=revision_list, from_date=from_date, to_date=to_date)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for state {0} for week '{1}'-'{2}'".format(state_name, from_date, to_date))    
    
    except Exception as e:
        log_error("fetchweekrevisions", e)
        return jsonify(message="There is Some Problem, Please contact SRLDC IT")

@app.route('/api/fetchweeklyrevisionsdata', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_week_revisions_data")
def fetchWeeklyRevisionsData():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        print(params)
        cursor.execute("select file_data, from_date, to_date, upload_time, uploaded_by, revision_no from week_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY') and revision_no={3}".format(params["state"], params["from_date"], params["to_date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        from_date = ''
        to_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            from_date = data[0][1].strftime('%d-%b-%Y')
            to_date = data[0][2].strftime('%d-%b-%Y')
            uploaded_time = data[0][3].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][4]
            revision_no = data[0][5]

            # print(len(file_data), len(file_data[0]))
            cursor.close()
            return jsonify(status="success", data=file_data, time=uploaded_time, from_date=from_date,to_date=to_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact SRLDC IT!")
            

        # return jsonify(message="Fetched Successfully")

    except Exception as e:
        log_error("weeklyrevisionsdata", e)
        return jsonify(message="There is Some Problem, Please contact SRLDC IT")
    

@app.route('/api/fetchintradayrevisionsdata', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_intraday_revisions_data")
def checkIntradayUploaded():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        params = request.get_json()
        # print("Received params:", params)

        # Extract state, date, and revision info
        state = params["state"]
        date_str = params["date"]    # e.g. "18/12/2024"
        revision_info = params["revision"]
        
        # revision_info is of the form: { "fetch_from": "intraday" | "day_ahead", "revision": <number> }
        fetch_from = revision_info["fetch_from"]  # "intraday" or "day_ahead"
        rev_no = int(revision_info["revision"])

        # Decide which table to query based on fetch_from
        # if fetch_from == "day_ahead":
        #     table_name = "file_uploads"
        # else:
        #     table_name = "intraday_file_uploads"

        table_name = 'intraday_file_uploads'

        # Build the query string
        query = f"""
            SELECT file_data, upload_date, upload_time, uploaded_by, revision_no
            FROM {table_name}
            WHERE state_id = %s
              AND upload_date = TO_DATE(%s, 'DD/MM/YYYY')
              AND revision_no = %s
        """

        # Execute the query
        cursor.execute(query, (state, date_str, rev_no))
        rows = cursor.fetchall()

        if not rows:
            # No data found
            cursor.close()
            return jsonify(
                status="failure",
                message="No data found for the requested state, date, and revision. Please contact SRLDC IT!"
            )

        # Expect only one row for a given (state, date, revision). If multiple, take the first.
        file_data, upload_date_db, upload_time_db, uploaded_by, db_revision_no = rows[0]

        # Format date/time
        uploaded_date_str = upload_date_db.strftime('%d-%b-%Y') if upload_date_db else ''
        # Use 12-hour format with AM/PM => '%I:%M:%S %p'
        uploaded_time_str = upload_time_db.strftime('%d-%b-%Y %I:%M:%S %p') if upload_time_db else ''

        cursor.close()

        # Return success with the data
        return jsonify(
            status="success",
            data=file_data,
            time=uploaded_time_str,
            date=uploaded_date_str,
            revision=db_revision_no,
            role=uploaded_by
        )

    except Exception as e:
        log_error("fetchintradayrevisionsdata", e)
        return jsonify(
            status="failure",
            message="There is a Problem in fetching the data. Please contact SRLDC IT!"
        )

# @app.route('/api/fetchintradayrevisionsdata', methods=['POST'])
# @jwt_required(locations=["cookies"])
# @token_required
# def checkIntradayUploaded():
#     try:
#         jwt_data = get_jwt()
#         user_role = jwt_data.get("role")
#         username = jwt_data.get("sub")   # "sub" contains username

#         # ---- ROLE-BASED ACCESS CONTROL ----
#         # Only admin or user roles are allowed
#         if user_role not in ["admin", "user"]:
#             return jsonify(status="failure", message="Unauthorized role"), 403

#         params = request.get_json()
#         state = params["state"]
#         date_str = params["date"]
#         revision_info = params["revision"]
#         fetch_from = revision_info["fetch_from"]
#         rev_no = int(revision_info["revision"])

#         conn = psycopg2.connect(**db_params)
#         cursor = conn.cursor()

#         # If role is "user", restrict to their own state
#         if user_role == "user":
#             cursor.execute("SELECT state_id FROM states WHERE username = %s", (username,))
#             allowed_state_id = cursor.fetchone()[0]
#             if str(state) != str(allowed_state_id):
#                 cursor.close()
#                 return jsonify(status="failure", message="Access Denied: Not your state"), 403

#         query = """
#             SELECT file_data, upload_date, upload_time, uploaded_by, revision_no
#             FROM intraday_file_uploads
#             WHERE state_id = %s
#               AND upload_date = TO_DATE(%s, 'DD/MM/YYYY')
#               AND revision_no = %s
#         """
#         cursor.execute(query, (state, date_str, rev_no))
#         rows = cursor.fetchall()
#         cursor.close()

#         if not rows:
#             return jsonify(status="failure", message="No data found")

#         file_data, upload_date_db, upload_time_db, uploaded_by, db_revision_no = rows[0]

#         return jsonify(
#             status="success",
#             data=file_data,
#             time=upload_time_db.strftime('%d-%b-%Y %I:%M:%S %p'),
#             date=upload_date_db.strftime('%d-%b-%Y'),
#             revision=db_revision_no,
#             role=uploaded_by
#         )

#     except Exception as e:
#         log_error("fetchintradayrevisionsdata", e)
#         return jsonify(status="failure", message="Server Error"), 500



@app.route('/api/fetchrevisionsdata', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_day_revisions_data")
def checkUploaded():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        print(params)
        cursor.execute("select file_data, upload_date, upload_time, uploaded_by, revision_no from file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY') and revision_no={2}".format(params["state"], params["date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        uploaded_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            uploaded_date = data[0][1].strftime('%d-%b-%Y')
            uploaded_time = data[0][2].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][3]
            revision_no = data[0][4]
            # print(type(uploaded_date), type(uploaded_time))
            # print(type(file_data), uploaded_time, uploaded_date, uploaded_by, revision_no)
            # print(file_data[-1])
            # print(len(file_data), len(file_data[0]))
            cursor.close()
            return jsonify(status="success", data=file_data, time=uploaded_time, date=uploaded_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact SRLDC IT!")
    except Exception as e:
        log_error("fetchrevisionsdata", e)
        
        

    # return jsonify(message="Fetched Successfully")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


state_revision_numbers = {}








@app.route('/api/parseintradayexcel', methods=['POST'])
@jwt_required(locations=["cookies"])
def parse_intraday_excel():
    try:
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})
        
        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        # Use Pandas to read the file
        # We DO NOT use fillna(0) here because we need to detect blanks specifically for Forecasted Demand
        df = pd.read_excel(file, header=None) 
        
        data_list = df.values.tolist()
        
        cleaned_data = []
        blanks_replaced_flag = False

        for row in data_list:
            # Check if first column is a number between 1 and 96
            try:
                # Skip if Block Number is missing/NaN
                if pd.isna(row[0]):
                    continue

                block_num = int(row[0])
                
                if 1 <= block_num <= 96:
                    # We will construct a new 'clean' row to append
                    # Indices: 0=Block, 1=Time, 2=Forecasted Demand, 3...=Others
                    original_slice = row[:21]
                    new_row = [block_num, original_slice[1]] 

                    # --- 1. VALIDATE FORECASTED DEMAND (Index 2) ---
                    demand_val = original_slice[2]
                    
                    # Check for Blank/NaN
                    if pd.isna(demand_val) or str(demand_val).strip() == '':
                        return jsonify({'error': f'Forecasted Demand at Block {block_num} cannot be blank.'})
                    
                    # Check for Zero
                    try:
                        demand_float = float(demand_val)
                        if demand_float == 0:
                            return jsonify({'error': f'Forecasted Demand at Block {block_num} cannot be 0.'})
                        new_row.append(demand_float)
                    except ValueError:
                        return jsonify({'error': f'Invalid numeric value for Forecasted Demand at Block {block_num}.'})

                    # --- 2. HANDLE OTHER COLUMNS (Index 3 to 20) ---
                    # These can be 0. If blank, replace with 0 and warn.
                    for i in range(3, 21):
                        val = original_slice[i]
                        
                        if pd.isna(val) or str(val).strip() == '':
                            new_row.append(0.0)
                            blanks_replaced_flag = True
                        else:
                            try:
                                new_row.append(float(val))
                            except ValueError:
                                return jsonify({'error': f'Invalid numeric data in Column {i+1} at Block {block_num}.'})
                    
                    cleaned_data.append(new_row)

            except (ValueError, TypeError):
                continue

        if len(cleaned_data) != 96:
             return jsonify({'error': f'Invalid Row Count. Expected 96 blocks, found {len(cleaned_data)}'})

        # --- 3. SUMMATION VALIDATION (Thermal + Gas + Hydro == Total Own Sources) ---
        # Based on structure:
        # Col 3 (Index 3): Thermal
        # Col 4 (Index 4): Gas
        # Col 5 (Index 5): Hydro
        # Col 6 (Index 6): Total (B)
        for row in cleaned_data:
            thermal = row[3]
            gas = row[4]
            hydro = row[5]
            total_own = row[6]
            
            # Check sum with a small tolerance for floating point arithmetic
            if abs((thermal + gas + hydro) - total_own) > 0.1:
                 return jsonify({'error': f'Summation Mismatch at Block {row[0]} ({row[1]}). Thermal ({thermal}) + Gas ({gas}) + Hydro ({hydro}) = {thermal+gas+hydro:.2f}, but Total is {total_own}'})

        # Construct Response Message
        success_msg = 'Parsed successfully, You can Upload the Data now!'
        if blanks_replaced_flag:
            success_msg = 'Parsed successfully. Warning: Some blank fields in generation columns were replaced with 0. Please review before uploading.'

        return jsonify({'message': success_msg, 'data': cleaned_data})

    except Exception as e:
        log_error("parse_intraday", e)
        return jsonify({'error': str(e)})



# API 2: SAVE DATA (Write to Format File)
@app.route('/api/uploadintradaydata', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("upload_intraday", enforce_state=True)
def upload_intraday_data():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(
            database="demand_forecast_states", user='prasadbabu',
            password='BabuPrasad#123', host='10.0.100.79', port='5432'
        )
        cursor = conn.cursor()

        # Get Form Data
        state = request.form.get('state')
        disabledDateRaw = request.form.get('disabledDate')
        data_json = request.form.get('data')
        
        if not data_json:
            return jsonify({'error': 'No data provided'})
            
        new_data = json.loads(data_json) # List of lists (96 rows)
        role = jwt_data['role']

        # Get State Name (Optional validation)
        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        res_state = cursor.fetchone()
        if not res_state:
             return jsonify({'error': 'Invalid State ID'})
        # state_name = res_state[0] 

        # --- ROBUST DATE PARSING ---
        formatted_date = None
        
        # List of potential date formats to try
        date_formats = [
            "%d %b, %Y",       # 05 Jan, 2026
            "%Y-%m-%d",        # 2026-01-05
            "%a %b %d %Y",     # Sun Jan 05 2026 (Common JS prefix)
            "%a %b %d %Y %H:%M:%S" # Full JS string start
        ]

        # Clean the input string (remove timezone info often found in JS strings like ' GMT+0530...')
        # This simplifies parsing significantly.
        clean_date_str = str(disabledDateRaw).split(' GMT')[0].strip()

        for fmt in date_formats:
            try:
                # Try parsing with the current format
                # If clean_date_str is longer than the format expects (e.g. has extra spaces), 
                # strptime usually handles it if the beginning matches, but for exactness:
                if len(clean_date_str) > 15 and fmt == "%a %b %d %Y":
                     # Take just the first 15 chars for "Sun Jan 05 2026"
                     formatted_date = datetime.strptime(clean_date_str[:15], fmt).strftime("%Y-%m-%d")
                else:
                     formatted_date = datetime.strptime(clean_date_str, fmt).strftime("%Y-%m-%d")
                break # If successful, stop trying
            except ValueError:
                continue
        
        # Fallback: If all parses fail, use the raw value (will likely error in SQL if invalid)
        if not formatted_date:
            formatted_date = disabledDateRaw

        # --- REVISION LOGIC (Keep Past Data) ---
        current_timestamp = datetime.now()
        current_time_in_minutes = current_timestamp.hour * 60 + current_timestamp.minute
        current_block_no = (current_time_in_minutes // 15) + 1

        # Fetch most recent revision to preserve past blocks
        cursor.execute("""
            SELECT revision_no, file_data 
            FROM intraday_file_uploads 
            WHERE upload_date = to_date(%s, 'YYYY-MM-DD') 
            AND state_id = %s 
            ORDER BY revision_no DESC LIMIT 1
        """, (formatted_date, state))
        
        recent_revision = cursor.fetchone()
        
        final_data = new_data # Default to new incoming data
        revision_no = 1

        if recent_revision:
            previous_revision_no = recent_revision[0]
            previous_data = recent_revision[1] # This is a JSON object (list of lists)
            revision_no = previous_revision_no + 1
            
            # Logic: If current time is past a block, we MUST keep the old value for that block
            # We cannot change history.
            for i, block in enumerate(previous_data):
                # block[0] is the block number
                if i < len(final_data) and int(block[0]) <= current_block_no:
                    final_data[i] = block # Overwrite new input with old retained data

        # --- DB INSERT (Saving Data Only) ---
        virtual_filename = f"DB_STORED_{formatted_date}_rev{revision_no}"

        cursor.execute("""
            INSERT INTO intraday_file_uploads 
            (state_id, upload_date, upload_time, file_name, revision_no, uploaded_by, file_data, file_type) 
            VALUES (%s, to_date(%s, 'YYYY-MM-DD'), %s, %s, %s, %s, %s, %s)
        """, (state, formatted_date, current_timestamp.strftime('%Y-%m-%d %H:%M:%S'), virtual_filename, revision_no, role, json.dumps(final_data), 'I'))
        
        conn.commit()

        return jsonify({'message': f'Data saved successfully to database. Revision-{revision_no} created.'})

    except Exception as e:
        if conn: conn.rollback()
        # log_error("upload_intraday_data", e) 
        print(f"Error in upload_intraday_data: {str(e)}")
        return jsonify({'error': str(e)})

    finally:
        if cursor: cursor.close()
        if conn: conn.close()




@app.route('/api/parsedayaheadexcel', methods=['POST'])
@jwt_required(locations=["cookies"])
def parse_dayahead_excel():
    try:
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})
        
        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        # Use Pandas to read the file
        df = pd.read_excel(file, header=None) 
        
        data_list = df.values.tolist()
        
        cleaned_data = []
        blanks_replaced_flag = False

        for row in data_list:
            # Check if first column is a number between 1 and 96
            try:
                if pd.isna(row[0]): continue
                block_num = int(row[0])
                
                if 1 <= block_num <= 96:
                    # Slice to first 25 columns (matches your frontend Jspreadsheet)
                    original_slice = row[:25] 
                    new_row = [block_num, original_slice[1]] 

                    # --- 1. VALIDATE FORECASTED DEMAND (Index 2) ---
                    demand_val = original_slice[2]
                    
                    if pd.isna(demand_val) or str(demand_val).strip() == '':
                        return jsonify({'error': f'Forecasted Demand at Block {block_num} cannot be blank.'})
                    
                    try:
                        demand_float = float(demand_val)
                        if demand_float == 0:
                            return jsonify({'error': f'Forecasted Demand at Block {block_num} cannot be 0.'})
                        new_row.append(demand_float)
                    except ValueError:
                        return jsonify({'error': f'Invalid numeric value for Demand at Block {block_num}.'})

                    # --- 2. HANDLE OTHER COLUMNS (Index 3 to 24) ---
                    for i in range(3, 25):
                        val = original_slice[i]
                        if pd.isna(val) or str(val).strip() == '':
                            new_row.append(0.0)
                            blanks_replaced_flag = True
                        else:
                            try:
                                new_row.append(float(val))
                            except ValueError:
                                return jsonify({'error': f'Invalid numeric data in Column {i+1} at Block {block_num}.'})
                    
                    cleaned_data.append(new_row)

            except (ValueError, TypeError):
                continue

        if len(cleaned_data) != 96:
             return jsonify({'error': f'Invalid Row Count. Expected 96 blocks, found {len(cleaned_data)}'})

        # --- 3. SUMMATION VALIDATION (Thermal + Gas + Hydro == Total Own Sources) ---
        # Col 3: Thermal, Col 4: Gas, Col 5: Hydro, Col 6: Total (B)
        for row in cleaned_data:
            thermal, gas, hydro, total_own = row[3], row[4], row[5], row[6]
            if abs((thermal + gas + hydro) - total_own) > 0.1:
                 return jsonify({'error': f'Summation Mismatch at Block {row[0]}. Thermal+Gas+Hydro ({thermal+gas+hydro:.2f}) != Total ({total_own})'})

        success_msg = 'Parsed successfully. You can Upload the Data now!'
        if blanks_replaced_flag:
            success_msg = 'Parsed successfully. Warning: Blank generation fields were replaced with 0.'

        return jsonify({'message': success_msg, 'data': cleaned_data})

    except Exception as e:
        # log_error("parse_dayahead", e)
        return jsonify({'error': str(e)})


@app.route('/api/uploaddayahead', methods=['POST']) # Kept route name same as service
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("upload_dayahead", enforce_state=True)
def upload_dayahead_data():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Get Form Data
        state = request.form.get('state')
        disabledDateRaw = request.form.get('disabledDate')
        data_json = request.form.get('data')
        
        if not data_json: return jsonify({'error': 'No data provided'})
        new_data = json.loads(data_json)
        role = jwt_data['role']

        # Date Parsing
        formatted_date = disabledDateRaw
        try:
             formatted_date = datetime.strptime(str(disabledDateRaw).split(' GMT')[0].strip(), "%a %b %d %Y %H:%M:%S").strftime("%Y-%m-%d")
        except:
             try: formatted_date = datetime.strptime(disabledDateRaw, "%d %b, %Y").strftime("%Y-%m-%d")
             except: pass

        # Get State Name
        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        state_name = cursor.fetchone()[0]

        # Calculate Revision
        cursor.execute("SELECT COUNT(*) FROM file_uploads WHERE upload_date = %s AND state_id = %s", (formatted_date, state))
        revision_no = cursor.fetchone()[0] # 0 for first upload, 1 for second, etc.
        # Note: If you want 1-based revision for filenames, use revision_no + 1

        virtual_filename = f"DB_STORED_{formatted_date}_{state_name}_rev{revision_no}"
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 1. Insert into Day Ahead Table
        cursor.execute("""
            INSERT INTO file_uploads (state_id, upload_date, upload_time, file_name, revision_no, uploaded_by, file_data) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (state, formatted_date, current_time, virtual_filename, revision_no, role, json.dumps(new_data)))

        # 2. Insert into Intraday Table (As Revision 0) - Standard Practice
        # Delete existing Rev 0 if any
        cursor.execute("DELETE FROM intraday_file_uploads WHERE upload_date = %s AND state_id = %s AND revision_no = 0", (formatted_date, state))
        
        cursor.execute("""
            INSERT INTO intraday_file_uploads (state_id, upload_date, upload_time, file_name, revision_no, uploaded_by, file_data, file_type) 
            VALUES (%s, %s, %s, %s, 0, %s, %s, 'D')
        """, (state, formatted_date, current_time, virtual_filename, role, json.dumps(new_data)))

        conn.commit()

        # --- Email Logic (Preserved from your snippet) ---
        # (Assuming fetch_max_revision_data and send_mail_with_attachment are defined elsewhere)
        # ... [Email Logic Here if needed] ...

        return jsonify({'message': f'Data saved successfully. Revision {revision_no} created.'})

    except Exception as e:
        if conn: conn.rollback()
        # log_error("upload_dayahead", e)
        print(f"Error: {e}")
        return jsonify({'error': str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@app.route("/api/downloaddayahead", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_dayahead", enforce_state=True)
def downloadDayAhead():
    conn = None
    cursor = None
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        data_json = request.get_json()
        state_id = data_json.get('state')
        upload_date = data_json.get('date')
        revision_no = data_json.get('revision')
        if isinstance(revision_no, dict): revision_no = revision_no.get('revision')

        # Fetch Data
        query = """
            SELECT t1.file_data, t2.state_name
            FROM public.file_uploads t1
            JOIN public.states t2 ON t1.state_id = t2.state_id
            WHERE t1.state_id = %s AND t1.upload_date = %s AND t1.revision_no = %s
        """
        cursor.execute(query, (state_id, upload_date, revision_no))
        result = cursor.fetchone()

        if not result: return jsonify({'status': 'failure', 'error': 'Data not found'}), 404
        
        file_data = result[0] if isinstance(result[0], list) else json.loads(result[0])
        state_name = result[1]

        # Load Format
        format_path = os.path.join(shared_drive_path, "FORMATS", "Day-Ahead Demand Forecast format (from States).xlsx")
        if not os.path.exists(format_path): return jsonify({'status': 'failure', 'error': 'Format file missing'}), 404

        workbook = load_workbook(format_path)
        sheet = workbook.active

        # Helper to safely write
        def safe_write(r, c, v):
            cell = sheet.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = v

        # Metadata (Adjust coordinates based on your specific Day Ahead Excel layout)
        safe_write(2, 3, state_name) # C2
        safe_write(3, 3, upload_date) # C3

        # Write Data
        start_row = 8
        for row_idx, row_data in enumerate(file_data):
            r = start_row + row_idx
            for col_idx, val in enumerate(row_data):
                safe_write(r, col_idx + 1, val)

        # Footer (Total MUs)
        footer_row = start_row + len(file_data)
        safe_write(footer_row, 2, "Total MUs")
        
        # Calculate MUs for columns 3 onwards (Index 2)
        if file_data:
            for col_idx in range(2, len(file_data[0])):
                total = sum(float(row[col_idx]) for row in file_data if row[col_idx] not in [None, ''])
                safe_write(footer_row, col_idx + 1, round(total/4000, 2))

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name=f"DayAhead_{state_name}_{upload_date}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'status': 'failure', 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    

    
@app.route("/api/downloadintraday", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_intraday", enforce_state=True)
def downloadIntraday():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Parse the JSON input from the frontend
        data_json = request.get_json()
        state_id = data_json.get('state')
        upload_date = data_json.get('date')  # Expected in 'YYYY-MM-DD' format
        revision_no_data = data_json.get('revision')

        # Handle revision number extraction
        if isinstance(revision_no_data, dict):
            revision_no = revision_no_data.get('revision')
        else:
            revision_no = revision_no_data

        # Query the database
        query = """
            SELECT t1.file_data, t2.state_name
            FROM public.intraday_file_uploads t1
            JOIN public.states t2 ON t1.state_id = t2.state_id
            WHERE t1.state_id = %s
            AND t1.upload_date = %s
            AND t1.revision_no = %s
        """
        cursor.execute(query, (state_id, upload_date, revision_no))
        result = cursor.fetchone()

        if not result:
            return jsonify({'status': 'failure', 'error': 'Data not found for the specified criteria'}), 404

        file_data_raw, state_name = result
        
        # Ensure file_data is a list
        if isinstance(file_data_raw, str):
            file_data = json.loads(file_data_raw)
        else:
            file_data = file_data_raw

        # Define the path to the Format File
        format_path = os.path.join(shared_drive_path, "FORMATS", "Intraday Demand Forecast format (from States).xlsx")

        if not os.path.exists(format_path):
            return jsonify({'status': 'failure', 'error': 'Format template not found on server'}), 404

        # Load the Format Workbook
        workbook = load_workbook(format_path)
        sheet = workbook.active

        # --- Helper Function to Safely Write ---
        def safe_write(row, col, value):
            cell = sheet.cell(row=row, column=col)
            # Only write if it is NOT a MergedCell (read-only part of a merge)
            if not isinstance(cell, MergedCell):
                cell.value = value

        # --- Fill Header Metadata ---
        # Using coordinate assignment directly handles merges gracefully usually, 
        # but safely accessing via cell() is safer for dynamic logic.
        safe_write(2, 3, state_name) # C2
        safe_write(3, 3, upload_date) # C3

        # --- Write Data Rows ---
        start_row = 8
        for row_index, row_data in enumerate(file_data):
            current_row = start_row + row_index
            for col_index, cell_value in enumerate(row_data):
                # Columns in openpyxl are 1-based. col_index is 0-based.
                safe_write(current_row, col_index + 1, cell_value)

        # --- Write Footer (Total MUs) ---
        footer_row = start_row + len(file_data)
        
        # Write Label "Total MUs" in Column B (Index 2)
        safe_write(footer_row, 2, "Total MUs")

        # Calculate and Write Sums/4000
        if file_data and len(file_data) > 0:
            num_cols = len(file_data[0])
            
            # Iterate through columns starting from index 2 (3rd column) to the end
            for col_idx in range(2, num_cols):
                col_sum = 0
                for row in file_data:
                    try:
                        val = float(row[col_idx])
                    except (ValueError, TypeError):
                        val = 0
                    col_sum += val
                
                calculated_mu = round(col_sum / 4000, 2)
                
                # Write to Excel
                safe_write(footer_row, col_idx + 1, calculated_mu)

        # Save to Buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        download_filename = f"Intraday_{state_name}_{upload_date}_Rev{revision_no}.xlsx"

        return send_file(
            output, 
            as_attachment=True, 
            download_name=download_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        # log_error("downloadintraday", e) 
        print(f"Error in downloadintraday: {str(e)}") # Helpful for debugging
        return jsonify({'status': 'failure', 'error': str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()







# @app.route('/api/pendingentries', methods=['GET'])
# @jwt_required(locations=["cookies"])
# @session_token_required
# @token_required
# @rbac_required("pending_entries")
# def pendingTimingEntries():
#     try:
#         response = requests.get("https://oms2.srldc.in/Codebook/getDueTimingEntryData")
#         data = response.json()["data"]["dueTiminingEntryList"]

#         # Read normalized role + username from session_token_required
#         user = request.user
#         username = user["username"]
#         role = user["role"]    # admin / nldc / t_user
#         # state_id = user["state_id"]  # not needed

#         states_entities_dict = {
#             "t_kar_state": ['kptcl', 'KTL'],
#             "t_ap_state": ['LKPPLSTG3', 'APTRANSCO'],
#             "t_tn_state": ['TN','TANTRANSCO'],
#             "t_ker_state": ['kseb', 'kseb_sd'],
#             "t_tg_state": ['tgtransco'],
#             "t_pondy_state": [],
#             "pgcil_sr_1": ['PGCIL SR-1'],
#             "pgcil_sr_2": ['PGCIL SR-2'],
#             "nldc": ['nldc'],
#             "greenko": ['Greenko_Kurnool']
#         }

#         entities_data = []
#         id = 1

#         # t_user → username is like t_kar_state, t_ap_state...
#         if role in ('t_user','nldc'):
#             entity_list = states_entities_dict.get(username, [])
#             for i in data:
#                 if (
#                     any(item.lower() in i["codeIssuedto"].lower().split(",") for item in entity_list)
#                     or i["codeRequestedby"] in entity_list
#                 ) and i["constituentEnteredTime"] == "":
#                     entities_data.append({
#                         "id": id,
#                         "codeIssueTime": i["codeIssuedTime"],
#                         "elementType": i["entityFeatureName"],
#                         "elementName": i["elementName"],
#                         "switching": i["end"],
#                         "nldcCode": i["nldcCode"] + "/" + i["otherRegionCode"],
#                         "srldcCode": i["codeNo"],
#                         "category": i["outageCategory"],
#                         "codeIssuedTo": i["codeIssuedto"],
#                         "codeRequestedBy": i["codeRequestedby"],
#                         "codeId": i["codeId"],
#                         "isSelected": False
#                     })
#                     id += 1

#         else:
#             # admin or nldc → full list
#             for i in data:
#                 if i["constituentEnteredTime"] == "":
#                     entities_data.append({
#                         "id": id,
#                         "codeIssueTime": i["codeIssuedTime"],
#                         "elementType": i["entityFeatureName"],
#                         "elementName": i["elementName"],
#                         "switching": i["end"],
#                         "nldcCode": i["nldcCode"] + "/" + i["otherRegionCode"],
#                         "srldcCode": i["codeNo"],
#                         "category": i["outageCategory"],
#                         "codeIssuedTo": i["codeIssuedto"],
#                         "codeRequestedBy": i["codeRequestedby"],
#                         "codeId": i["codeId"],
#                         "isSelected": False
#                     })
#                     id += 1

#         if not entities_data:
#             return jsonify(status="failure", message="Data is empty!")

#         return jsonify(status="success", message="Fetched successfully!", data=entities_data)

#     except Exception as e:
#         log_error("pendingtimingentries", e)
#         return jsonify(status="failure", message="There is a problem in fetching the data!")



# --- Imports needed at the top ---
import ssl
import requests
from flask import jsonify, request
from requests.adapters import HTTPAdapter
from urllib3.poolmanager import PoolManager
from urllib3.util.ssl_ import create_urllib3_context

# --- The Adapter Class ---
class LegacySSLAdapter(HTTPAdapter):
    def init_poolmanager(self, connections, maxsize, block=False):
        ctx = create_urllib3_context()
        ctx.load_default_certs()
        # Enable legacy renegotiation (value 0x4)
        ctx.options |= 0x4 
        self.poolmanager = PoolManager(
            num_pools=connections,
            maxsize=maxsize,
            block=block,
            ssl_context=ctx
        )

# --- Your Route ---
@app.route('/api/pendingentries', methods=['GET'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("pending_entries")
def pendingTimingEntries():
    try:
        # Create a session and mount the custom adapter
        session = requests.Session()
        session.mount("https://", LegacySSLAdapter())
        
        # Use the session to make the request
        response = session.get("https://oms2.srldc.in/Codebook/getDueTimingEntryData")
        
        # Proceed as normal
        data = response.json()["data"]["dueTiminingEntryList"]

        # Read normalized role + username from session_token_required
        user = request.user
        username = user["username"]
        role = user["role"]    # admin / nldc / t_user
        
        states_entities_dict = {
            "t_kar_state": ['kptcl', 'KTL'],
            "t_ap_state": ['LKPPLSTG3', 'APTRANSCO'],
            "t_tn_state": ['TN','TANTRANSCO'],
            "t_ker_state": ['kseb', 'kseb_sd'],
            "t_tg_state": ['tgtransco'],
            "t_pondy_state": [],
            "pgcil_sr_1": ['PGCIL SR-1'],
            "pgcil_sr_2": ['PGCIL SR-2'],
            "nldc": ['nldc'],
            "greenko": ['Greenko_Kurnool']
        }

        entities_data = []
        id = 1

        # t_user → username is like t_kar_state, t_ap_state...
        if role in ('t_user','nldc'):
            entity_list = states_entities_dict.get(username, [])
            for i in data:
                if (
                    any(item.lower() in i["codeIssuedto"].lower().split(",") for item in entity_list)
                    or i["codeRequestedby"] in entity_list
                ) and i["constituentEnteredTime"] == "":
                    entities_data.append({
                        "id": id,
                        "codeIssueTime": i["codeIssuedTime"],
                        "elementType": i["entityFeatureName"],
                        "elementName": i["elementName"],
                        "switching": i["end"],
                        "nldcCode": i["nldcCode"] + "/" + i["otherRegionCode"],
                        "srldcCode": i["codeNo"],
                        "category": i["outageCategory"],
                        "codeIssuedTo": i["codeIssuedto"],
                        "codeRequestedBy": i["codeRequestedby"],
                        "codeId": i["codeId"],
                        "isSelected": False
                    })
                    id += 1

        else:
            # admin or nldc → full list
            for i in data:
                if i["constituentEnteredTime"] == "":
                    entities_data.append({
                        "id": id,
                        "codeIssueTime": i["codeIssuedTime"],
                        "elementType": i["entityFeatureName"],
                        "elementName": i["elementName"],
                        "switching": i["end"],
                        "nldcCode": i["nldcCode"] + "/" + i["otherRegionCode"],
                        "srldcCode": i["codeNo"],
                        "category": i["outageCategory"],
                        "codeIssuedTo": i["codeIssuedto"],
                        "codeRequestedBy": i["codeRequestedby"],
                        "codeId": i["codeId"],
                        "isSelected": False
                    })
                    id += 1

        if not entities_data:
            return jsonify(status="failure", message="Data is empty!")

        return jsonify(status="success", message="Fetched successfully!", data=entities_data)

    except Exception as e:
        log_error("pendingtimingentries", e)
        return jsonify(status="failure", message="There is a problem in fetching the data!")



@app.route('/api/getelementpreviouscodes', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_previous_codes", enforce_state=True)
def getPreviousCodes():
    try:
        params = request.get_json()

        print(params)

        params = params["params"]
        from_date = params["from_date"]
        to_date = params["to_date"]
        feature_name = params.get("feature_name", "")  # Provide a default value if not present
        pwcEntityId = params.get("pwcEntityId", "")  # Provide a default value if not present


        jwt_data = get_jwt()


        # Create the payload to send in the POST request
        payload = json.dumps({
            "startDate": from_date,
            "endDate": to_date,
            "pwcEntityId": "",
            "elementId": "",
            "featureName": ""
        })

        headers = {
            "Content-Type": "application/json"
            # "Authorization": "Bearer your_token"  # If authentication is required
        }

        response = requests.post("https://oms2.srldc.in/Codebook/getElementPreviousCodes", data=payload, headers=headers)

        # print(response.status_code)
        # print(response.json())

        header_data = dict(request.headers)
        # print(header_data)
        # token = header_data['Authorization'].split()[1]
        # x = decode_token(token, csrf_value=None, allow_expired=False)

        username = jwt_data['sub']   # username in database

        # username = jwt_data['sub']   # username in database

        role = jwt_data['role']       # role in database


        # print("Username:", username,"Role:", role)



        states_entities_dict = {}
        states_entities_dict["t_kar_state"] = ['kptcl', 'KTL']
        states_entities_dict["t_ap_state"] = ['LKPPLSTG3', 'APTRANSCO']
        states_entities_dict["t_tn_state"] = ['TN','TANTRANSCO']
        states_entities_dict["t_ker_state"] = []
        states_entities_dict["t_tg_state"] = ['TSTRANSCO']
        states_entities_dict["t_pondy_state"] = []
        states_entities_dict["pgcil_sr_1"] = ['PGCIL SR-1']
        states_entities_dict["pgcil_sr_2"] = ['PGCIL SR-2']
        states_entities_dict["greenko"] = ['Greenko_Kurnool']
        states_entities_dict["nldc"] = ['nldc']


        admin_states_list = []

        for key, value in states_entities_dict.items():
            for v in value:
                admin_states_list.append(v)


        # PGCIL SR-2, PGCIL SR-1

        # print(states_entities_dict[username])



        if response.status_code == 200:
            data = response.json()["data"]["previousCodes"]
            # print(data)
            # print(data['data']["previousCodes"].keys())
                    
            entities_data = []

            # print(data)

            id = 1

            if role in ('t_user','nldc'):
                for i in data:
                    if (i["codeIssuedto"] in states_entities_dict[username] or i["codeRequestedby"] in states_entities_dict[username]):
                        entities_data.append({"id": id, "codeIssueTime": i["codeIssuedTime"], "elementType": i["entityFeatureName"], "elementName": i["elementName"], "switching": i["end"], "nldcCode": i["nldcCode"], "srldcCode": i["codeNo"], "category": i["outageCategory"], "codeIssuedTo": i["codeIssuedto"], "codeRequestedBy": i["codeRequestedby"], "codeId": i["codeId"], "isSelected": False})
                        id = id + 1
            
            else:
                for i in data:
                    # print(i.keys())
                    # print('constituentEnteredTime' in i.keys())
                    if (i["codeIssuedto"] in admin_states_list or i["codeRequestedby"] in admin_states_list):
                        entities_data.append({"id": id, "codeIssueTime": i["codeIssuedTime"], "elementType": i["entityFeatureName"], "elementName": i["elementName"], "switching": i["end"], "nldcCode": i["nldcCode"], "srldcCode": i["codeNo"], "category": i["outageCategory"], "codeIssuedTo": i["codeIssuedto"], "codeRequestedBy": i["codeRequestedby"], "codeId": i["codeId"], "isSelected": False})
                        id = id + 1

            # print(type(entities_data[0]["codeIssueTime"]))
            # formatted_data_list = [{'date_key': datetime.strptime(item['date_key'], '%Y-%m-%dT%H:%M:%S.%f').strftime('%d/%m/%Y, %H:%M') for item in data_list]
                                
            for i in range(len(entities_data)):
                entities_data[i]["codeIssueTime"] = datetime.strptime(entities_data[i]["codeIssueTime"], '%Y-%m-%dT%H:%M:%S.%f').isoformat()
            
            # print(entities_data[0]["codeIssueTime"])


            if len(entities_data) == 0:
                return jsonify(status="failure", message="Data is Empty!")
        else:
            # print(response.status_code)
            return jsonify(status="failure", message="There is a problem in fetching data, Please contact SRLDC IT!")
        return jsonify(status="success", message="Message fetched successfully!", data=entities_data )
    except Exception as e:
        # print(e)
        log_error("getpreviouscodes", e)
        # cursor.close()
        return jsonify(status="failure", message="There is a problem in fetching the data, Please contact SRLDC IT!")
 









# @app.route('/api/genoutages', methods=['POST'])
# @jwt_required(locations=["cookies"])
# @token_required
# @session_token_required
# def getGenOutages():
#     try:
#         jwt_data = get_jwt()
#         parameter = request.get_json()


        

#         username = jwt_data['sub']   # username in database

#         role = jwt_data['role']       # role in database

#         print("Username:", username,"Role:", role)

#         states_entities_dict = {}
#         states_entities_dict["t_kar_state"] = ['Karnataka', 'ISGS', 'Regional IPPS']
#         states_entities_dict["t_ap_state"] = ['Andhra Pradesh', 'ISGS', 'Regional IPPS']
#         states_entities_dict["t_tn_state"] = ['Tamilnadu', 'ISGS', 'Regional IPPS']
#         states_entities_dict["t_ker_state"] = ['Kerala', 'ISGS', 'Regional IPPS']
#         states_entities_dict["t_tg_state"] = ['Telangana', 'ISGS', 'Regional IPPS']
#         states_entities_dict["t_pondy_state"] = []
#         states_entities_dict["pgcil_sr_1"] = ['PGCIL SR-1']
#         states_entities_dict["pgcil_sr_2"] = ['PGCIL SR-2']
#         states_entities_dict["nldc"] = ['nldc']



#         if parameter["type"] == "thermal":
#             api_url = "https://oms2.srldc.in/Codebook/getScadaGenDisplayOutagesData?displayType=Thermal"
#         elif parameter["type"] == "hydro":
#             api_url = "https://oms2.srldc.in/Codebook/getScadaGenDisplayOutagesData?displayType=Hydro"
        
#         thermal_data = fetch_thermal_outages_data(api_url)

#         if not thermal_data:
#             return jsonify(status="failure", message="Failed to retrieve thermal outages data.")

#         classifications = thermal_data.get("data", {}).get("classifications", {})
#         formatted_data = []

#         def extract_unit_sizes(units_string):
#             """Extracts unit sizes like 3x200 and 500 and expands them to list of integers."""
#             clean = re.sub(r'<.*?>', '', units_string)  # Remove HTML tags
#             parts = clean.replace(" ", "").split("+")
#             ic_list = []

#             for part in parts:
#                 if "x" in part:
#                     match = re.match(r'(\d+)x(\d+)', part)
#                     if match:
#                         count, size = map(int, match.groups())
#                         ic_list.extend([size] * count)
#                 elif part.isdigit():
#                     ic_list.append(int(part))

#             return ic_list

#         for state, stations in classifications.items():
#             for station in stations:
#                 station_name = station.get("Station Name", "")
#                 units_string = station.get("Units", "")
#                 unit_sizes = extract_unit_sizes(units_string)
#                 units = station.get("UnitNames", [])
#                 statuses = station.get("unitsStatus", [])

#                 for i, unit in enumerate(units):
#                     status_info = statuses[i] if i < len(statuses) and isinstance(statuses[i], dict) else {}
#                     days_in_raw = status_info.get("No of Days In", "") or status_info.get("No of Days Out", "")
#                     days_in_cleaned = str(days_in_raw).strip().lstrip(',').strip()


#                     # Check if user is 'admin'
#                     if username == "admin":
#                         formatted_data.append({
#                             "state": state,
#                             "station_name": station_name,
#                             "Units": re.sub(r'<.*?>', '', station.get("Units", "")),
#                             "unit_name": unit,
#                             "ic": unit_sizes[i] if i < len(unit_sizes) else "",
#                             "status": status_info.get("status", ""),
#                             "elementId": status_info.get("elementId", ""),
#                             "color": status_info.get("color", ""),
#                             "outageCategory": status_info.get("outageCategory", ""),
#                             "details": status_info.get("Details", ""),
#                             "days_in": days_in_cleaned,
#                             "time_dday": status_info.get("TimeDday", "")
#                         })
#                     # For non-admin users, apply filtering based on states_entities_dict
#                     elif any(item.lower() in state.lower().split(",") for item in states_entities_dict.get(username, [])):
#                         formatted_data.append({
#                             "state": state,
#                             "station_name": station_name,
#                             "Units": re.sub(r'<.*?>', '', station.get("Units", "")),
#                             "unit_name": unit,
#                             "ic": unit_sizes[i] if i < len(unit_sizes) else "",
#                             "status": status_info.get("status", ""),
#                             "elementId": status_info.get("elementId", ""),
#                             "color": status_info.get("color", ""),
#                             "outageCategory": status_info.get("outageCategory", ""),
#                             "details": status_info.get("Details", ""),
#                             "days_in": days_in_cleaned,
#                             "time_dday": status_info.get("TimeDday", "")
#                         })


#         return jsonify(status="success", message="Data Fetched Successfully!", data=formatted_data)

#     except Exception as e:
#         # print("Error in getGenOutages:", e)
#         log_error("genoutages", e)
#         return jsonify(status="failure", message="Problem in Fetching data, Please contact SRLDC IT!")



@app.route('/api/genoutages', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("gen_outages", enforce_state=True)
def getGenOutages():
    try:


        jwt_data = get_jwt()
        parameter = request.get_json()
        username = jwt_data['sub']
        role = jwt_data['role']

        states_entities_dict = {
            "t_kar_state": ['Karnataka', 'ISGS', 'Regional IPPS'],
            "t_ap_state": ['Andhra Pradesh', 'ISGS', 'Regional IPPS'],
            "t_tn_state": ['Tamilnadu', 'ISGS', 'Regional IPPS'],
            "t_ker_state": ['Kerala', 'ISGS', 'Regional IPPS'],
            "t_tg_state": ['Telangana', 'ISGS', 'Regional IPPS'],
            "t_pondy_state": [],
            "pgcil_sr_1": ['PGCIL SR-1'],
            "pgcil_sr_2": ['PGCIL SR-2'],
            "nldc": ['nldc']
        }

        if parameter["type"] == "thermal":
            api_url = "https://oms2.srldc.in/Codebook/getScadaGenDisplayOutagesData?displayType=Thermal"
        elif parameter["type"] == "hydro":
            api_url = "https://oms2.srldc.in/Codebook/getScadaGenDisplayOutagesData?displayType=Hydro"

        thermal_data = fetch_thermal_outages_data(api_url)
        if not thermal_data:
            return jsonify(status="failure", message="Failed to retrieve thermal outages data.")

        classifications = thermal_data.get("data", {}).get("classifications", {})
        formatted_data = []

        def extract_unit_sizes(units_string):
            clean = re.sub(r'<.*?>', '', units_string)
            parts = clean.replace(" ", "").split("+")
            ic_list = []

            for part in parts:
                if "x" in part:
                    match = re.match(r'(\d+)x(\d+)', part)
                    if match:
                        count, size = map(int, match.groups())
                        ic_list.extend([size] * count)
                elif part.isdigit():
                    ic_list.append(int(part))
            return ic_list

        for state, stations in classifications.items():
            for station in stations:
                station_name = station.get("Station Name", "")
                units_string = station.get("Units", "")
                full_station_key = f"{station_name} ({units_string})"  # Unique for merged stations

                unit_sizes = extract_unit_sizes(units_string)
                units = station.get("UnitNames", [])
                statuses = station.get("unitsStatus", [])

                # Clean and pair only valid statuses
                valid_statuses = [s for s in statuses if isinstance(s, dict) and s.get("status")]
                mapped_unit_indices = 0

                for status_info in valid_statuses:
                    if mapped_unit_indices >= len(units):
                        break

                    unit = units[mapped_unit_indices]
                    days_in_raw = status_info.get("No of Days In", "") or status_info.get("No of Days Out", "")
                    days_in_cleaned = str(days_in_raw).strip().lstrip(',').strip()

                    # Check user role
                    if username == "admin" or any(item.lower() in state.lower().split(",") for item in states_entities_dict.get(username, [])):
                        formatted_data.append({
                            "state": state,
                            "station_name": station_name,
                            "Units": re.sub(r'<.*?>', '', station.get("Units", "")),
                            "unit_name": unit,
                            "ic": unit_sizes[mapped_unit_indices] if mapped_unit_indices < len(unit_sizes) else "",
                            "status": status_info.get("status", ""),
                            "elementId": status_info.get("elementId", ""),
                            "color": status_info.get("color", ""),
                            "outageCategory": status_info.get("outageCategory", ""),
                            "details": status_info.get("Details", ""),
                            "days_in": days_in_cleaned,
                            "time_dday": status_info.get("TimeDday", "")
                        })
                    mapped_unit_indices += 1

        return jsonify(status="success", message="Data Fetched Successfully!", data=formatted_data)

    except Exception as e:
        log_error("genoutages", e)
        return jsonify(status="failure", message="Problem in Fetching data, Please contact SRLDC IT!")



@app.route('/api/submitpartialoutagedetail', methods=['POST'])
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("submit_partial_outage", enforce_state=True)
def submitOutageDetail():
    try:
        OMS2_BASE_URL = "https://oms2.srldc.in/Codebook/partial_outage_by_date"
        wrapper = request.get_json()
        body = wrapper.get("data", {})  #  your frontend is sending {'data': { 'date':..., 'data': [...] }}

        date = body.get("date")
        updates = body.get("data", [])

        print("Date received:", date)
        print("Updates received:", updates)

        if not date:
            return jsonify(status="failure", message="Date is required.")

        success_ids = []
        failed_updates = []

        headers = {
            "Content-Type": "application/json"
        }

        for entry in updates:
            try:
                payload = {
                    "id": entry["id"],
                    "reason_not_attaining_full_generation": entry["reason_not_attaining_full_generation"]
                }

                res = requests.post(f"{OMS2_BASE_URL}/{date}/", json=payload, headers=headers)
                res_data = res.json()

                #  Now check message also if status is not 'success'
                if res.status_code == 200:
                    if res_data.get("status") == "success":
                        success_ids.append(entry["id"])
                    elif "updated successfully" in res_data.get("message", "").lower():
                        success_ids.append(entry["id"])
                    else:
                        failed_updates.append({
                            "id": entry["id"],
                            "error": res_data.get("message", "Unknown error")
                        })
                else:
                    failed_updates.append({
                        "id": entry["id"],
                        "error": f"HTTP {res.status_code}"
                    })

            except Exception as e:
                failed_updates.append({
                    "id": entry.get("id"),
                    "error": str(e)
                })

        if failed_updates:
            return jsonify(status="partial_success", success_ids=success_ids, failed=failed_updates)

        return jsonify(status="success", message="All reasons updated successfully.")

    except Exception as e:
        log_error("submitoutagedetail", e)
        return jsonify(status="failure", message="Problem in submitting data, please contact SRLDC IT.")

@app.route('/api/getpartialoutagedetail', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_partial_outage", enforce_state=True)
def submit_outage_detail():
    try:
        # Get the posted date from frontend
        data = request.get_json()
        date = data.get('date')  # expecting format "YYYY-MM-DD"
        
        if not date:
            return jsonify({"status": "error", "message": "No date provided"}), 400

        # Now call the external API
        api_url = f"https://oms2.srldc.in/Codebook/partial_outage_by_date/{date}/"
        response = requests.get(api_url, timeout=10)

        if response.status_code == 200:
            outages_data = response.json()  # directly forward the API JSON response
            return jsonify({"status": "success", "data": outages_data})
        else:
            return jsonify({"status": "error", "message": "Failed to fetch data from SRLDC"}), 500

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/weekaheadformat')
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_weekahead_format")
def weekAheadFormat():
    try:
        jwt_data = get_jwt()
        # Function to find the date of the next Monday
        def next_monday(d):
            # Calculate the number of days until the next Monday
            days_til_monday = (7 - d.weekday()) % 7
            if days_til_monday == 0:
                # If today is Monday, add 7 days to get the next Monday
                days_til_monday = 7
            return d + timedelta(days=days_til_monday)

        # Function to create the 2D list
        def create_2d_list(start_date, num_days, num_blocks_per_day):
            data = []

            for day in range(num_days):
                current_date = start_date + timedelta(days=day)
                for block in range(num_blocks_per_day):
                    current_time = (datetime.min + timedelta(minutes=block * 15)).time()
                    next_time = (datetime.min + timedelta(minutes=(block + 1) * 15)).time()
                    timestamp = f"{current_time:%H:%M} - {next_time:%H:%M}"
                    row = [current_date.strftime('%Y-%m-%d'), block + 1, timestamp] + [0] * 19
                    data.append(row)

            return data

        # Find the next Monday
        today = datetime.now()
        next_monday_date = next_monday(today)

        # Create the 2D list for 7 days with 96 blocks each
        num_days = 7
        num_blocks_per_day = 96
        data = create_2d_list(next_monday_date, num_days, num_blocks_per_day)

        
        return jsonify(status="success", data=data)
    
    except Exception as e:
        log_error("weekaheadformat", e)
        # cursor.close()
        return jsonify(msg="Problem in fetching the data, Please contact SRLDC IT", status="failure")


@app.route('/api/monthaheadformat')
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_monthahead_format")
def monthAheadFormat():

    try:
        jwt_data = get_jwt()
        def create_2d_list(start_date, end_date, num_blocks_per_day):
            data = []
            
            while start_date <= end_date:
                for block in range(num_blocks_per_day):
                    current_time = (datetime.min + timedelta(minutes=block * 15)).time()
                    next_time = (datetime.min + timedelta(minutes=(block + 1) * 15)).time()
                    timestamp = f"{current_time:%H:%M} - {next_time:%H:%M}"
                    row = [start_date.strftime('%Y-%m-%d'), block + 1, timestamp] + [0] * 19
                    data.append(row)
                
                start_date += timedelta(days=1)
            
            return data

        today = datetime.now()

        # Calculate the start of the next month, handling month rollover if necessary
        next_month = (today.month % 12) + 1
        next_year = today.year + (today.month // 12)

        start_of_next_month = datetime(next_year, next_month, 1)

        # Calculate the end of the next month by getting the first day of the month after next, then subtracting one day
        following_month = (next_month % 12) + 1
        following_year = next_year + (next_month // 12)

        end_of_next_month = datetime(following_year, following_month, 1) - timedelta(days=1)


        # Create the 2D list for the entire next month with 96 blocks each day
        num_blocks_per_day = 96
        data = create_2d_list(start_of_next_month, end_of_next_month, num_blocks_per_day)

        # print(len(data))

        return jsonify(data=data, status="success")
    except Exception as e:
        log_error("monthaheadformat", e)
        # cursor.close()
        return jsonify(status="failure", msg="Problem in Fetching the data, Please contact SRLDC IT!")



@app.route('/api/yearaheadformat')
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_yearahead_format")
def yearheadFormat():
    try:
        jwt_data = get_jwt()
        def create_2d_list_for_year(start_date, end_date, num_blocks_per_day):
            data = []

            while start_date <= end_date:
                for block in range(num_blocks_per_day):
                    current_time = (datetime.min + timedelta(minutes=block * 60)).time()
                    next_time = (datetime.min + timedelta(minutes=(block + 1) * 60)).time()
                    row = [start_date.strftime('%Y-%m-%d'), block + 1] + [0] * 19
                    data.append(row)

                start_date += timedelta(days=1)

            return data

        # Get the current date
        today = datetime.now()
        
        # Define the financial year starting from the next April 1st
        if today.month >= 4:
            start_of_financial_year = datetime(today.year + 1, 4, 1)
            end_of_financial_year = datetime(today.year + 2, 3, 31)
        else:
            start_of_financial_year = datetime(today.year, 4, 1)
            end_of_financial_year = datetime(today.year + 1, 3, 31)

        # num_blocks_per_day = 24
        data_for_financial_year = create_2d_list_for_year(start_of_financial_year, end_of_financial_year, num_blocks_per_day=24)

        return jsonify(data=data_for_financial_year, status="success")
    
    except Exception as e:
        log_error("yearaheadformat", e)
        cursor.close()
        return jsonify(status="failure", msg="Problem in Fetching the data, Please contact SRLDC IT!")

    
@app.route('/api/submitentries', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("submit_entries", enforce_state=True)
def submitTimingEntries():
    try:
        jwt_data = get_jwt()
        data = request.get_json()

        # data = json.loads(data)

        data = data["data"]

        # import pdb
        # pdb.set_trace()




        for i in data:

            if 'T' in i["codeIssueTime"]:
                i["codeIssueTime"] = datetime.fromisoformat(i["codeIssueTime"]).strftime("%Y-%m-%d %H:%M")
            else:
                i["codeIssueTime"] = datetime.strptime(i["codeIssueTime"], "%d/%m/%Y, %H:%M")

                # Format the datetime object as "yyyy-mm-dd hh:mm"
                i["codeIssueTime"] = i["codeIssueTime"].strftime("%Y-%m-%d %H:%M")

            
            # import pdb
            # pdb.set_trace()

            response = requests.get("https://oms2.srldc.in/Codebook/EnterConstituentTime?codeId={0}&enteredTime={1}".format(i["codeId"], i["codeIssueTime"]))

            # print(response)

        return jsonify(status="success", message="Sent Successfully!")
    except Exception as e:
        log_error("submittimingentries", e)
        cursor.close()
        # print(error)
        return jsonify(status="failure", message="There is a problem, Please contact SRLDC IT!")
    
    

@app.route('/api/weeekaheadforecast', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("process_weekahead", enforce_state=True)
def weekAheadForecast():
    try: 
        jwt_data = get_jwt()
        return jsonify(msg="Successful!", status="failure")
    except Exception as error:
        return jsonify(msg="There is some problem, Please contact SRLDC IT!")
    

@app.route('/api/parseweekaheadexcel', methods=['POST'])
@jwt_required(locations=["cookies"])
def parse_weekahead_excel():
    try:
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})
        
        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        # Use Pandas to read the file
        df = pd.read_excel(file, header=None) 
        data_list = df.values.tolist()
        
        cleaned_data = []
        blanks_replaced_flag = False

        # Column Mapping for Error Reporting (Indices based on array 0-based index)
        # Excel: Date(A), Block(B), Period(C), Demand(D), Thermal(E)...
        # List:  0        1         2          3          4
        col_names = {
            3: "Forecasted Demand", 4: "Thermal", 5: "Gas", 6: "Hydro", 7: "Total Own Gen",
            8: "Solar", 9: "Wind", 10: "Other RES", 11: "Total RES", 12: "ISGS/LTA",
            13: "Bilateral", 14: "Total Avail", 15: "Gap", 16: "Bilateral Proc",
            17: "Power Exchange", 18: "Shortage", 19: "Relief", 20: "Load Shedding", 21: "Reactive"
        }

        for row_idx, row in enumerate(data_list):
            try:
                # --- Row Identification ---
                # Check if Block Number (index 1) exists. If NaN, skip (likely header or empty end row)
                if pd.isna(row[1]): continue 
                
                try:
                    block_num = int(row[1])
                except:
                    continue # Skip if block is not an integer (header)

                # Context for error messages
                # Handle Pandas Timestamp or string for Date
                raw_date = row[0]
                if isinstance(raw_date, pd.Timestamp):
                    row_date = raw_date.strftime('%d/%m/%Y')
                elif not pd.isna(raw_date):
                     row_date = str(raw_date)
                else:
                     return jsonify({'error': f'Missing Date at Block {block_num} (Row {row_idx+1}).'})

                row_time = str(row[2]) if not pd.isna(row[2]) else f"Block {block_num}"

                if 1 <= block_num <= 96:
                    # Slice to first 22 columns
                    original_slice = row[:22] 
                    
                    new_row = [row_date, block_num, original_slice[2]]

                    # --- 1. VALIDATE FORECASTED DEMAND (Index 3) ---
                    # Constraint: Cannot be Blank or Zero
                    demand_val = original_slice[3]
                    col_name_demand = col_names.get(3, "Demand")

                    if pd.isna(demand_val) or str(demand_val).strip() == '':
                        return jsonify({'error': f"'{col_name_demand}' is MISSING at Date: {row_date}, Time: {row_time}."})
                    
                    try:
                        demand_float = float(demand_val)
                        if demand_float == 0:
                            return jsonify({'error': f"'{col_name_demand}' cannot be ZERO at Date: {row_date}, Time: {row_time}."})
                        new_row.append(demand_float)
                    except ValueError:
                        return jsonify({'error': f"Invalid numeric value for '{col_name_demand}' at Date: {row_date}, Time: {row_time}."})

                    # --- 2. HANDLE OTHER COLUMNS (Index 4 to 21) ---
                    # Constraint: Can be 0. If Blank -> Replace with 0 + Warn.
                    for i in range(4, 22):
                        val = original_slice[i]
                        current_col_name = col_names.get(i, f"Column {i}")

                        if pd.isna(val) or str(val).strip() == '':
                            new_row.append(0.0)
                            blanks_replaced_flag = True
                        else:
                            try:
                                new_row.append(float(val))
                            except ValueError:
                                return jsonify({'error': f"Invalid numeric data in '{current_col_name}' at Date: {row_date}, Time: {row_time}."})
                    
                    cleaned_data.append(new_row)

            except Exception as e:
                return jsonify({'error': f"Row Processing Error at Row {row_idx+1}: {str(e)}"})

        # --- 3. MISSING ROWS VALIDATION ---
        # Expected rows: 96 blocks * 7 days = 672
        if len(cleaned_data) != 672:
            if len(cleaned_data) == 0:
                return jsonify({'error': 'File appears empty or format is incorrect (No valid blocks found).'})
            
            # Logic to find WHICH dates/blocks are missing
            try:
                # Infer Start Date from the first valid row found
                first_date_str = cleaned_data[0][0] # Format DD/MM/YYYY
                start_dt = datetime.strptime(first_date_str, '%d/%m/%Y')
                
                # Create a set of found (Date, Block) keys
                found_keys = set((r[0], r[1]) for r in cleaned_data)
                
                missing_info = []
                
                # Generate Expected Keys for 7 Days
                for day_offset in range(7):
                    curr_date = start_dt + timedelta(days=day_offset)
                    curr_date_str = curr_date.strftime('%d/%m/%Y')
                    
                    for b in range(1, 97):
                        if (curr_date_str, b) not in found_keys:
                            missing_info.append(f"{curr_date_str} Block {b}")
                            if len(missing_info) >= 3: break # Limit error msg length
                    if len(missing_info) >= 3: break
                
                if missing_info:
                     return jsonify({'error': f"Invalid Row Count ({len(cleaned_data)}/672). Data is MISSING for: {', '.join(missing_info)}..."})
                else:
                     # Fallback if count is wrong but duplicates exist or extra days
                     return jsonify({'error': f"Invalid Row Count ({len(cleaned_data)}). Expected 672. Check for duplicate or extra rows."})

            except ValueError:
                # If date format in excel wasn't DD/MM/YYYY, simple count error
                return jsonify({'error': f"Invalid Row Count. Expected 672 rows, found {len(cleaned_data)}."})


        # --- 4. SUMMATION VALIDATION (Thermal + Gas + Hydro == Total Own Sources) ---
        # Indices in cleaned_data: 0=Date, 1=Block, 2=Period, 3=Demand, 4=Thermal, 5=Gas, 6=Hydro, 7=Total(B)
        for row in cleaned_data:
            try:
                date_s, time_s = row[0], row[2]
                thermal, gas, hydro, total_own = row[4], row[5], row[6], row[7]
                
                if abs((thermal + gas + hydro) - total_own) > 0.1:
                     return jsonify({'error': f"Summation Mismatch at {date_s} ({time_s}). Thermal({thermal}) + Gas({gas}) + Hydro({hydro}) != Total({total_own})"})
            except:
                pass 

        success_msg = 'Parsed successfully. You can Upload the Data now!'
        if blanks_replaced_flag:
            success_msg = 'Parsed successfully. Warning: Some blank generation fields were automatically replaced with 0.'

        return jsonify({'message': success_msg, 'data': cleaned_data})

    except Exception as e:
        return jsonify({'error': f"Server Error: {str(e)}"})



@app.route('/api/uploadweekahead', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("upload_weekahead", enforce_state=True)
def upload_weekahead_data():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        state = request.form.get('state')
        fromDate = request.form.get('fromDate') # 'DD/MM/YYYY'
        toDate = request.form.get('toDate')     # 'DD/MM/YYYY'
        data_json = request.form.get('data') # Expecting data string here now, not file
        role = jwt_data['role']

        if not data_json: return jsonify({"error": "No data provided"}), 400
        new_data = json.loads(data_json)

        # Get State Name
        cursor.execute("select state_name from states where state_id=%s", (state,))
        state_name = cursor.fetchone()[0]

        # Calculate Revision
        cursor.execute("""
            SELECT COUNT(*) FROM week_ahead_file_uploads
            WHERE from_date = to_date(%s, 'DD/MM/YYYY')
              AND to_date = to_date(%s, 'DD/MM/YYYY')
              AND state_id = %s
        """, (fromDate, toDate, state))
        
        revision_no = cursor.fetchone()[0]

        # --- REVISION LOGIC (Preserve Past Data) ---
        # If re-uploading for same week, verify current time against row dates/blocks
        # Logic: If current time > Row Date+Block Time, keep old value.
        # Fetch previous revision
        if revision_no > 0:
            cursor.execute("""
                SELECT file_data FROM week_ahead_file_uploads
                WHERE from_date = to_date(%s, 'DD/MM/YYYY')
                  AND to_date = to_date(%s, 'DD/MM/YYYY')
                  AND state_id = %s
                  AND revision_no = %s
            """, (fromDate, toDate, state, revision_no - 1))
            res = cursor.fetchone()
            if res:
                previous_data = res[0]
                current_dt = datetime.datetime.now()
                
                for i, row in enumerate(previous_data):
                    # row[0] is Date string (DD/MM/YYYY), row[1] is Block
                    try:
                        row_date_str = row[0]
                        # Create a timestamp for the end of that block on that day
                        # Simplifying: If the Date is in the past (yesterday or before), keep old data
                        # If Date is today, check block.
                        # (Robust implementation requires parsing date string format from Jspreadsheet)
                        # Assuming row[0] matches the input format or standard ISO
                        pass # Implement strict time comparison if needed
                        
                        # Basic Preservation: If index matches, keep old. 
                        # REAL LOGIC: You likely want to overwrite WHOLE week for week-ahead usually, 
                        # or just overwrite future. Leaving as strict overwrite for now unless specific logic requested.
                    except: pass
        
        # Virtual Filename
        virtual_filename = f"DB_STORED_{fromDate}_{toDate}_{state_name}_rev{revision_no}"

        cursor.execute("""
            INSERT INTO week_ahead_file_uploads
            (state_id, from_date, to_date, upload_time, file_name, revision_no, uploaded_by, file_data)
            VALUES (%s, to_date(%s,'DD/MM/YYYY'), to_date(%s,'DD/MM/YYYY'),
                    NOW(), %s, %s, %s, %s)
        """, (state, fromDate, toDate, virtual_filename, revision_no, role, json.dumps(new_data)))

        conn.commit()

        return jsonify({
            "message": f"Data saved successfully. Revision-{revision_no} created.",
            "status": "success"
        })

    except Exception as e:
        if conn: conn.rollback()
        # log_error("uploadweekahead", e)
        return jsonify(message=str(e), status="failure")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    
@app.route("/api/downloadweekahead", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_weekahead", enforce_state=True)
def downloadWeekAhead():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        data_json = request.get_json()
        state_id = data_json.get('state')
        upload_from_date = data_json.get('from_date') 
        upload_to_date = data_json.get('to_date')
        revision_no = data_json.get('revision')

        # Fetch Data
        query = """
            SELECT t1.file_data, t2.state_name
            FROM public.week_ahead_file_uploads t1
            JOIN public.states t2 ON t1.state_id = t2.state_id
            WHERE t1.state_id = %s 
            AND t1.from_date = %s
            AND t1.to_date = %s 
            AND t1.revision_no = %s
        """
        cursor.execute(query, (state_id, upload_from_date, upload_to_date, revision_no))
        result = cursor.fetchone()

        if not result: return jsonify({'status': 'failure', 'error': 'Data not found'}), 404

        file_data = result[0] if isinstance(result[0], list) else json.loads(result[0])
        state_name = result[1]

        # Load Format
        format_path = os.path.join(shared_drive_path, "FORMATS", "Week-Ahead Demand Forecast format (from States).xlsx")
        if not os.path.exists(format_path): return jsonify({'status': 'failure', 'error': 'Format file missing'}), 404

        workbook = load_workbook(format_path)
        sheet = workbook.active

        # Helper
        def safe_write(r, c, v):
            cell = sheet.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = v

        # Metadata (Adjust coordinates)
        safe_write(2, 3, state_name) # C2
        safe_write(3, 3, f"{upload_from_date} to {upload_to_date}") # C3

        # Write Data
        start_row = 8
        for row_idx, row_data in enumerate(file_data):
            r = start_row + row_idx
            for col_idx, val in enumerate(row_data):
                safe_write(r, col_idx + 1, val)

        # Footer (Total MUs)
        footer_row = start_row + len(file_data)
        safe_write(footer_row, 3, "Total MUs") # Col 3 label

        # Calculate MUs for columns 4 onwards (Index 3: Demand)
        if file_data:
            for col_idx in range(3, len(file_data[0])):
                total = sum(float(row[col_idx]) for row in file_data if row[col_idx] not in [None, ''])
                safe_write(footer_row, col_idx + 1, round(total/4000, 2))

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"WeekAhead_{state_name}_{upload_from_date}_Rev{revision_no}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'status': 'failure', 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    



@app.route('/api/uploadmonthahead', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("upload_monthahead", enforce_state=True)
def upload_monthahead_data():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # 1. Get Form Fields
        state = request.form.get('state')
        fromDate = request.form.get('fromDate')
        toDate = request.form.get('toDate')
        role = jwt_data['role']

        # 2. Get JSON Data (Fixed logic for large files)
        # We check for 'dataFile' (file upload) instead of 'data' (form text)
        if 'dataFile' not in request.files:
            return jsonify({"error": "No data provided"}), 400
            
        json_file = request.files['dataFile']
        raw_json = json_file.read().decode("utf-8") # Read file content
        new_data = json.loads(raw_json)             # Parse JSON

        # 3. Get State Name
        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        state_name = cursor.fetchone()[0]

        # 4. Calculate Revision
        cursor.execute("""
            SELECT COUNT(*) FROM month_ahead_file_uploads
            WHERE from_date = to_date(%s,'DD/MM/YYYY')
              AND to_date = to_date(%s,'DD/MM/YYYY')
              AND state_id = %s
        """, (fromDate, toDate, state))
        revision_no = cursor.fetchone()[0]

        # 5. Virtual Filename
        virtual_filename = f"DB_STORED_{fromDate}_{toDate}_{state_name}_rev{revision_no}"

        # 6. Insert into DB
        cursor.execute("""
            INSERT INTO month_ahead_file_uploads
            (state_id, from_date, to_date, upload_time, file_name, revision_no, uploaded_by, file_data)
            VALUES (%s, to_date(%s,'DD/MM/YYYY'), to_date(%s,'DD/MM/YYYY'), NOW(), %s, %s, %s, %s)
        """, (state, fromDate, toDate, virtual_filename, revision_no, role, json.dumps(new_data)))

        conn.commit()

        return jsonify({
            "message": f"Data saved successfully. Revision-{revision_no} created.",
            "status": "success"
        })

    except Exception as e:
        if conn: conn.rollback()
        # log_error("uploadmonthahead", e)
        return jsonify(message=str(e), status="failure")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@app.route('/api/parsemonthaheadexcel', methods=['POST'])
@jwt_required(locations=["cookies"])
def parse_monthahead_excel():
    try:
        # 1. File Validation
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})
        
        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        # 2. Get Date Range from Frontend
        from_date_str = request.form.get('fromDate') # Expected 'DD/MM/YYYY'
        to_date_str = request.form.get('toDate')

        if not from_date_str or not to_date_str:
            return jsonify({'error': 'Date range (fromDate/toDate) missing. Cannot validate month length.'})

        # Parse Dates
        try:
            try:
                start_dt = datetime.strptime(from_date_str, '%d/%m/%Y')
                end_dt = datetime.strptime(to_date_str, '%d/%m/%Y')
            except ValueError:
                # Fallback for YYYY-MM-DD
                start_dt = datetime.strptime(from_date_str, '%Y-%m-%d')
                end_dt = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Invalid Date Format. Expected DD/MM/YYYY.'})

        # 3. Calculate Exact Expected Rows for the Selected Range
        num_days = (end_dt - start_dt).days + 1
        expected_rows = num_days * 96

        # 4. Read Excel
        df = pd.read_excel(file, header=None) 
        all_rows = df.values.tolist()

        # 5. FILTER: Extract only valid data rows (Block is an integer 1-96)
        valid_data_rows = []
        for row in all_rows:
            try:
                # Check column 1 (Block Number)
                if pd.isna(row[1]): continue
                block_val = int(row[1])
                if 1 <= block_val <= 96:
                    valid_data_rows.append(row)
            except (ValueError, TypeError):
                continue # Skip rows where Block is not a number (like headers "Block")

        # 6. STRICT ROW COUNT CHECK
        # Must match exactly. No less, No more.
        if len(valid_data_rows) != expected_rows:
             return jsonify({'error': f"Row Count Mismatch! The selected range ({num_days} days) requires exactly {expected_rows} blocks, but the uploaded file contains {len(valid_data_rows)} valid blocks. Please check your file."})

        # 7. Process Data (Overwrite Dates & Validate)
        # Since count matches, we process all valid rows.
        cleaned_data = []
        blanks_replaced_flag = False

        col_names = {
            3: "Forecasted Demand", 4: "Thermal", 5: "Gas", 6: "Hydro", 7: "Total Own Gen",
            8: "Solar", 9: "Wind", 10: "Other RES", 11: "Total RES", 12: "ISGS/LTA",
            13: "Bilateral", 14: "Total Avail", 15: "Gap", 16: "Bilateral Proc",
            17: "Power Exchange", 18: "Shortage", 19: "Relief", 20: "Load Shedding", 21: "Reactive"
        }

        for row_idx, row in enumerate(valid_data_rows):
            try:
                # --- Auto-Fix Date and Block ---
                day_offset = row_idx // 96
                block_num_calculated = (row_idx % 96) + 1
                
                # Generate correct date based on selected Start Date + Offset
                current_date_obj = start_dt + timedelta(days=day_offset)
                current_date_str = current_date_obj.strftime('%d/%m/%Y')
                
                # Get Time string from file if available (Column C/Index 2)
                time_str = str(row[2]) if len(row) > 2 and not pd.isna(row[2]) else ""

                if 1 <= block_num_calculated <= 96:
                    # Slice to first 22 columns
                    original_slice = row[:22] 
                    
                    # Force the calculated Date and Block into the new row
                    new_row = [current_date_str, block_num_calculated, time_str]

                    # --- Validate Demand (Index 3) ---
                    demand_val = original_slice[3]
                    col_name_demand = col_names.get(3, "Demand")

                    if pd.isna(demand_val) or str(demand_val).strip() == '':
                        return jsonify({'error': f"'{col_name_demand}' is MISSING at Date: {current_date_str}, Block: {block_num_calculated}."})
                    
                    try:
                        demand_float = float(demand_val)
                        if demand_float == 0:
                            return jsonify({'error': f"'{col_name_demand}' cannot be ZERO at Date: {current_date_str}, Block: {block_num_calculated}."})
                        new_row.append(demand_float)
                    except ValueError:
                        return jsonify({'error': f"Invalid numeric value for '{col_name_demand}' at Date: {current_date_str}, Block: {block_num_calculated}."})

                    # --- Validate Other Columns (Index 4 to 21) ---
                    for i in range(4, 22):
                        val = original_slice[i]
                        current_col_name = col_names.get(i, f"Column {i}")

                        if pd.isna(val) or str(val).strip() == '':
                            new_row.append(0.0)
                            blanks_replaced_flag = True
                        else:
                            try:
                                new_row.append(float(val))
                            except ValueError:
                                return jsonify({'error': f"Invalid numeric data in '{current_col_name}' at Date: {current_date_str}, Block: {block_num_calculated}."})
                    
                    cleaned_data.append(new_row)

            except Exception as e:
                return jsonify({'error': f"Row Processing Error at Row {row_idx+1}: {str(e)}"})

        # --- 8. Summation Validation ---
        for row in cleaned_data:
            try:
                date_s, time_s = row[0], row[2]
                thermal, gas, hydro, total_own = row[4], row[5], row[6], row[7]
                
                if abs((thermal + gas + hydro) - total_own) > 0.1:
                     return jsonify({'error': f"Summation Mismatch at {date_s} ({time_s}). Thermal({thermal}) + Gas({gas}) + Hydro({hydro}) != Total({total_own})"})
            except: pass 

        success_msg = 'Parsed successfully. Dates aligned to selected range.'
        if blanks_replaced_flag:
            success_msg += ' (Warning: Blank fields replaced with 0)'

        return jsonify({'message': success_msg, 'data': cleaned_data})

    except Exception as e:
        return jsonify({'error': f"Server Error: {str(e)}"})



@app.route("/api/downloadmonthahead", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_monthahead", enforce_state=True)
def downloadMonthAhead():
    conn = None
    cursor = None
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        data_json = request.get_json()
        state_id = data_json.get('state')
        upload_from_date = data_json.get('from_date')
        upload_to_date = data_json.get('to_date')
        revision_no = data_json.get('revision')

        # Fetch saved data
        query = """
            SELECT t1.file_data, t2.state_name
            FROM public.month_ahead_file_uploads t1
            JOIN public.states t2 ON t1.state_id = t2.state_id
            WHERE t1.state_id = %s 
            AND t1.from_date = %s
            AND t1.to_date = %s 
            AND t1.revision_no = %s
        """
        cursor.execute(query, (state_id, upload_from_date, upload_to_date, revision_no))
        result = cursor.fetchone()

        if not result: return jsonify({'status': 'failure', 'error': 'Data not found'}), 404

        file_data = result[0] if isinstance(result[0], list) else json.loads(result[0])
        state_name = result[1]

        # Load Format File
        format_path = os.path.join(shared_drive_path, "FORMATS", "Month-Ahead Demand Forecast format (from States).xlsx")
        if not os.path.exists(format_path): return jsonify({'status': 'failure', 'error': 'Format file missing'}), 404

        workbook = load_workbook(format_path)
        sheet = workbook.active

        def safe_write(r, c, v):
            cell = sheet.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = v

        # Write Metadata
        safe_write(2, 3, state_name) 
        safe_write(3, 3, f"{upload_from_date} to {upload_to_date}")

        # Write Data
        # 'file_data' is just the data rows (sliced and cleaned during upload)
        start_row = 8
        for row_idx, row_data in enumerate(file_data):
            r = start_row + row_idx
            for col_idx, val in enumerate(row_data):
                safe_write(r, col_idx + 1, val)

        # --- Dynamic Footer Calculation ---
        # The footer is written immediately AFTER the last row of data.
        # This handles the variable length (28/29/30/31 days) correctly.
        footer_row = start_row + len(file_data)
        
        # Write Label "Total MUs" in Column C (Index 3 in Excel 1-based, here column index 3 matches 'Period' col usually, verify your col mapping)
        # Based on WeekAhead/Intraday, usually Period is Col 2 or 3. 
        # In Month Ahead code provided:
        # Col 1: Date, Col 2: Block, Col 3: Period. 
        # Total MUs label usually goes in Period column (Col 3).
        safe_write(footer_row, 3, "Total MUs") 

        # Calculate Total MUs for numeric columns
        # Demand starts at index 3 in our list (Col 4 in Excel)
        if file_data:
            # Iterate columns from index 3 to 21 (end of data)
            num_columns = len(file_data[0])
            for col_idx in range(3, num_columns):
                total = 0
                for row in file_data:
                    try:
                        val = float(row[col_idx])
                    except (ValueError, TypeError):
                        val = 0
                    total += val
                
                # Write calculated sum/4000
                safe_write(footer_row, col_idx + 1, round(total/4000, 2))

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"MonthAhead_{state_name}_{upload_from_date}_Rev{revision_no}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'status': 'failure', 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()



@app.route('/api/fetchmonthrevisions', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_month_revisions")
def fetchMonthRevisions():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        from_date = params["from_date"]
        to_date = params["to_date"]
        state = params["state"]
        cursor.execute("select revision_no from month_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY')".format(params["state"], from_date, to_date))
        revisions_data = cursor.fetchall()
        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for '{0}'-'{1}'".format(from_date, to_date), revisions=revision_list, from_date=from_date, to_date=to_date)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for state {0} for the month '{1}'-'{2}'".format(state_name, from_date, to_date)) 

    except Exception as e:
        log_error("fetchmonthrevisions", e)
        cursor.close()
        return jsonify(message="There is some problem in fetching the revisions! Please contact SRLDC IT", status="failure")
    

@app.route('/api/fetchmonthlyrevisionsdata', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_month_revisions_data")
def fetchMonthlyRevisionsData():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)
        cursor.execute("select file_data, from_date, to_date, upload_time, uploaded_by, revision_no from month_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY') and revision_no={3}".format(params["state"], params["from_date"], params["to_date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        from_date = ''
        to_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            from_date = data[0][1].strftime('%d-%b-%Y')
            to_date = data[0][2].strftime('%d-%b-%Y')
            uploaded_time = data[0][3].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][4]
            revision_no = data[0][5]

            cursor.close()
            # print(len(file_data), len(file_data[0]))
            return jsonify(status="success", data=file_data, time=uploaded_time, from_date=from_date,to_date=to_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact SRLDC IT!")
            

        # return jsonify(message="Fetched Successfully")

    except Exception as e:
        log_error("fetchmonthlyrevisionsdata", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact SRLDC IT", status="failure")


@app.route('/api/parseyearaheadexcel', methods=['POST'])
@jwt_required(locations=["cookies"])
def parse_yearahead_excel():
    try:
        # 1. File Validation
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})
        
        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        # 2. Get Date Range from Frontend
        from_date_str = request.form.get('fromDate') # DD/MM/YYYY
        to_date_str = request.form.get('toDate')

        if not from_date_str or not to_date_str:
            return jsonify({'error': 'Date range (fromDate/toDate) missing. Cannot validate year length.'})

        # Parse Dates
        try:
            try:
                start_dt = datetime.strptime(from_date_str, '%d/%m/%Y')
                end_dt = datetime.strptime(to_date_str, '%d/%m/%Y')
            except ValueError:
                start_dt = datetime.strptime(from_date_str, '%Y-%m-%d')
                end_dt = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Invalid Date Format. Expected DD/MM/YYYY.'})

        # 3. Calculate Expected Rows (Days * 24 hours)
        num_days = (end_dt - start_dt).days + 1
        expected_rows = num_days * 24

        # 4. Read Excel
        df = pd.read_excel(file, header=None) 
        all_rows = df.values.tolist()

        # 5. Filter for Valid Data Rows (Hour 1-24)
        valid_data_rows = []
        for row in all_rows:
            try:
                if pd.isna(row[1]): continue
                hour_val = int(row[1])
                if 1 <= hour_val <= 24:
                    valid_data_rows.append(row)
            except (ValueError, TypeError):
                continue

        # 6. Row Count Check
        if len(valid_data_rows) != expected_rows:
             return jsonify({'error': f"Row Count Mismatch! The selected year ({num_days} days) requires exactly {expected_rows} rows, but the file contains {len(valid_data_rows)} valid rows."})

        # 7. Process Data (Overwrite Dates)
        cleaned_data = []
        blanks_replaced_flag = False

        # Column Mapping (Indices 0-20, based on 21 cols in Year Ahead)
        # 0:Date, 1:Hour, 2:Demand...
        col_names = {
            2: "Forecasted Demand", 3: "Thermal", 4: "Gas", 5: "Hydro", 6: "Total Own Gen",
            7: "Solar", 8: "Wind", 9: "Other RES", 10: "Total RES", 11: "ISGS/LTA",
            12: "Bilateral", 13: "Total Avail", 14: "Gap", 15: "Bilateral Proc",
            16: "Power Exchange", 17: "Shortage", 18: "Relief", 19: "Load Shedding", 20: "Reactive"
        }

        for row_idx, row in enumerate(valid_data_rows):
            try:
                # Calculate Date and Hour
                day_offset = row_idx // 24
                hour_val = (row_idx % 24) + 1
                
                current_date_obj = start_dt + timedelta(days=day_offset)
                current_date_str = current_date_obj.strftime('%d/%m/%Y')

                if 1 <= hour_val <= 24:
                    # Slice to first 21 columns (Year Ahead usually has fewer cols/no Period col compared to Month Ahead)
                    # Verify column count in your specific template. Assuming 21 cols based on your TS code.
                    original_slice = row[:21] 
                    
                    # Force Date and Hour
                    # New Row Structure: [Date, Hour, Demand, ...]
                    new_row = [current_date_str, hour_val] 

                    # --- Validate Demand (Index 2) ---
                    demand_val = original_slice[2]
                    col_name_demand = col_names.get(2, "Demand")

                    if pd.isna(demand_val) or str(demand_val).strip() == '':
                        return jsonify({'error': f"'{col_name_demand}' is MISSING at Date: {current_date_str}, Hour: {hour_val}."})
                    
                    try:
                        demand_float = float(demand_val)
                        if demand_float == 0:
                            return jsonify({'error': f"'{col_name_demand}' cannot be ZERO at Date: {current_date_str}, Hour: {hour_val}."})
                        new_row.append(demand_float)
                    except ValueError:
                        return jsonify({'error': f"Invalid numeric value for '{col_name_demand}' at Date: {current_date_str}, Hour: {hour_val}."})

                    # --- Validate Other Columns (Index 3 to 20) ---
                    for i in range(3, 21):
                        val = original_slice[i]
                        current_col_name = col_names.get(i, f"Column {i}")

                        if pd.isna(val) or str(val).strip() == '':
                            new_row.append(0.0)
                            blanks_replaced_flag = True
                        else:
                            try:
                                new_row.append(float(val))
                            except ValueError:
                                return jsonify({'error': f"Invalid numeric data in '{current_col_name}' at Date: {current_date_str}, Hour: {hour_val}."})
                    
                    cleaned_data.append(new_row)

            except Exception as e:
                return jsonify({'error': f"Row Processing Error at Row {row_idx+1}: {str(e)}"})

        # --- Summation Validation ---
        for row in cleaned_data:
            try:
                date_s, hour_s = row[0], row[1]
                # Indices: 0:Date, 1:Hour, 2:Demand, 3:Thermal, 4:Gas, 5:Hydro, 6:Total(B)
                thermal, gas, hydro, total_own = row[3], row[4], row[5], row[6]
                
                if abs((thermal + gas + hydro) - total_own) > 0.1:
                     return jsonify({'error': f"Summation Mismatch at {date_s} (Hour {hour_s}). Thermal({thermal}) + Gas({gas}) + Hydro({hydro}) != Total({total_own})"})
            except: pass 

        success_msg = 'Parsed successfully. Dates aligned to selected range.'
        if blanks_replaced_flag:
            success_msg += ' (Warning: Blank fields replaced with 0)'

        return jsonify({'message': success_msg, 'data': cleaned_data})

    except Exception as e:
        return jsonify({'error': f"Server Error: {str(e)}"})


@app.route('/api/uploadyearahead', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("upload_yearahead", enforce_state=True)
def uploadYearAheadDataAndFile():
    conn = None
    cursor = None
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        state = request.form.get('state')
        fromDate = request.form.get('fromDate')
        toDate = request.form.get('toDate')
        role = jwt_data['role']

        # --- FIX: Read JSON from File ---
        if 'dataFile' not in request.files:
            return jsonify({"error": "No data provided"}), 400
            
        json_file = request.files['dataFile']
        raw_json = json_file.read().decode("utf-8")
        new_data = json.loads(raw_json)
        # -------------------------------

        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        state_name = cursor.fetchone()[0]

        # Calculate Revision
        cursor.execute("""
            SELECT COUNT(*) FROM year_ahead_file_uploads
            WHERE from_date = to_date(%s,'DD/MM/YYYY')
              AND to_date = to_date(%s,'DD/MM/YYYY')
              AND state_id = %s
        """, (fromDate, toDate, state))
        revision_no = cursor.fetchone()[0]

        # Virtual Filename
        virtual_filename = f"DB_STORED_{fromDate}_{toDate}_{state_name}_rev{revision_no}"

        cursor.execute("""
            INSERT INTO year_ahead_file_uploads
            (state_id, from_date, to_date, upload_time, file_name, revision_no, uploaded_by, file_data)
            VALUES (%s, to_date(%s,'DD/MM/YYYY'), to_date(%s,'DD/MM/YYYY'), NOW(), %s, %s, %s, %s)
        """, (state, fromDate, toDate, virtual_filename, revision_no, role, json.dumps(new_data)))

        conn.commit()

        return jsonify({
            "message": f"Data saved successfully. Revision-{revision_no} created.",
            "status": "success"
        })

    except Exception as e:
        if conn: conn.rollback()
        return jsonify(message=str(e), status="failure")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@app.route("/api/downloadyearahead", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_yearahead", enforce_state=True)
def downloadYearAhead():
    conn = None
    cursor = None
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        data_json = request.get_json()
        state_id = data_json.get('state')
        upload_from_date = data_json.get('from_date')
        upload_to_date = data_json.get('to_date')
        revision_no = data_json.get('revision')

        # Fetch Data
        query = """
            SELECT t1.file_data, t2.state_name
            FROM public.year_ahead_file_uploads t1
            JOIN public.states t2 ON t1.state_id = t2.state_id
            WHERE t1.state_id = %s 
            AND t1.from_date = %s
            AND t1.to_date = %s 
            AND t1.revision_no = %s
        """
        cursor.execute(query, (state_id, upload_from_date, upload_to_date, revision_no))
        result = cursor.fetchone()

        if not result: 
            return jsonify({'status': 'failure', 'error': 'Data not found'}), 404

        file_data = result[0] if isinstance(result[0], list) else json.loads(result[0])
        state_name = result[1]

        # --- FIX: Correct Format File Name ---
        format_path = os.path.join(shared_drive_path, "FORMATS", "Year-Ahead Demand Forecast format (from States).xlsx")
        
        if not os.path.exists(format_path): 
            return jsonify({'status': 'failure', 'error': 'Format file missing on server'}), 404

        workbook = load_workbook(format_path)
        sheet = workbook.active

        def safe_write(r, c, v):
            cell = sheet.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): 
                cell.value = v

        # Write Metadata
        safe_write(2, 3, state_name) 
        safe_write(3, 3, f"{upload_from_date} to {upload_to_date}")

        # Write Data Rows
        # start_row = 8 means Excel Row 8 (1-based index)
        start_row = 8
        for row_idx, row_data in enumerate(file_data):
            r = start_row + row_idx
            for col_idx, val in enumerate(row_data):
                # col_idx + 1 because Excel is 1-based, data list is 0-based
                safe_write(r, col_idx + 1, val)

        # --- Dynamic Footer Calculation ---
        # Footer goes immediately after the last data row
        footer_row = start_row + len(file_data)
        
        # Write "Total MUs" label in Column 2 (Index 1) -> Excel Column 'B' (Hour) or 'C' depending on merge
        # In Year Ahead format, Col 1=Date, Col 2=Hour.
        safe_write(footer_row, 2, "Total MUs") 

        # Calculate Totals for Numeric Columns
        # In Year Ahead data list: 0=Date, 1=Hour, 2=Demand, 3=Thermal ...
        if file_data:
            num_columns = len(file_data[0])
            # Start summing from index 2 (Demand) up to the last column
            for col_idx in range(2, num_columns):
                total = 0
                for row in file_data:
                    try:
                        val = float(row[col_idx])
                    except (ValueError, TypeError): 
                        val = 0
                    total += val
                
                # Write Sum/4000
                safe_write(footer_row, col_idx + 1, round(total/4000, 2))

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"YearAhead_{state_name}_{upload_from_date}_Rev{revision_no}.xlsx"
        return send_file(
            output, 
            as_attachment=True, 
            download_name=filename, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'status': 'failure', 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/api/fetchyearlyrevisionsdata', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_year_revisions_data")
def fetchYearlyRevisionsData():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)
        cursor.execute("select file_data, from_date, to_date, upload_time, uploaded_by, revision_no from year_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY') and revision_no={3}".format(params["state"], params["from_date"], params["to_date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        from_date = ''
        to_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            from_date = data[0][1].strftime('%d-%b-%Y')
            to_date = data[0][2].strftime('%d-%b-%Y')
            uploaded_time = data[0][3].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][4]
            revision_no = data[0][5]
            cursor.close()
            # print(len(file_data), len(file_data[0]))
            # cursor.c
            return jsonify(status="success", data=file_data, time=uploaded_time, from_date=from_date,to_date=to_date, revision=revision_no, role=uploaded_by)
            
        else:
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact SRLDC IT!")
            

        return jsonify(message="Fetched Successfully")
    
    except Exception as e:
        log_error("fetchyearlyrevisionsdata", e)
        cursor.close()
        return jsonify(message="There is some problem in fetching the revisions data! Please contact SRLDC IT", status="failure")



@app.route('/api/fetchyearrevisions', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_year_revisions")
def fetchYearRevisions():
    try:
        jwt_data = get_jwt()
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        from_date = params["from_date"]
        to_date = params["to_date"]
        state = params["state"]
        cursor.execute("select revision_no from year_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY')".format(params["state"], from_date, to_date))
        revisions_data = cursor.fetchall()
        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for '{0}'-'{1}'".format(from_date, to_date), revisions=revision_list, from_date=from_date, to_date=to_date)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for state {0} for the year '{1}'-'{2}'".format(state_name, from_date, to_date))    
    
    except Exception as e:
        log_error("fetchyearrevisions", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact SRLDC IT", status="failure")






@app.route('/api/uploadstatus')
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_upload_status")
def scatterPlotUploadStatus():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Day Ahead Data Processing
        header_data = dict(request.headers)
        state = request.form.get('state')
        # print(header_data)

        # token = header_data['Authorization'].split()[1]
        # x = decode_token(token, csrf_value=None, allow_expired=False)


        # print("Cookies Received:", request.cookies)  # 🔥 Debug cookies
        jwt_data = get_jwt()  #  Extract JWT
        # print("JWT Payload:", jwt_data)  #  Debug JWT payload




        state_id_query = """select state_id from public.states where username = %s"""
        cursor.execute(state_id_query, (jwt_data['sub'],))
        state_id = cursor.fetchall()[0][0]

        # print(state)
        end_date = datetime.now() + timedelta(days=1)
        start_date = end_date - timedelta(days=30)

        sql_query = """
                    WITH min_revision AS (
                    SELECT 
                        state_id, 
                        upload_date, 
                        MIN(revision_no) AS min_revision_no
                    FROM 
                        file_uploads
                    GROUP BY 
                        state_id, 
                        upload_date
                )
                SELECT 
                    states.state_name,
                    COALESCE(file_uploads.upload_date, %s) AS upload_date,
                    COALESCE(min_uploads.upload_time, NULL) AS upload_time,
                    CASE
                        WHEN min_uploads.upload_time IS NULL THEN 2  -- Not Uploaded
                        WHEN min_uploads.upload_time < (file_uploads.upload_date - INTERVAL '1 day' + INTERVAL '10 hours') THEN 1  -- Proper Upload before 10 AM on the previous day
                        ELSE 0  -- Late Upload
                    END AS upload_status_code,
                    COUNT(file_uploads.state_id) AS upload_count
                FROM 
                    states
                LEFT JOIN 
                    file_uploads ON states.state_id = file_uploads.state_id
                    AND file_uploads.upload_date BETWEEN %s AND %s
                LEFT JOIN (
                    SELECT 
                        file_uploads.state_id, 
                        file_uploads.upload_date, 
                        file_uploads.upload_time
                    FROM 
                        file_uploads
                    INNER JOIN min_revision ON 
                        file_uploads.state_id = min_revision.state_id
                        AND file_uploads.upload_date = min_revision.upload_date
                        AND file_uploads.revision_no = min_revision.min_revision_no
                ) AS min_uploads ON 
                    file_uploads.state_id = min_uploads.state_id 
                    AND file_uploads.upload_date = min_uploads.upload_date
                WHERE 
                    states.state_id IN (1,2,3,4,5,7)
                GROUP BY 
                    states.state_name, 
                    file_uploads.upload_date,
                    min_uploads.upload_time  
                ORDER BY
                    states.state_name,
                    upload_date DESC;
        """
        cursor.execute(sql_query, (end_date, start_date, end_date))
        results = cursor.fetchall()

        day_data = []
        date_range = [end_date - timedelta(days=i) for i in range(30)]
        state_names = set(result[0] for result in results)

        for state_name in state_names:
            state_data = {"name": state_name, "data": []}
            for date in date_range:
                date_str = date.strftime('%Y-%m-%d')
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1].strftime("%Y-%m-%d") == date_str:
                        upload_count = result[4]
                        upload_time = result[2].strftime("%Y-%m-%d %H:%M:%S") if result[2] is not None else None
                        upload_status_code = result[3]
                        found_result = True
                        break
                if not found_result:
                    upload_count = 0
                    upload_time = None
                    upload_status_code = 2  # Not Uploaded
                state_data["data"].append({
                    'x': date_str, 
                    'y': upload_status_code, 
                    'upload_time': upload_time, 
                    'upload_count': upload_count
                })
            day_data.append(state_data)

        day_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }

        mape_comp_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%d/%m/%Y'),
            "end_date": end_date.strftime('%d/%m/%Y')
        }


        # for i in day_data:
        #     print(i["name"])

        #######################################################################################################################################
        #######################################################################################################################################
        # Week Ahead Data Processing
                


        # Week Ahead Data Processing
        # Calculate date ranges
        today = datetime.now()
        # Get the start of this week (Monday)
        this_week_start = today - timedelta(days=today.weekday())

        # Setting date ranges
        start_date = this_week_start - timedelta(weeks=2)  # Start date two weeks before this week
        end_date = this_week_start + timedelta(weeks=3) - timedelta(days=1)  # End date two weeks after this week

        # Adjusting end_date to ensure it ends on the last day of the intended week
        if end_date.weekday() != 6:  # Check if end_date is not a Sunday
            end_date = end_date - timedelta(days=end_date.weekday() + 1)  # Adjust to the last Sunday before or on end_date

        sql_query = """
            WITH min_revision AS (
                SELECT 
                    state_id, 
                    from_date, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    week_ahead_file_uploads
                WHERE
                    from_date BETWEEN %s AND %s
                GROUP BY 
                    state_id, 
                    from_date
            ), min_uploads AS (
                SELECT 
                    fu.state_id, 
                    fu.from_date, 
                    fu.upload_time,
                    fu.revision_no
                FROM 
                    week_ahead_file_uploads fu
                INNER JOIN min_revision mr ON
                    fu.state_id = mr.state_id
                    AND fu.from_date = mr.from_date
                    AND fu.revision_no = mr.min_revision_no
            )
            SELECT 
                states.state_name,
                COALESCE(mu.from_date, %s) AS week_start_date,
                COALESCE(mu.upload_time, NULL) AS upload_time,
                CASE
                    WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                    WHEN mu.upload_time < DATE_TRUNC('week', mu.from_date - INTERVAL '1 week')  + INTERVAL '1 day' THEN 1  -- Uploaded on time, before the first working day of the week
                    ELSE 0  -- Late Upload
                END AS upload_status_code,
                COUNT(mu.state_id) AS upload_count
            FROM 
                states
            LEFT JOIN 
                min_uploads mu ON states.state_id = mu.state_id
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                mu.from_date,
                mu.upload_time  
            ORDER BY
                states.state_name,
                mu.from_date DESC;
        """
        cursor.execute(sql_query, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        results = cursor.fetchall()


        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        week_range = sorted([(start_date + timedelta(weeks=i)).date() for i in range(5)], reverse=True)  # Generate Mondays for 6 weeks

        week_data = []
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for week_start in week_range:
                week_start_date = week_start
                week_end_date = week_start_date + timedelta(days=6)  # End on Sunday
                week_range_str = f"{week_start_date.strftime('%Y-%m-%d')} to {week_end_date.strftime('%Y-%m-%d')}"

                upload_count = 0
                upload_status_code = 2  # Default to Not Uploaded
                upload_time = None
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1] == week_start_date:
                        upload_count = result[4]
                        upload_time = result[2].strftime("%Y-%m-%d %H:%M:%S") if result[2] is not None else None
                        upload_status_code = result[3]
                        found_result = True
                        break

                if not found_result:
                    upload_count = 0
                    upload_status_code = 2  # Not Uploaded
                state_data["data"].append({'x': week_range_str, 'y': upload_status_code, 'upload_time': upload_time, 'upload_count': upload_count})
            week_data.append(state_data)

        week_dates = {
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }






        #######################################################################################################################################
        #######################################################################################################################################
        # Month Ahead Data Processing
        # Initialize current date and calculate relevant month starts
        today = datetime.now()
        current_month_start = today.replace(day=1)
        previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        next_month_start = (current_month_start + timedelta(days=31)).replace(day=1)

        # Define the date ranges
        start_date = previous_month_start
        end_date = next_month_start.replace(day=monthrange(next_month_start.year, next_month_start.month)[1])

        # Prepare the SQL query to fetch the data
        sql_query = """
        WITH min_revision AS (
                SELECT 
                    state_id, 
                    DATE_TRUNC('month', from_date) AS from_month, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    month_ahead_file_uploads
                WHERE
                    from_date BETWEEN %s AND %s
                GROUP BY 
                    state_id, 
                    from_month
            ), min_uploads AS (
                SELECT 
                    fu.state_id, 
                    DATE_TRUNC('month', fu.from_date) AS from_month, 
                    fu.upload_time,
                    fu.revision_no
                FROM 
                    month_ahead_file_uploads fu
                INNER JOIN min_revision mr ON
                    fu.state_id = mr.state_id
                    AND DATE_TRUNC('month', fu.from_date) = mr.from_month
                    AND fu.revision_no = mr.min_revision_no
            )
            SELECT 
                states.state_name,
                COALESCE(mu.from_month, %s) AS month_start_date,
                COALESCE(mu.upload_time, NULL) AS upload_time,
                CASE
                    WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                    WHEN mu.upload_time < DATE_TRUNC('month', mu.from_month) - INTERVAL '1 month' + INTERVAL '5 day' THEN 1  -- Uploaded on time
                    ELSE 0  -- Late Upload
                END AS upload_status_code,
                COUNT(mu.state_id) AS upload_count
            FROM 
                states
            LEFT JOIN 
                min_uploads mu ON states.state_id = mu.state_id
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                mu.from_month,
                mu.upload_time  
            ORDER BY
                states.state_name,
                mu.from_month DESC;
        """
        cursor.execute(sql_query, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Fetch the state names from the database
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Initialize month range and data structure
        month_range = [next_month_start, current_month_start, previous_month_start]
        month_data = []

        # Process results and sort by latest month first
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for month_start in sorted(month_range, reverse=True):
                _, last_day_of_month = monthrange(month_start.year, month_start.month)
                month_end = month_start.replace(day=last_day_of_month)
                month_range_str = f"{month_start.strftime('%Y-%m-%d')} to {month_end.strftime('%Y-%m-%d')}"

                upload_count = 0
                upload_status_code = 2  # Default to Not Uploaded
                upload_time = None
                found_result = False

                # Check each result for the current state and month
                for result in results:
                    result_state, result_month_start, result_upload_time, result_status, result_count = result
                    if result_state == state_name and result_month_start.strftime('%Y-%m-%d') == month_start.strftime('%Y-%m-%d'):
                        upload_count = result_count
                        upload_time = result_upload_time.strftime("%Y-%m-%d %H:%M:%S") if result_upload_time else None
                        upload_status_code = result_status
                        found_result = True
                        break

                if not found_result:
                    upload_count = 0
                    upload_status_code = 2  # Not Uploaded

                state_data["data"].append({
                    'x': month_range_str,
                    'y': upload_status_code,
                    'upload_time': upload_time,
                    'upload_count': upload_count
                })

            month_data.append(state_data)

        # Prepare JSON for frontend
        month_dates = {
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }


        ############### Year Ahead Forecast Status

        # Financial year calculation
        
        today = datetime.now()

        # Determine the current and next financial years
        if today.month < 4:
            current_financial_year_start = datetime(today.year - 1, 4, 1)
        else:
            current_financial_year_start = datetime(today.year, 4, 1)
        current_financial_year_end = datetime(current_financial_year_start.year + 1, 3, 31)
        next_financial_year_start = datetime(current_financial_year_start.year + 1, 4, 1)
        next_financial_year_end = datetime(next_financial_year_start.year + 1, 3, 31)

        # SQL query (no CASE, just fetch data)
        sql_query = """
        SELECT 
            states.state_name,
            yafu.from_date,
            yafu.to_date,
            yafu.upload_time,
            COUNT(yafu.state_id) AS upload_count
        FROM 
            states
        LEFT JOIN 
            year_ahead_file_uploads yafu ON states.state_id = yafu.state_id
        WHERE 
            states.state_id IN (1,2,3,4,5,7)
            AND (yafu.from_date BETWEEN %s AND %s OR yafu.from_date IS NULL)
        GROUP BY 
            states.state_name, yafu.from_date, yafu.to_date, yafu.upload_time
        ORDER BY
            states.state_name, yafu.from_date DESC;
        """
        cursor.execute(
            sql_query,
            (current_financial_year_start.strftime('%Y-%m-%d'),
            next_financial_year_end.strftime('%Y-%m-%d'))
        )
        results = cursor.fetchall()

        # Fetch state names to ensure coverage of all states
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Organize the fetched data by state and financial year
        year_data = []
        financial_years = [current_financial_year_start, next_financial_year_start]
        from datetime import date
        from dateutil.relativedelta import relativedelta
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for financial_year in financial_years:
                found = False
                for result in results:
                    result_state, from_date, to_date, upload_time, upload_count = result

                    # normalize dates
                    if isinstance(from_date, date) and not isinstance(from_date, datetime):
                        from_date = datetime.combine(from_date, datetime.min.time())
                    if isinstance(to_date, date) and not isinstance(to_date, datetime):
                        to_date = datetime.combine(to_date, datetime.min.time())

                    if result_state == state_name and from_date and from_date.date() == financial_year.date():
                        found = True

                        # === Apply business rules ===
                        if not upload_time:
                            upload_status = 2  # Not uploaded
                        else:
                            sept30_preceding = datetime(from_date.year - 1, 9, 30)
                            if upload_time > sept30_preceding:
                                upload_status = 0  # Late (after Sept 30 of preceding year)
                            elif upload_time <= from_date + relativedelta(months=+5, days=-1):
                                upload_status = 1  # On time (within 5 months)
                            else:
                                upload_status = 0  # Late

                        upload_time_formatted = upload_time.strftime("%Y-%m-%d %H:%M:%S") if upload_time else None
                        state_data["data"].append({
                            'x': f"{from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')}",
                            'y': upload_status,
                            'upload_time': upload_time_formatted,
                            'upload_count': upload_count
                        })
                        break
                if not found:
                    state_data["data"].append({
                        'x': f"{financial_year.strftime('%Y-%m-%d')} to {financial_year.replace(year=financial_year.year + 1, month=3, day=31).strftime('%Y-%m-%d')}",
                        'y': 2,  # Not Uploaded
                        'upload_time': None,
                        'upload_count': 0
                    })
            year_data.append(state_data)

        # Prepare JSON data for the frontend
        year_dates = {
            "start_date": current_financial_year_start.strftime('%Y-%m-%d'),
            "end_date": next_financial_year_end.strftime('%Y-%m-%d')
        }



        # url = host_url +"/mapechart"
        # headers = {
        #     'Authorization': request.headers.get('Authorization'),
        #     'Content-Type': 'application/json'
        # }
        # payload = {
        #     "params": {
        #         "fromDate": from_date,
        #         "toDate": to_date,
        #         "state": state_id
        #     }
        # }

        # print(header_data)

        # token = header_data['Authorization'].split()[1]
        # x = decode_token(token, csrf_value=None, allow_expired=False)

        # username = jwt_data['sub']
        # # print(username, "username")
        # # role = jwt_data['role']

        # # print(state_dict)

        # if username in state_dict.keys():
        #     url = host_url +"/mapechart"
        #     headers = {
        #         'Authorization': request.headers.get('Authorization'),
        #         'Content-Type': 'application/json'
        #     }
        #     payload = {
        #         "params": {
        #             "fromDate": mape_comp_dates['start_date'],
        #             "toDate": mape_comp_dates['end_date'],
        #             "state": state_dict[username]
        #         }
        #     }



        #     response = requests.post(url, json=payload, headers=headers)
        #     cursor.close()

        #     if response.status_code == 200:
        #         # pass
        #         # print(response.json().keys())
        #         res_data = response.json()
        #         print(res_data.keys())
        #         print(res_data['title'])
        #         if res_data['status'] == 'success':    
        #             return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates, mape_data= res_data['data'],mape_title = res_data['title'],comp_data = res_data['comp_data'],  status="success")
        #         else:
        #             return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates,  status="success")

        #     else:
        #         print("Response not recieved")
        #         return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates,  status="success")


        ################### Intraday Processing

        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        sql_query = """
                                WITH min_revision AS (
                SELECT 
                    state_id, 
                    upload_date, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    intraday_file_uploads
                GROUP BY 
                    state_id, 
                    upload_date
            ),
            file_count AS (
                SELECT 
                    state_id, 
                    upload_date, 
                    COUNT(revision_no) AS revision_count,
                    MAX(CASE WHEN file_type = 'D' THEN 1 ELSE 0 END) AS has_file_type_D
                FROM 
                    intraday_file_uploads
                GROUP BY 
                    state_id, 
                    upload_date
            )
            SELECT 
                states.state_name,
                COALESCE(intraday_file_uploads.upload_date, %s) AS upload_date,
                CASE
                    WHEN file_count.revision_count = 1 AND file_count.has_file_type_D = 1 THEN 0  -- Single revision and file_type is 'D' → Not Uploaded
                    WHEN MIN(intraday_file_uploads.revision_no) IS NULL THEN 0  -- No Uploads → Not Uploaded
                    ELSE 1  -- Uploaded
                END AS upload_status_code
            FROM 
                states
            LEFT JOIN 
                intraday_file_uploads ON states.state_id = intraday_file_uploads.state_id
                AND intraday_file_uploads.upload_date BETWEEN %s AND %s
            LEFT JOIN 
                file_count ON states.state_id = file_count.state_id 
                AND intraday_file_uploads.upload_date = file_count.upload_date
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                intraday_file_uploads.upload_date, 
                file_count.revision_count, 
                file_count.has_file_type_D
            ORDER BY
                states.state_name,
                upload_date DESC;

        """
        cursor.execute(sql_query, (end_date, start_date, end_date))
        results = cursor.fetchall()

        intraday_data = []
        date_range = [end_date - timedelta(days=i) for i in range(30)]
        state_names = set(result[0] for result in results)

        for state_name in state_names:
            state_data = {"name": state_name, "data": []}
            for date in date_range:
                date_str = date.strftime('%Y-%m-%d')
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1].strftime("%Y-%m-%d") == date_str:
                        upload_status_code = result[2]
                        found_result = True
                        break
                if not found_result:
                    upload_status_code = 0  # Not Uploaded
                state_data["data"].append({
                    'x': date_str, 
                    'y': upload_status_code
                })
            intraday_data.append(state_data)

        intraday_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }

        ########## NLDC Mail STatius

        # from datetime import datetime, timedelta

        # Set end and start dates

        mail_status_data = []
        end_date = datetime.now().date() + timedelta(days=1)
        start_date = end_date - timedelta(days=30)

        # Query for forecast_mail_status
        # Query to get all sent entries within range
        sql_query = """
            SELECT forecast_type, from_date, to_date
            FROM forecast_mail_status
            WHERE forecast_type IN ('day', 'week', 'month')
              AND from_date <= %s
              AND to_date >= %s
        """
        cursor.execute(sql_query, (end_date, start_date))
        results = cursor.fetchall()  # [(forecast_type, from_date, to_date), ...]

        # Create date sets for each type
        sent_dates = {
            'day': set(),
            'week': set(),
            'month': set()
        }

        for f_type, from_date, to_date in results:
            for d in range((to_date - from_date).days + 1):
                sent_dates[f_type].add(from_date + timedelta(days=d))

        # Generate date range
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        mail_status_data = []
        for f_type in ['day', 'week', 'month']:
            type_data = {"name": f_type, "data": []}
            for date in date_range:
                status = 1 if date in sent_dates[f_type] else 0
                type_data["data"].append({
                    "x": date.strftime('%Y-%m-%d'),
                    "y": status
                })
            mail_status_data.append(type_data)

        # Date range for frontend
        mail_status_dates = {
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }

        # Output:
        # forecast_data → used in heatmap
        # forecast_dates → used for showing selected period







        


        return jsonify(day=day_data, week=week_data, month=month_data, year = year_data,intraday = intraday_data,intraday_dates = intraday_dates, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates, mail_status_dates=mail_status_dates, mail_status=mail_status_data, status="success")
    
    except Exception as e:
        # print(error)

        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print(exc_type, fname, exc_tb.tb_lineno)
        log_error("uploadstatus", e)
        # cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact SRLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact SRLDC IT!", status="failure")











@app.route('/api/mapechart', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_upload_status", enforce_state=True)
def mapeChart():
    try:
        conn = psycopg2.connect(
        database="demand_forecast_states", user='prasadbabu', 
        password='BabuPrasad#123', host='10.0.100.79', port='5432'
        )
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)

        jwt_data = get_jwt()

        from_date = params["params"]["fromDate"]
        to_date = params["params"]["toDate"]  
        state_id = int(params["params"]["state"])

        # print(state_id, type(state_id))

        cursor.execute("select state_name from states where state_id = {0}".format(state_id))
        state_name = cursor.fetchall()[0][0]

        region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
        selected_state_id = state_id  # Set to the desired single state or region (use 6 for the entire region)

        # Define expected state IDs based on selected_state_id
        expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

        date_range = pd.date_range(start=datetime.strptime(from_date, "%d/%m/%Y"), end=datetime.strptime(to_date, "%d/%m/%Y"))
        date_df = pd.DataFrame(date_range, columns=["D_F_DATE"])



        # Fetch actual data and filter by expected state IDs
        cursor.execute('''
            select s.state_id, process_date, demand_met, scada_demand 
            from states s 
            join actual_demand ad on s.state_id = ad.state_id 
            where process_date between to_date('{0}', 'DD/MM/YYYY') and to_date('{1}', 'DD/MM/YYYY')
            order by process_date
        '''.format(from_date, to_date))

        sr_actual_data_result = cursor.fetchall()

        # Aggregate actual data by date for selected state or region
        daily_totals = defaultdict(lambda: {"DAY_ACTUAL": [0] * 96, "SCADA_DEMAND": [0] * 96, "STATE_IDS": set()})
        for record in sr_actual_data_result:
            state_id, process_date, demand_met_list, scada_demand_list = record
            if state_id not in expected_state_ids:  # Filter out any unexpected states
                continue
            date_str = process_date.strftime("%Y-%m-%d")

            if demand_met_list is None:
                demand_met_list = [0] * 96
            if scada_demand_list is None:
                scada_demand_list = [0] * 96
            
            daily_totals[date_str]["STATE_IDS"].add(state_id)
            
            if (not demand_met_list or all(x == 0 for x in demand_met_list)):
                demand_met_list = scada_demand_list

            daily_totals[date_str]["DAY_ACTUAL"] = [x + y for x, y in zip(daily_totals[date_str]["DAY_ACTUAL"], demand_met_list)]
            daily_totals[date_str]["SCADA_DEMAND"] = [x + y for x, y in zip(daily_totals[date_str]["SCADA_DEMAND"], scada_demand_list)]

        sr_day_actual_json_list = [
            {"A_DATE": date, "DAY_ACTUAL": totals["DAY_ACTUAL"], "SCADA_DEMAND": totals["SCADA_DEMAND"], "STATE_IDS": list(totals["STATE_IDS"])}
            for date, totals in daily_totals.items()
        ]
        sr_actual_day_df = pd.DataFrame(sr_day_actual_json_list)

        # print("mape api")
        # print(sr_actual_day_df)
        # print(sr_actual_day_df.dtypes)


        # print(sr_actual_day_df['A_DATE'].unique())

        def intraday_mape(from_date, to_date, sr_actual_day_df):
            cursor.execute('''
                WITH RECURSIVE DateRange AS (
                    SELECT to_date('{0}', 'DD/MM/YYYY')::date AS upload_date
                    UNION ALL
                    SELECT (upload_date + INTERVAL '1 day')::date
                    FROM DateRange
                    WHERE upload_date + INTERVAL '1 day' <= to_date('{1}', 'DD/MM/YYYY')
                ),
                MaxRevisions AS (
                    SELECT state_id, upload_date, MAX(revision_no) AS max_revision
                    FROM intraday_file_uploads
                    WHERE upload_date BETWEEN to_date('{0}', 'DD/MM/YYYY') AND to_date('{1}', 'DD/MM/YYYY')
                    AND file_type = 'I'
                    GROUP BY state_id, upload_date
                )
                SELECT dr.upload_date, t.state_id, t.revision_no, t.file_data
                FROM DateRange dr
                LEFT JOIN intraday_file_uploads t
                    ON dr.upload_date = t.upload_date
                    AND t.revision_no = (
                        SELECT max_revision
                        FROM MaxRevisions mr
                        WHERE mr.state_id = t.state_id AND mr.upload_date = t.upload_date
                    )
                    AND t.file_type = 'I'
                ORDER BY dr.upload_date, t.state_id;
            '''.format(from_date, to_date))

            sr_day_data = cursor.fetchall()

            # Aggregate forecast data by date for selected state or region
            sr_day_forecast_dict = defaultdict(lambda: {"INTRADAY_FORECAST": [0] * 96, "STATE_IDS": set()})
            for record in sr_day_data:
                upload_date, state_id, revision_no, file_data = record
                if state_id not in expected_state_ids:
                    continue
                date_str = upload_date.strftime("%Y-%m-%d")
                
                if file_data:
                    day_forecast_values = [forecast[2] for forecast in file_data]
                    sr_day_forecast_dict[date_str]["STATE_IDS"].add(state_id)
                    sr_day_forecast_dict[date_str]["INTRADAY_FORECAST"] = [x + y for x, y in zip(sr_day_forecast_dict[date_str]["INTRADAY_FORECAST"], day_forecast_values)]

            sr_day_forecast_json_list = [
                {"D_F_DATE": date, "INTRADAY_FORECAST": data["INTRADAY_FORECAST"], "STATE_IDS": list(data["STATE_IDS"])}
                for date, data in sr_day_forecast_dict.items()
            ]
            # Create the DataFrame
            if sr_day_forecast_json_list:  # If the list is not empty
                sr_forecast_day_df = pd.DataFrame(sr_day_forecast_json_list)
            else:  # If the list is empty, create an empty DataFrame with required columns
                sr_forecast_day_df = pd.DataFrame(columns=["D_F_DATE", "INTRADAY_FORECAST", "STATE_IDS"])

            formatted_forecast_data, formatted_actual_data = [], []
            sr_mape_dict = {"name": "Intraday", "data": []}

            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            if "D_F_DATE" in sr_forecast_day_df.columns:
                sr_forecast_day_df["D_F_DATE"] = pd.to_datetime(sr_forecast_day_df["D_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_forecast_day_df, on="D_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            for _, row in sr_actual_day_df.iterrows():
                day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                base_date = row['A_DATE']
                for i, actual_value in enumerate(day_actual_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(actual_value)
                        })




            for _, row in sr_result_df.iterrows():
                temp = {"x": fmt_date(row["A_DATE"])}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()
                forecast_day_values = row["INTRADAY_FORECAST"]
                if not isinstance(forecast_day_values, list):
                    forecast_day_values = [0] * 96

                if forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                elif (len(row["INTRADAY_FORECAST"]) != 96 or pd.isna(row["INTRADAY_FORECAST"]).any()):
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                else:
                    mape_value = calculate_mape(day_actual_values, row["INTRADAY_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None
                
                    # Populate formatted_actual_data and formatted_forecast_data
                    base_date = row['D_F_DATE']
                    # print(row["A_DATE"])
                    forecast_day_values = row["INTRADAY_FORECAST"]
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })
                    
                    
                sr_mape_dict["data"].append(temp)

            return formatted_forecast_data, sr_mape_dict
        
        def day_ahead_mape(from_date, to_date, sr_actual_day_df):
            cursor.execute('''
                WITH RECURSIVE DateRange AS (
                    SELECT to_date('{0}', 'DD/MM/YYYY')::date AS upload_date
                    UNION ALL
                    SELECT (upload_date + INTERVAL '1 day')::date
                    FROM DateRange
                    WHERE upload_date + INTERVAL '1 day' <= to_date('{1}', 'DD/MM/YYYY')
                ),
                MaxRevisions AS (
                    SELECT state_id, upload_date, MAX(revision_no) AS max_revision
                    FROM file_uploads
                    WHERE upload_date BETWEEN to_date('{0}', 'DD/MM/YYYY') AND to_date('{1}', 'DD/MM/YYYY')
                    GROUP BY state_id, upload_date
                )
                SELECT dr.upload_date, t.state_id, t.revision_no, t.file_data
                FROM DateRange dr
                LEFT JOIN file_uploads t
                    ON dr.upload_date = t.upload_date
                    AND t.revision_no = (SELECT max_revision FROM MaxRevisions mr WHERE mr.state_id = t.state_id AND mr.upload_date = t.upload_date)
                ORDER BY dr.upload_date, t.state_id;
            '''.format(from_date, to_date))

            sr_day_data = cursor.fetchall()

            # Aggregate forecast data by date for selected state or region
            sr_day_forecast_dict = defaultdict(lambda: {"DAY_FORECAST": [0] * 96, "STATE_IDS": set()})
            for record in sr_day_data:
                upload_date, state_id, revision_no, file_data = record
                if state_id not in expected_state_ids:
                    continue
                date_str = upload_date.strftime("%Y-%m-%d")
                # print(date_str, state_id)
                
                if file_data:
                    day_forecast_values = [forecast[2] for forecast in file_data]
                    sr_day_forecast_dict[date_str]["STATE_IDS"].add(state_id)
                    sr_day_forecast_dict[date_str]["DAY_FORECAST"] = [
                        (float(x) if str(x).strip() not in ['', 'None', 'nan', 'NaN'] else 0.0) +
                        (float(y) if str(y).strip() not in ['', 'None', 'nan', 'NaN'] else 0.0)
                        for x, y in zip(sr_day_forecast_dict[date_str]["DAY_FORECAST"], day_forecast_values)
                    ]

            sr_day_forecast_json_list = [
                {"D_F_DATE": date, "DAY_FORECAST": data["DAY_FORECAST"], "STATE_IDS": list(data["STATE_IDS"])}
                for date, data in sr_day_forecast_dict.items()
            ]
            sr_forecast_day_df = pd.DataFrame(sr_day_forecast_json_list)

            formatted_forecast_data, formatted_actual_data = [], []
            sr_mape_dict = {"name": "Day Ahead", "data": []}


            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            # print(sr_forecast_day_df)
            if sr_forecast_day_df.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_forecast_day_df["D_F_DATE"] = pd.NaT
                sr_forecast_day_df["STATE_IDS"] = None
                sr_forecast_day_df["DAY_FORECAST"] = None
            else:
                sr_forecast_day_df["D_F_DATE"] = pd.to_datetime(sr_forecast_day_df["D_F_DATE"])
            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_forecast_day_df, on="D_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # print(sr_actual_day_df.shape)

            for _, row in sr_actual_day_df.iterrows():
                # print(row)
                day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                base_date = row['A_DATE']
                # print(day_actual_values)
                for i, actual_value in enumerate(day_actual_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(actual_value)
                        })




            for _, row in sr_result_df.iterrows():
                # print(row)
                temp = {"x": row["D_F_DATE"].strftime("%Y-%m-%d")}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()
                # print()
                forecast_day_values = row["DAY_FORECAST"]

                

                if sr_actual_day_df.empty:
                    for i in range(96):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })                    

                    

                if not isinstance(row["DAY_FORECAST"], list):
                    for i in range(96):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })


                elif forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    # print("If satisfied")
                    
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                elif (len(row["DAY_FORECAST"]) != 96 or pd.isna(row["DAY_FORECAST"]).any()):
                    # print("Elif satisfied")
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                else:
                    # print(row)
                    # print(day_actual_values)
                    mape_value = calculate_mape(row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL'], row["DAY_FORECAST"])
                    # print(row["D_F_DATE"],mape_value)
                    # print(row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL'])
                    # print(row["D_F_DATE"])
                    # print(row["DAY_FORECAST"])
                    # print(day_actual_values)
                    # print(row["DAY_ACTUAL"])

                    # print(mape_value)
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None
                
                    # Populate formatted_actual_data and formatted_forecast_data
                    base_date = row['D_F_DATE']
                    # print(row["A_DATE"])
                    forecast_day_values = row["DAY_FORECAST"]
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })
                    
                    
                sr_mape_dict["data"].append(temp)
            
            # print(formatted_actual_data)
            # if isinstance(formatted_actual_data, list):
            #     if len(formatted_actual_data) == 0:


            return formatted_forecast_data, formatted_actual_data, sr_mape_dict

        

        def week_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id):
            # Define expected state IDs based on selected_state_id
            region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
            expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

            cursor.execute('''
                WITH MaxRevisions AS (
                    SELECT
                        state_id,
                        from_date,
                        to_date,
                        MAX(revision_no) AS max_revision
                    FROM week_ahead_file_uploads
                    WHERE
                        (from_date <= to_date('{1}', 'DD/MM/YYYY') AND to_date >= to_date('{0}', 'DD/MM/YYYY'))
                    GROUP BY state_id, from_date, to_date
                )
                SELECT
                    t.state_id,
                    t.from_date,
                    t.to_date,
                    t.revision_no,
                    t.file_data
                FROM week_ahead_file_uploads t
                INNER JOIN MaxRevisions mr
                    ON t.state_id = mr.state_id
                    AND t.from_date = mr.from_date
                    AND t.to_date = mr.to_date
                    AND t.revision_no = mr.max_revision
                WHERE
                    t.state_id IN ({2})
                ORDER BY t.from_date;
            '''.format(from_date, to_date, ','.join(map(str, expected_state_ids))))

            sr_week_data = cursor.fetchall()

            # Expand week-ahead forecast data by date range for each state
            expanded_week_forecast = []
            for record in sr_week_data:
                state_id, from_date, to_date, _, file_data = record
                if state_id not in expected_state_ids:
                    continue
                if file_data:
                    week_forecast_values = [forecast[3] for forecast in file_data]

                    current_date = from_date
                    while current_date <= to_date:
                        forecast_index = (current_date - from_date).days * 96  # Start index for 96 intervals
                        daily_forecast_96 = week_forecast_values[forecast_index:forecast_index + 96]
                        
                        if len(daily_forecast_96) == 96:
                            expanded_week_forecast.append({
                                "W_F_DATE": current_date.strftime("%Y-%m-%d"),
                                "WEEK_FORECAST": daily_forecast_96,
                                "STATE_ID": state_id
                            })
                        else:
                            print(f"Warning: Insufficient forecast data for {current_date.strftime('%Y-%m-%d')}")
                        
                        current_date += timedelta(days=1)

            # Convert expanded forecast data to DataFrame
            sr_expanded_week_forecast_df = pd.DataFrame(expanded_week_forecast)

            # print(sr_expanded_week_forecast_df.columns)

            

            # Group by date and sum forecasts across states for each day
            # if not sr_expanded_week_forecast_df.empty and "W_F_DATE" in sr_expanded_week_forecast_df.columns:
            #     sr_week_forecast_daily = (
            #         sr_expanded_week_forecast_df.groupby("W_F_DATE").apply(
            #             lambda group: {
            #                 "WEEK_FORECAST": np.sum(np.vstack(group["WEEK_FORECAST"]), axis=0).tolist(),
            #                 "STATE_IDS": list(group["STATE_ID"])
            #             }
            #         )
            #     ).apply(pd.Series).reset_index()
            # else:
            #     print("sr_expanded_week_forecast_df is empty or missing W_F_DATE. Returning an empty result.")
            #     sr_week_forecast_daily = pd.DataFrame(columns=["W_F_DATE", "WEEK_FORECAST", "STATE_IDS"])

            if not sr_expanded_week_forecast_df.empty and "W_F_DATE" in sr_expanded_week_forecast_df.columns:
                def safe_convert_forecast(forecast_list):
                    try:
                        # Convert to float and ensure it's a list of length 96
                        arr = np.array([float(x) for x in forecast_list if str(x).strip() not in ['', 'nan', 'None']], dtype=np.float64)
                        return arr if len(arr) == 96 else None
                    except:
                        return None

                sr_week_forecast_daily = (
                    sr_expanded_week_forecast_df.groupby("W_F_DATE").apply(
                        lambda group: {
                            "WEEK_FORECAST": np.sum(
                                np.vstack([
                                    arr for arr in [safe_convert_forecast(f) for f in group["WEEK_FORECAST"]] if arr is not None
                                ]), axis=0
                            ).tolist(),
                            "STATE_IDS": list(group["STATE_ID"])
                        }
                    )
                ).apply(pd.Series).reset_index()
            else:
                print("sr_expanded_week_forecast_df is empty or missing W_F_DATE. Returning an empty result.")
                sr_week_forecast_daily = pd.DataFrame(columns=["W_F_DATE", "WEEK_FORECAST", "STATE_IDS"])



            formatted_week_forecast = []

            # Merge data frames
            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            # sr_week_forecast_daily["W_F_DATE"] = pd.to_datetime(sr_week_forecast_daily["W_F_DATE"])
            # sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            if sr_week_forecast_daily.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_week_forecast_daily["W_F_DATE"] = pd.NaT
                sr_week_forecast_daily["STATE_IDS"] = None
                sr_week_forecast_daily["WEEK_FORECAST"] = None
            else:
                sr_week_forecast_daily["W_F_DATE"] = pd.to_datetime(sr_week_forecast_daily["W_F_DATE"])

            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            # print(date_df.dtypes)
            # print(sr_week_forecast_daily.dtypes)

            sr_result_df = pd.merge(date_df, sr_week_forecast_daily, left_on="D_F_DATE", right_on="W_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # Calculate MAPE with complete state data only
            sr_mape_dict = {"name": "Week Ahead", "data": []}
            # print(sr_result_df.head())
            for _, row in sr_result_df.iterrows():
                temp = {"x": fmt_date(row["D_F_DATE"]) }
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()

                # print(forecast_state_ids)
                # print(actual_state_ids)
                # print(expected_state_ids)
                # print(row["WEEK_FORECAST"], "Z")

                if actual_state_ids != expected_state_ids:
                    if forecast_state_ids != expected_state_ids:
                        # print("A")
                        temp["y"] = None
                        for i in range(96):
                            timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                            formatted_week_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": 0
                            })
                    else:
                        for i, forecast_value in enumerate(row["WEEK_FORECAST"]):
                            timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                            formatted_week_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": round(forecast_value)
                            })

            

                elif (not isinstance(row["WEEK_FORECAST"], list) or len(row["WEEK_FORECAST"]) != 96 or pd.isna(row["WEEK_FORECAST"]).any()):
                    # print("B")
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                else:
                    # print("C")
                    day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                    mape_value = calculate_mape(day_actual_values, row["WEEK_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None

                    base_date = row['A_DATE']
                    for i, forecast_value in enumerate(row["WEEK_FORECAST"]):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })

                sr_mape_dict["data"].append(temp)

            return formatted_week_forecast, sr_mape_dict

        def month_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id):
            # Define expected state IDs based on selected_state_id
            region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
            expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

            cursor.execute('''
                WITH MaxRevisions AS (
                    SELECT
                        state_id,
                        from_date,
                        to_date,
                        MAX(revision_no) AS max_revision
                    FROM month_ahead_file_uploads
                    WHERE
                        (from_date <= to_date('{1}', 'DD/MM/YYYY') AND to_date >= to_date('{0}', 'DD/MM/YYYY'))
                    GROUP BY state_id, from_date, to_date
                )
                SELECT
                    t.state_id,
                    t.from_date,
                    t.to_date,
                    t.revision_no,
                    t.file_data
                FROM month_ahead_file_uploads t
                INNER JOIN MaxRevisions mr
                    ON t.state_id = mr.state_id
                    AND t.from_date = mr.from_date
                    AND t.to_date = mr.to_date
                    AND t.revision_no = mr.max_revision
                WHERE
                    t.state_id IN ({2})
                ORDER BY t.from_date;
            '''.format(from_date, to_date, ','.join(map(str, expected_state_ids))))

            sr_month_data = cursor.fetchall()

            # Expand month-ahead forecast data by date range for each state
            expanded_month_forecast = []
            for record in sr_month_data:
                state_id, from_date, to_date, _, file_data = record
                if state_id not in expected_state_ids:
                    continue
                if file_data:
                    month_forecast_values = [forecast[3] for forecast in file_data]

                    current_date = from_date
                    while current_date <= to_date:
                        forecast_index = (current_date - from_date).days * 96
                        daily_forecast_96 = month_forecast_values[forecast_index:forecast_index + 96]
                        
                        if len(daily_forecast_96) == 96:
                            expanded_month_forecast.append({
                                "M_F_DATE": current_date.strftime("%Y-%m-%d"),
                                "MONTH_FORECAST": daily_forecast_96,
                                "STATE_ID": state_id
                            })
                        else:
                            print(f"Warning: Insufficient forecast data for {current_date.strftime('%Y-%m-%d')}")
                        
                        current_date += timedelta(days=1)

            # Convert expanded forecast data to DataFrame
            sr_expanded_month_forecast_df = pd.DataFrame(expanded_month_forecast)

            # Group by date and sum forecasts across states for each day
            sr_month_forecast_daily = (
                sr_expanded_month_forecast_df.groupby("M_F_DATE").apply(
                    lambda group: {
                        "MONTH_FORECAST": np.sum(np.vstack(group["MONTH_FORECAST"]), axis=0).tolist(),
                        "STATE_IDS": list(group["STATE_ID"])
                    }
                )
            ).apply(pd.Series).reset_index()

            formatted_month_forecast = []

            # Merge data frames
            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])

            if sr_month_forecast_daily.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_month_forecast_daily["M_F_DATE"] = pd.NaT
                sr_month_forecast_daily["STATE_IDS"] = None
                sr_month_forecast_daily["MONTH_FORECAST"] = None
            else:
                sr_month_forecast_daily["M_F_DATE"] = pd.to_datetime(sr_month_forecast_daily["M_F_DATE"])

            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])
            
            
            sr_month_forecast_daily["M_F_DATE"] = pd.to_datetime(sr_month_forecast_daily["M_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_month_forecast_daily, left_on="D_F_DATE", right_on="M_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # Calculate MAPE with complete state data only
            sr_mape_dict = {"name": "Month Ahead", "data": []}
            for _, row in sr_result_df.iterrows():
                temp = {"x":  fmt_date(row["D_F_DATE"])}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()

                if actual_state_ids != expected_state_ids:
                    if forecast_state_ids != expected_state_ids:
                        # print("A")
                        temp["y"] = None
                        for i in range(96):
                            timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                            formatted_month_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": 0
                            })
                    else:
                        for i, forecast_value in enumerate(row["MONTH_FORECAST"]):
                            timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                            formatted_month_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": round(forecast_value)
                            })

                elif (len(row["MONTH_FORECAST"]) != 96 or pd.isna(row["MONTH_FORECAST"]).any()):
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                else:
                    day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                    mape_value = calculate_mape(day_actual_values, row["MONTH_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None

                    base_date = row['D_F_DATE']
                    for i, forecast_value in enumerate(row["MONTH_FORECAST"]):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })

                sr_mape_dict["data"].append(temp)

            return formatted_month_forecast, sr_mape_dict


        day_forecast_data, actual_data, day_mape = day_ahead_mape(from_date, to_date, sr_actual_day_df)
        week_forecast, week_mape = week_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id)  # Replace with desired state ID or 6 for region
        month_forecast, month_mape = month_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id)  # Replace with desired state ID or 6 for region
        intraday_forecast, intra_mape = intraday_mape(from_date, to_date, sr_actual_day_df)  # Replace with desired state ID or 6 for region

        # import pdb
        # pdb.set_trace()


        mape_list = []
        mape_list.append(day_mape)
        mape_list.append(week_mape)
        mape_list.append(month_mape)
        mape_list.append(intra_mape)

        # print(week_forecast)

        forecast_list = []
        forecast_list.append({"name": "Day Ahead", "data": day_forecast_data})
        forecast_list.append({"name": "Week Ahead", "data": week_forecast})
        forecast_list.append({"name": "Month Ahead", "data": month_forecast})
        forecast_list.append({"name": "Actual", "data": actual_data})
        forecast_list.append({"name": "Intraday", "data": intraday_forecast})

        final_forecast_dict = {}
        final_forecast_dict['data'] = forecast_list 
        final_forecast_dict['title'] = f"Comparison between Actual, Intraday, Day Ahead, Week Ahead and Month Ahead data  from {from_date} to {to_date} for {state_name if state_name != 'SRLDC' else 'SR'}"
        
        # print(mape_list)

        # comp_data = {}  # Replace this with your `final_forecast_dict`

        # # Prepare the JSON object
        # json_content = {
        #     "status": "success",
        #     "data": mape_list,
        #     "title": "MAPE for the data between {0} and {1} for {2}".format(from_date, to_date, state_name if state_name != 'SRLDC' else 'SR'),
        #     "comp_data": final_forecast_dict
        # }

        # # File path to save the JSON file
        # file_path = "mape_data.json"  # Update the path as needed

        # # Write to the JSON file
        # with open(file_path, "w") as json_file:
        #     json.dump(json_content, json_file, indent=4)

        # print(f"JSON file saved at: {file_path}")


        

        return jsonify(status="success", data=mape_list, title="MAPE for the data between {0} and {1} for {2}".format(from_date, to_date, state_name if state_name != 'SRLDC' else 'SR' ), comp_data=final_forecast_dict)        

    except Exception as e:
        log_error("mapechart", e)
        import traceback
        print(traceback.format_exc())
        cursor.close()
        return jsonify(message="There is some problem in Fetching the Data! Please contact SRLDC IT", status="failure")






@app.route('/api/breakupactualforecast', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_breakup_actual_forecast")
def breakupActualForecast():
    try:
        conn = psycopg2.connect(
        database="demand_forecast_states", user='prasadbabu', 
        password='BabuPrasad#123', host='10.0.100.79', port='5432'
        )
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)

        # jwt_data = get_jwt()

        region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region

        # Define expected state IDs based on selected_state_id
        # expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

        from_date = params["params"]["fromDate"]
        to_date = params["params"]["toDate"]  
        state_id = int(params["params"]["state"])
        selected_option = int(params["params"]["indexNumber"])

        # print(params, selected_option)

        cursor.execute("select state_name from states where state_id = {0}".format(state_id))
        state_name = cursor.fetchall()[0][0]

        region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
        selected_state_id = state_id  # Set to the desired single state or region (use 6 for the entire region)

        # Define expected state IDs based on selected_state_id
        expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

        date_range = pd.date_range(start=datetime.strptime(from_date, "%d/%m/%Y"), end=datetime.strptime(to_date, "%d/%m/%Y"))
        date_df = pd.DataFrame(date_range, columns=["D_F_DATE"])


        # Function to fetch actual demand data (Only for selected_option == 2)
        def fetch_actual_demand(cursor, from_date, to_date, state_id):
            print(from_date, to_date, state_id)
            query = f"""
                SELECT s.state_id, process_date, demand_met, scada_demand
                FROM states s
                JOIN actual_demand ad ON s.state_id = ad.state_id
                WHERE process_date BETWEEN TO_DATE('{from_date}', 'DD/MM/YYYY') AND TO_DATE('{to_date}', 'DD/MM/YYYY')
                ORDER BY process_date;
            """
            cursor.execute(query)
            # print("Funtion hit!")
            return cursor.fetchall()

        # Function to fetch SCADA data and aggregate into 15-minute intervals
        def fetch_scada_data(scada_key, from_date, to_date):
            
            scada_conn = psycopg2.connect(**scada_db_params)
            scada_cursor = scada_conn.cursor()
            # print(from_date, to_date, scada_key)
            # print("before Query")
            query = f"""
                WITH time_series AS (
                    SELECT generate_series(
                        date_trunc('day', '{datetime.strptime(from_date, "%d/%m/%Y").strftime("%Y-%m-%d")}'::timestamptz),
                        date_trunc('day', '{datetime.strptime(to_date, "%d/%m/%Y").strftime("%Y-%m-%d")}'::timestamptz) + interval '23 hours 45 minutes',
                        interval '15 minutes'
                    ) AS interval_start
                ),
                agg_data AS (
                    SELECT 
                        time_bucket('15 minutes', timestamp) AS interval_start,
                        ROUND(AVG(d_val))::bigint AS avg_d_val
                    FROM scada.measurements_optimized
                    WHERE scada_key = '{scada_key}'
                    AND timestamp >= '{datetime.strptime(from_date, "%d/%m/%Y").strftime("%Y-%m-%d")}'::timestamptz
                    AND timestamp <= '{datetime.strptime(to_date, "%d/%m/%Y").strftime("%Y-%m-%d")}'::timestamptz + interval '23 hours 45 minutes'
                    GROUP BY interval_start
                )
                SELECT 
                    date_trunc('day', ts.interval_start) AS day,
                    ARRAY_AGG(COALESCE(ad.avg_d_val, 0) ORDER BY ts.interval_start) AS avg_d_val_list
                FROM time_series ts
                LEFT JOIN agg_data ad
                    ON ts.interval_start = ad.interval_start
                GROUP BY day
                ORDER BY day;
            """



            # print(query)
            
            scada_cursor.execute(query)
            scada_data = scada_cursor.fetchall()

            # print(scada_data)


            # print("After Query")
            # Convert to DataFrame

            # print(scada_df.tail())

            scada_data = [(day, [float(val) for val in avg_d_val_list]) for day, avg_d_val_list in scada_data]

            # print(scada_df.head())
            # print(scada_df.shape)

            # print("scada data", scada_data[0])

            scada_cursor.close()
            scada_conn.close()

            return scada_data


        


        # Fetch Actual Data

        
        # Connect to forecast database
        forecast_conn = psycopg2.connect(**db_params)
        forecast_cursor = forecast_conn.cursor()

        # Connect to SCADA database
        # scada_conn = psycopg2.connect(**scada_db_params)
        # scada_cursor = scada_conn.cursor()


        # print("Connected Both databases!")

        # Initialize daily totals structure
        # Initialize daily_totals structure
        daily_totals = defaultdict(lambda: {
            "DAY_ACTUAL": [0] * 96,
            "SCADA_DEMAND": [0] * 96,
            "SCADA_DEMAND": [0] * 96,  # New field for SCADA-based generation types
            "STATE_IDS": set()
        })


        # print("before if")

        # print("Selected Option: ", selected_option)

        # Case 1: Fetch Actual Demand (Forecasted Demand)
        if selected_option == 2:
            # print("before Actual data")
            actual_data = fetch_actual_demand(forecast_cursor, from_date, to_date, state_id)
            
            # print("After Actual Data")
            # print(actual_data)

            for record in actual_data:
                state_id, process_date, demand_met_list, scada_demand_list = record
                if state_id not in expected_state_ids:  # Filter out any unexpected states
                    continue
                date_str = process_date.strftime("%Y-%m-%d")

                if demand_met_list is None:
                    demand_met_list = [0] * 96
                if scada_demand_list is None:
                    scada_demand_list = [0] * 96

                daily_totals[date_str]["STATE_IDS"].add(state_id)

                # print(demand_met_list)
                # print("#####")
                # print(scada_demand_list)

                # If state 3 has missing demand data, replace it with SCADA demand
                if (not demand_met_list or all(x == 0 for x in demand_met_list)):
                    demand_met_list = scada_demand_list

                daily_totals[date_str]["DAY_ACTUAL"] = [x + y for x, y in zip(daily_totals[date_str]["DAY_ACTUAL"], demand_met_list)]
                daily_totals[date_str]["SCADA_DEMAND"] = [x + y for x, y in zip(daily_totals[date_str]["SCADA_DEMAND"], scada_demand_list)]

        # Case 2: Fetch SCADA Data (Thermal, Solar, Wind, etc.)
        elif selected_option in range(3,9):
            # print("Entered else")
            if selected_option in keys_dict and state_id in keys_dict[selected_option]:
                scada_key = keys_dict[selected_option][state_id]
                # print(scada_key)
                scada_data = fetch_scada_data( scada_key, from_date, to_date)

                # print("Before Scada")
                # print(scada_data)
                # print("After Scada")

                for record in scada_data:
                    process_date, scada_demand_list = record
                    # if state_id not in expected_state_ids:  # Filter out any unexpected states
                    #     continue
                    date_str = process_date.strftime("%Y-%m-%d")



                    # if demand_met_list is None:
                    #     demand_met_list = [0] * 96
                    if scada_demand_list is None:
                        scada_demand_list = [0] * 96

                    daily_totals[date_str]["STATE_IDS"] = region_state_ids if state_id == 6 else {state_id}



                    # print(demand_met_list)
                    # print("#####")
                    # print(scada_demand_list)

                    # If state 3 has missing demand data, replace it with SCADA demand
                    # If state 3 has missing demand data, replace it with SCADA demand
                    # if state_id == 3 and (not demand_met_list or all(x == 0 for x in demand_met_list)):
                    #     demand_met_list = scada_demand_list

                    # daily_totals[date_str]["DAY_ACTUAL"] = [x + y for x, y in zip(daily_totals[date_str]["DAY_ACTUAL"], demand_met_list)]
                    daily_totals[date_str]["SCADA_DEMAND"] = [x + y for x, y in zip(daily_totals[date_str]["SCADA_DEMAND"], scada_demand_list)]

        elif selected_option == 9:
            pass # Others Demand - All others 

        else:
            pass # WBES 


                # # Ensure 96 values per day
                # if len(scada_data) != 96:
                #     scada_data = [0] * 96  # Default to 0 if data is incomplete

                # for date in daily_totals:
                #     daily_totals[date]["SCADA_DEMAND"] = scada_data


        # print(daily_totals)

        # print("Rinnegan")

        # Convert results into JSON format
        sr_day_actual_json_list = [
            {
                "A_DATE": date,
                "DAY_ACTUAL": totals["DAY_ACTUAL"] if selected_option == 2 else None  ,
                "SCADA_DEMAND": totals["SCADA_DEMAND"],
                # "SCADA_GENERATION": totals["SCADA_DEMAND"],
                "STATE_IDS": list(totals["STATE_IDS"])
            }
            for date, totals in daily_totals.items()
        ]




        sr_actual_day_df = pd.DataFrame(sr_day_actual_json_list)
        # print(sr_actual_day_df.head())
        # import pdb
        # pdb.set_trace()

        # print(sr_actual_day_df.head())
        
        
        def intraday_mape(from_date, to_date, sr_actual_day_df):
            cursor.execute('''
                WITH RECURSIVE DateRange AS (
                    SELECT to_date('{0}', 'DD/MM/YYYY')::date AS upload_date
                    UNION ALL
                    SELECT (upload_date + INTERVAL '1 day')::date
                    FROM DateRange
                    WHERE upload_date + INTERVAL '1 day' <= to_date('{1}', 'DD/MM/YYYY')
                ),
                MaxRevisions AS (
                    SELECT state_id, upload_date, MAX(revision_no) AS max_revision
                    FROM intraday_file_uploads
                    WHERE upload_date BETWEEN to_date('{0}', 'DD/MM/YYYY') AND to_date('{1}', 'DD/MM/YYYY')
                    AND file_type = 'I'
                    GROUP BY state_id, upload_date
                )
                SELECT dr.upload_date, t.state_id, t.revision_no, t.file_data
                FROM DateRange dr
                LEFT JOIN intraday_file_uploads t
                    ON dr.upload_date = t.upload_date
                    AND t.revision_no = (
                        SELECT max_revision
                        FROM MaxRevisions mr
                        WHERE mr.state_id = t.state_id AND mr.upload_date = t.upload_date
                    )
                    AND t.file_type = 'I'
                ORDER BY dr.upload_date, t.state_id;
            '''.format(from_date, to_date))

            sr_day_data = cursor.fetchall()

            # Aggregate forecast data by date for selected state or region
            sr_day_forecast_dict = defaultdict(lambda: {"INTRADAY_FORECAST": [0] * 96, "STATE_IDS": set()})
            for record in sr_day_data:
                upload_date, state_id, revision_no, file_data = record
                if state_id not in expected_state_ids:
                    continue
                date_str = upload_date.strftime("%Y-%m-%d")
                
                if file_data:
                    day_forecast_values = [
                        float(forecast[selected_option]) if forecast[selected_option] not in [None, '', ' '] else 0.0
                        for forecast in file_data
                    ]
                    sr_day_forecast_dict[date_str]["INTRADAY_FORECAST"] = [x + y for x, y in zip(sr_day_forecast_dict[date_str]["INTRADAY_FORECAST"], day_forecast_values)]

            sr_day_forecast_json_list = [
                {"D_F_DATE": date, "INTRADAY_FORECAST": data["INTRADAY_FORECAST"], "STATE_IDS": list(data["STATE_IDS"])}
                for date, data in sr_day_forecast_dict.items()
            ]
            # Create the DataFrame
            if sr_day_forecast_json_list:  # If the list is not empty
                sr_forecast_day_df = pd.DataFrame(sr_day_forecast_json_list)
            else:  # If the list is empty, create an empty DataFrame with required columns
                sr_forecast_day_df = pd.DataFrame(columns=["D_F_DATE", "INTRADAY_FORECAST", "STATE_IDS"])

            formatted_forecast_data, formatted_actual_data = [], []
            sr_mape_dict = {"name": "Intraday", "data": []}

            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            if sr_forecast_day_df.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_forecast_day_df["D_F_DATE"] = pd.NaT
                sr_forecast_day_df["STATE_IDS"] = None
                sr_forecast_day_df["INTRADAY_FORECAST"] = None
            else:
                sr_forecast_day_df["D_F_DATE"] = pd.to_datetime(sr_forecast_day_df["D_F_DATE"])
            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_forecast_day_df, on="D_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            for _, row in sr_actual_day_df.iterrows():
                day_actual_values = row['SCADA_DEMAND'] if row["DAY_ACTUAL"] is None or all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                base_date = row['A_DATE']
                for i, actual_value in enumerate(day_actual_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(actual_value or 0)
                        })




            for _, row in sr_result_df.iterrows():
                # print(row)
                temp = {"x": row["D_F_DATE"].strftime("%Y-%m-%d")}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()
                # print()
                forecast_day_values = row["INTRADAY_FORECAST"]

                

                if sr_actual_day_df.empty:
                    for i in range(96):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })                    

                    

                if not isinstance(row["INTRADAY_FORECAST"], list):
                    for i in range(96):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None

                elif forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    # print("If satisfied")
                    
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                elif (len(row["INTRADAY_FORECAST"]) != 96 or pd.isna(row["INTRADAY_FORECAST"]).any()):
                    # print("Elif satisfied")
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                else:
                    # print(row)
                    # print(day_actual_values)
                    mape_value = calculate_mape(row['SCADA_DEMAND'] if row["DAY_ACTUAL"] is None or all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL'], row["INTRADAY_FORECAST"])
                    # print(row["D_F_DATE"],mape_value)
                    # print(row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL'])
                    # print(row["D_F_DATE"])
                    # print(row["DAY_FORECAST"])
                    # print(day_actual_values)
                    # print(row["DAY_ACTUAL"])

                    # print(mape_value)
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None
                
                    # Populate formatted_actual_data and formatted_forecast_data
                    base_date = row['D_F_DATE']
                    # print(row["A_DATE"])
                    forecast_day_values = row["INTRADAY_FORECAST"]
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value or 0)
                        })
                    
                    
                sr_mape_dict["data"].append(temp)

            return formatted_forecast_data, sr_mape_dict
 
        def day_ahead_mape(from_date, to_date, sr_actual_day_df, selected_option):
            cursor.execute('''
                WITH RECURSIVE DateRange AS (
                    SELECT to_date('{0}', 'DD/MM/YYYY')::date AS upload_date
                    UNION ALL
                    SELECT (upload_date + INTERVAL '1 day')::date
                    FROM DateRange
                    WHERE upload_date + INTERVAL '1 day' <= to_date('{1}', 'DD/MM/YYYY')
                ),
                MaxRevisions AS (
                    SELECT state_id, upload_date, MAX(revision_no) AS max_revision
                    FROM file_uploads
                    WHERE upload_date BETWEEN to_date('{0}', 'DD/MM/YYYY') AND to_date('{1}', 'DD/MM/YYYY')
                    GROUP BY state_id, upload_date
                )
                SELECT dr.upload_date, t.state_id, t.revision_no, t.file_data
                FROM DateRange dr
                LEFT JOIN file_uploads t
                    ON dr.upload_date = t.upload_date
                    AND t.revision_no = (SELECT max_revision FROM MaxRevisions mr WHERE mr.state_id = t.state_id AND mr.upload_date = t.upload_date)
                ORDER BY dr.upload_date, t.state_id;
            '''.format(from_date, to_date))

            sr_day_data = cursor.fetchall()

            # Aggregate forecast data by date for selected state or region

            # print("SR day data")
            sr_day_forecast_dict = defaultdict(lambda: {"DAY_FORECAST": [0] * 96, "STATE_IDS": set()})
            for record in sr_day_data:
                upload_date, state_id, revision_no, file_data = record
                if state_id not in expected_state_ids:
                    continue
                date_str = upload_date.strftime("%Y-%m-%d")
                
                if file_data:
                    day_forecast_values = [
                        float(forecast[selected_option]) if forecast[selected_option] not in [None, '', ' '] else 0.0
                        for forecast in file_data
                    ]
                    sr_day_forecast_dict[date_str]["STATE_IDS"].add(state_id)
                    sr_day_forecast_dict[date_str]["DAY_FORECAST"] = [x + y for x, y in zip(sr_day_forecast_dict[date_str]["DAY_FORECAST"], day_forecast_values)]

            sr_day_forecast_json_list = [
                {"D_F_DATE": date, "DAY_FORECAST": data["DAY_FORECAST"], "STATE_IDS": list(data["STATE_IDS"])}
                for date, data in sr_day_forecast_dict.items()
            ]
            sr_forecast_day_df = pd.DataFrame(sr_day_forecast_json_list)

            formatted_forecast_data, formatted_actual_data = [], []
            sr_mape_dict = {"name": "Day Ahead", "data": []}


            # print("Empty SR Mape dict")


            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            # print(sr_forecast_day_df)
            if sr_forecast_day_df.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_forecast_day_df["D_F_DATE"] = pd.NaT
                sr_forecast_day_df["STATE_IDS"] = None
                sr_forecast_day_df["DAY_FORECAST"] = None
            else:
                sr_forecast_day_df["D_F_DATE"] = pd.to_datetime(sr_forecast_day_df["D_F_DATE"])
            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_forecast_day_df, on="D_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # print(sr_actual_day_df.shape)

            # print("Before sr_actual_day_df")

            for _, row in sr_actual_day_df.iterrows():
                # print(row)
                # day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                day_actual_values = row['SCADA_DEMAND'] if row["DAY_ACTUAL"] is None or all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']

                base_date = row['A_DATE']
                # print(day_actual_values)
                for i, actual_value in enumerate(day_actual_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(actual_value or 0)
                        })

            # print("Before SR result df")


            for _, row in sr_result_df.iterrows():
                # print(row)
                temp = {"x": row["D_F_DATE"].strftime("%Y-%m-%d")}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()
                # print()
                forecast_day_values = row["DAY_FORECAST"]

                

                if sr_actual_day_df.empty:
                    for i in range(96):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })                    

                    

                if not isinstance(row["DAY_FORECAST"], list):
                    for i in range(96):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })


                elif forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    # print("If satisfied")
                    
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                elif (len(row["DAY_FORECAST"]) != 96 or pd.isna(row["DAY_FORECAST"]).any()):
                    # print("Elif satisfied")
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                else:
                    # print(row)
                    # print(day_actual_values)
                    mape_value = calculate_mape(row['SCADA_DEMAND'] if row["DAY_ACTUAL"] is None or all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL'], row["DAY_FORECAST"])
                    # print(row["D_F_DATE"],mape_value)
                    # print(row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL'])
                    # print(row["D_F_DATE"])
                    # print(row["DAY_FORECAST"])
                    # print(day_actual_values)
                    # print(row["DAY_ACTUAL"])

                    # print(mape_value)
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None
                
                    # Populate formatted_actual_data and formatted_forecast_data
                    base_date = row['D_F_DATE']
                    # print(row["A_DATE"])
                    forecast_day_values = row["DAY_FORECAST"]
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value or 0)
                        })
                    
                    
                sr_mape_dict["data"].append(temp)
            
            # print(formatted_actual_data)
            # if isinstance(formatted_actual_data, list):
            #     if len(formatted_actual_data) == 0:

            # import pdb
            # pdb.set_trace()
            
            # print("After sr_result_day_df")

            return formatted_forecast_data, formatted_actual_data, sr_mape_dict


       


        def week_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id, selected_option):
            # Define expected state IDs based on selected_state_id
            region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
            expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

            cursor.execute('''
                WITH MaxRevisions AS (
                    SELECT
                        state_id,
                        from_date,
                        to_date,
                        MAX(revision_no) AS max_revision
                    FROM week_ahead_file_uploads
                    WHERE
                        (from_date <= to_date('{1}', 'DD/MM/YYYY') AND to_date >= to_date('{0}', 'DD/MM/YYYY'))
                    GROUP BY state_id, from_date, to_date
                )
                SELECT
                    t.state_id,
                    t.from_date,
                    t.to_date,
                    t.revision_no,
                    t.file_data
                FROM week_ahead_file_uploads t
                INNER JOIN MaxRevisions mr
                    ON t.state_id = mr.state_id
                    AND t.from_date = mr.from_date
                    AND t.to_date = mr.to_date
                    AND t.revision_no = mr.max_revision
                WHERE
                    t.state_id IN ({2})
                ORDER BY t.from_date;
            '''.format(from_date, to_date, ','.join(map(str, expected_state_ids))))

            sr_week_data = cursor.fetchall()

            # Expand week-ahead forecast data by date range for each state
            expanded_week_forecast = []
            for record in sr_week_data:
                state_id, from_date, to_date, _, file_data = record
                if state_id not in expected_state_ids:
                    continue
                if file_data:
                    # week_forecast_values = [forecast[selected_option + 1] for forecast in file_data]

                    week_forecast_values = [
                        float(forecast[selected_option + 1]) if forecast[selected_option + 1] not in [None, '', ' '] else 0.0
                        for forecast in file_data
                    ]

                    current_date = from_date
                    while current_date <= to_date:
                        forecast_index = (current_date - from_date).days * 96  # Start index for 96 intervals
                        daily_forecast_96 = week_forecast_values[forecast_index:forecast_index + 96]
                        
                        if len(daily_forecast_96) == 96:
                            expanded_week_forecast.append({
                                "W_F_DATE": current_date.strftime("%Y-%m-%d"),
                                "WEEK_FORECAST": daily_forecast_96,
                                "STATE_ID": state_id
                            })
                        else:
                            print(f"Warning: Insufficient forecast data for {current_date.strftime('%Y-%m-%d')}")
                        
                        current_date += timedelta(days=1)

            # Convert expanded forecast data to DataFrame
            sr_expanded_week_forecast_df = pd.DataFrame(expanded_week_forecast)

            # print(sr_expanded_week_forecast_df.columns)

            # print("before SR expanded")

            # Group by date and sum forecasts across states for each day
            if not sr_expanded_week_forecast_df.empty and "W_F_DATE" in sr_expanded_week_forecast_df.columns:
                def convert_to_numeric(arr):
                    """ Convert list elements to float, handling non-numeric values """
                    try:
                        return [float(x) if isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).isdigit() else 0 for x in arr]
                    except:
                        return [0] * 96  # Default to a zero-filled list if conversion fails

                sr_week_forecast_daily = (
                    sr_expanded_week_forecast_df.groupby("W_F_DATE").apply(
                        lambda group: {
                            "WEEK_FORECAST": np.sum(
                                np.vstack(
                                    [convert_to_numeric(x) for x in group["WEEK_FORECAST"] if isinstance(x, list)]
                                ), axis=0
                            ).tolist(),
                            "STATE_IDS": list(group["STATE_ID"])
                        }
                    )
                ).apply(pd.Series).reset_index()
            else:
                print("sr_expanded_week_forecast_df is empty or missing W_F_DATE. Returning an empty result.")
                sr_week_forecast_daily = pd.DataFrame(columns=["W_F_DATE", "WEEK_FORECAST", "STATE_IDS"])

            formatted_week_forecast = []

            # Merge data frames
            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            # sr_week_forecast_daily["W_F_DATE"] = pd.to_datetime(sr_week_forecast_daily["W_F_DATE"])
            # sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            if sr_week_forecast_daily.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_week_forecast_daily["W_F_DATE"] = pd.NaT
                sr_week_forecast_daily["STATE_IDS"] = None
                sr_week_forecast_daily["WEEK_FORECAST"] = None
            else:
                sr_week_forecast_daily["W_F_DATE"] = pd.to_datetime(sr_week_forecast_daily["W_F_DATE"])

            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            # print(date_df.dtypes)
            # print(sr_week_forecast_daily.dtypes)

            sr_result_df = pd.merge(date_df, sr_week_forecast_daily, left_on="D_F_DATE", right_on="W_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # Calculate MAPE with complete state data only
            sr_mape_dict = {"name": "Week Ahead", "data": []}
            # print(sr_result_df.head())
            # print("Before SR result df")

            for _, row in sr_result_df.iterrows():
                temp = {"x": row["D_F_DATE"].strftime('%Y-%m-%d')}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()

                # print(forecast_state_ids)
                # print(actual_state_ids)
                # print(expected_state_ids)
                # print(row["WEEK_FORECAST"], "Z")

                if actual_state_ids != expected_state_ids:
                    if forecast_state_ids != expected_state_ids:
                        # print("A")
                        temp["y"] = None
                        for i in range(96):
                            timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                            formatted_week_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": 0
                            })
                    else:
                        for i, forecast_value in enumerate(row["WEEK_FORECAST"]):
                            timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                            formatted_week_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": round(forecast_value or 0)
                            })

            

                elif (not isinstance(row["WEEK_FORECAST"], list) or len(row["WEEK_FORECAST"]) != 96 or pd.isna(row["WEEK_FORECAST"]).any()):
                    # print("B")
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                else:
                    # print("C")
                    day_actual_values = row['SCADA_DEMAND'] if row["DAY_ACTUAL"] is None or all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                    mape_value = calculate_mape(day_actual_values, row["WEEK_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None

                    base_date = row['A_DATE']
                    for i, forecast_value in enumerate(row["WEEK_FORECAST"]):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value or 0)
                        })

                sr_mape_dict["data"].append(temp)

            return formatted_week_forecast, sr_mape_dict

        def month_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id, selected_option):
            # Define expected state IDs based on selected_state_id
            region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
            expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

            cursor.execute('''
                WITH MaxRevisions AS (
                    SELECT
                        state_id,
                        from_date,
                        to_date,
                        MAX(revision_no) AS max_revision
                    FROM month_ahead_file_uploads
                    WHERE
                        (from_date <= to_date('{1}', 'DD/MM/YYYY') AND to_date >= to_date('{0}', 'DD/MM/YYYY'))
                    GROUP BY state_id, from_date, to_date
                )
                SELECT
                    t.state_id,
                    t.from_date,
                    t.to_date,
                    t.revision_no,
                    t.file_data
                FROM month_ahead_file_uploads t
                INNER JOIN MaxRevisions mr
                    ON t.state_id = mr.state_id
                    AND t.from_date = mr.from_date
                    AND t.to_date = mr.to_date
                    AND t.revision_no = mr.max_revision
                WHERE
                    t.state_id IN ({2})
                ORDER BY t.from_date;
            '''.format(from_date, to_date, ','.join(map(str, expected_state_ids))))

            sr_month_data = cursor.fetchall()

            # Expand month-ahead forecast data by date range for each state
            expanded_month_forecast = []
            for record in sr_month_data:
                state_id, from_date, to_date, _, file_data = record
                if state_id not in expected_state_ids:
                    continue
                if file_data:
                    # month_forecast_values = [forecast[selected_option+1] for forecast in file_data]

                    month_forecast_values = [
                        float(forecast[selected_option+1]) if forecast[selected_option+1] not in [None, '', ' '] else 0.0
                        for forecast in file_data
                    ]

                    current_date = from_date
                    while current_date <= to_date:
                        forecast_index = (current_date - from_date).days * 96
                        daily_forecast_96 = month_forecast_values[forecast_index:forecast_index + 96]
                        
                        if len(daily_forecast_96) == 96:
                            expanded_month_forecast.append({
                                "M_F_DATE": current_date.strftime("%Y-%m-%d"),
                                "MONTH_FORECAST": daily_forecast_96,
                                "STATE_ID": state_id
                            })
                        else:
                            print(f"Warning: Insufficient forecast data for {current_date.strftime('%Y-%m-%d')}")
                        
                        current_date += timedelta(days=1)

            # Convert expanded forecast data to DataFrame
            sr_expanded_month_forecast_df = pd.DataFrame(expanded_month_forecast)

            # Group by date and sum forecasts across states for each day


            def convert_to_numeric(arr):
                """ Convert list elements to float, handling non-numeric values """
                try:
                    return [float(x) if isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).isdigit() else 0 for x in arr]
                except:
                    return [0] * 96  # Default to a zero-filled list if conversion fails

            sr_month_forecast_daily = (
                sr_expanded_month_forecast_df.groupby("M_F_DATE").apply(
                    lambda group: {
                        "MONTH_FORECAST": np.sum(
                            np.vstack(
                                [convert_to_numeric(x) for x in group["MONTH_FORECAST"] if isinstance(x, list)]
                            ), axis=0
                        ).tolist(),
                        "STATE_IDS": list(group["STATE_ID"])
                    }
                )
            ).apply(pd.Series).reset_index()
                        

            formatted_month_forecast = []

            # Merge data frames
            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])

            if sr_month_forecast_daily.empty:
                # print("sr_forecast_day_df is empty. Creating D_F_DATE column as a placeholder.")
                sr_month_forecast_daily["M_F_DATE"] = pd.NaT
                sr_month_forecast_daily["STATE_IDS"] = None
                sr_month_forecast_daily["MONTH_FORECAST"] = None
            else:
                sr_month_forecast_daily["M_F_DATE"] = pd.to_datetime(sr_month_forecast_daily["M_F_DATE"])

            if sr_actual_day_df.empty:
                sr_actual_day_df["A_DATE"] = pd.NaT
                sr_actual_day_df["STATE_IDS"] = None
            else:
                sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])
            
            
            sr_month_forecast_daily["M_F_DATE"] = pd.to_datetime(sr_month_forecast_daily["M_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_month_forecast_daily, left_on="D_F_DATE", right_on="M_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # Calculate MAPE with complete state data only
            sr_mape_dict = {"name": "Month Ahead", "data": []}
            for _, row in sr_result_df.iterrows():
                temp = {"x": row["D_F_DATE"].strftime('%Y-%m-%d')}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()

                if actual_state_ids != expected_state_ids:
                    if forecast_state_ids != expected_state_ids:
                        # print("A")
                        temp["y"] = None
                        for i in range(96):
                            timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                            formatted_month_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": 0
                            })
                    else:
                        for i, forecast_value in enumerate(row["MONTH_FORECAST"]):
                            timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                            formatted_month_forecast.append({
                                "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                                "y": round(forecast_value or 0)
                            })

                elif (len(row["MONTH_FORECAST"]) != 96 or pd.isna(row["MONTH_FORECAST"]).any()):
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                else:
                    day_actual_values = row['SCADA_DEMAND'] if row["DAY_ACTUAL"] is None or all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                    mape_value = calculate_mape(day_actual_values, row["MONTH_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None

                    base_date = row['D_F_DATE']
                    for i, forecast_value in enumerate(row["MONTH_FORECAST"]):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value or 0)
                        })

                sr_mape_dict["data"].append(temp)

            return formatted_month_forecast, sr_mape_dict


        day_forecast_data, actual_data, day_mape = day_ahead_mape(from_date, to_date, sr_actual_day_df, selected_option)
        week_forecast, week_mape = week_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id, selected_option)  # Replace with desired state ID or 6 for region
        month_forecast, month_mape = month_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id, selected_option)  # Replace with desired state ID or 6 for region
        intraday_forecast, intra_mape = intraday_mape(from_date, to_date, sr_actual_day_df)  # Replace with desired state ID or 6 for region

        # import pdb
        # pdb.set_trace()


        mape_list = []
        mape_list.append(day_mape)
        mape_list.append(week_mape)
        mape_list.append(month_mape)
        mape_list.append(intra_mape)

        # print(week_forecast)

        forecast_list = []
        forecast_list.append({"name": "Day Ahead", "data": day_forecast_data})
        forecast_list.append({"name": "Week Ahead", "data": week_forecast})
        forecast_list.append({"name": "Month Ahead", "data": month_forecast})
        forecast_list.append({"name": "Actual", "data": actual_data})
        forecast_list.append({"name": "Intraday", "data": intraday_forecast})


        # print(mape_list)

        option_mapping = {
            2: "Forecasted Demand",
            3: "Thermal",
            4: "Gas",
            5: "Hydro",
            7: "Solar",
            8: "Wind",
            9: "Other RES (biomass etc.)",
            10: "From ISGS & Other LTA & MTOA",
            11: "From Bilateral Transaction (Advance + FCFS)"
        }

        selected_label = option_mapping.get(selected_option, "Unknown Category")

   

        final_forecast_dict = {}
        final_forecast_dict['data'] = forecast_list 
        final_forecast_dict['title'] = f"Comparison between Actual, Day Ahead, Week Ahead and Month Ahead {selected_label} data  from {from_date} to {to_date} for {state_name if state_name != 'SRLDC' else 'SR'}"


        # import pdb
        # pdb.set_trace()








        forecast_cursor.close()
        forecast_conn.close()
        # scada_cursor.close()
        # scada_conn.close()




        







        # print("Hit API")

        return jsonify(status="success", data=mape_list, title="MAPE for the  {3} data between {0} and {1} for {2}".format(from_date, to_date, state_name if state_name != 'SRLDC' else 'SR', selected_label ), comp_data=final_forecast_dict)        

    except Exception as e:
        error_details = traceback.format_exc()  # Get the full traceback
        # log_error("breakupforecast", f"{e}\n{error_details}")  # Log error with details
        print(f"Error in breakupforecast: {e}\n{error_details}")  # Print error with traceback
        log_error("breakupforecast", e)
        return jsonify(status="failure")





@app.route('/api/getdatastatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_data_status")
def getDataStatus():
    try:
        import psycopg2.extras
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        payload = request.get_json() or {}
        params = payload.get("params", payload)
        # if nested (some callers send { params: { params: {...} }})
        if isinstance(params, dict) and "params" in params:
            params = params.get("params", params)

        # print("DEBUG: getDataStatus called with params:", params)

        # Basic inputs
        state_raw = params.get("state")
        data_type = params.get("dataType")
        from_date = params.get("fromDate")
        to_date = params.get("toDate")

        if not all([data_type, from_date, to_date]):
            print("ERROR: Missing required parameters")
            return jsonify({"error": "Missing required parameters"}), HTTPStatus.BAD_REQUEST

        # Parse state id if present
        state_id = None
        if state_raw not in [None, '', 'null']:
            try:
                state_id = int(state_raw)
            except Exception as ex:
                # print(f"WARN: could not parse state id '{state_raw}': {ex}")
                state_id = None

        # Which table / offsets / columns
        if data_type == "dayahead":
            table_name = "file_uploads"
            start_index = 2
            forecast_columns = BASE_FORECAST_COLUMNS + DAY_AHEAD_EXTRA_COLUMNS
        elif data_type == "weekahead":
            table_name = "week_ahead_file_uploads"
            start_index = 3
            forecast_columns = BASE_FORECAST_COLUMNS
        elif data_type == "monthahead":
            table_name = "month_ahead_file_uploads"
            start_index = 3
            forecast_columns = BASE_FORECAST_COLUMNS
        elif data_type == "intraday":
            table_name = "intraday_file_uploads"
            start_index = 2
            forecast_columns = BASE_FORECAST_COLUMNS
        else:
            print(f"ERROR: Invalid dataType {data_type}")
            return jsonify({"error": f"Invalid dataType {data_type}"}), HTTPStatus.BAD_REQUEST

        # prepare date list (for weekahead, align weeks if needed)
        start = datetime.strptime(from_date, "%d/%m/%Y").date()
        end = datetime.strptime(to_date, "%d/%m/%Y").date()
        if end < start:
            # print("ERROR: toDate must be >= fromDate")
            return jsonify({"error": "toDate must be >= fromDate"}), HTTPStatus.BAD_REQUEST

        def monday(d): return d - timedelta(days=d.weekday())
        def sunday(d): return d + timedelta(days=(6 - d.weekday()))

        if data_type == "weekahead":
            # build list of all dates in requested range but iterating week-chunks as required
            cur = monday(start)
            all_dates = []
            while cur <= end:
                for i in range(7):
                    day = cur + timedelta(days=i)
                    if day >= start and day <= end:
                        all_dates.append(day)
                cur = cur + timedelta(days=7)
            all_dates = sorted(list(set(all_dates)))
        else:
            all_dates = [start + timedelta(days=i) for i in range((end - start).days + 1)]

        # print(f"DEBUG: processing {len(all_dates)} dates from {start} to {end}")

        # helpers
        def fetch_latest_for_date(table, s_id, the_date):
            cursor.execute(f'''
                SELECT file_data
                FROM {table}
                WHERE state_id = %s AND upload_date = %s
                ORDER BY revision_no DESC
                LIMIT 1;
            ''', (s_id, the_date))
            r = cursor.fetchone()
            return r[0] if r and r[0] else None

        def extract_column_values(file_data, col_index_in_row, start_idx):
            """
            Try strategies to extract numeric values for a column. Return list of floats (may be empty).
            Extensive prints are added at call sites (above) — this function keeps compact.
            """
            vals = []
            if not file_data:
                return vals

            # If first element is list/tuple
            try:
                first = file_data[0]
            except Exception:
                return vals

            # day-list-mode: each element is 96-length list
            if isinstance(first, (list, tuple)) and all(isinstance(x, (list, tuple)) and len(x) == 96 for x in file_data):
                # flatten all numeric values from all day-blocks
                for day_block in file_data:
                    for v in day_block:
                        try:
                            vv = float(v) if v not in [None, '', ' '] else 0.0
                        except Exception:
                            vv = 0.0
                        vals.append(vv)
                return vals

            # row-mode: file_data is list-of-rows and we index into each row using col_index_in_row
            if isinstance(first, (list, tuple)):
                try:
                    for r in file_data:
                        try:
                            raw = r[col_index_in_row]
                        except Exception:
                            raw = None
                        try:
                            vv = float(raw) if raw not in [None, '', ' '] else 0.0
                        except Exception:
                            vv = 0.0
                        vals.append(vv)
                    return vals
                except Exception:
                    pass

            # flat list: if len == 96 treat as single block
            if not isinstance(first, (list, tuple)):
                if len(file_data) == 96:
                    try:
                        for v in file_data:
                            vv = float(v) if v not in [None, '', ' '] else 0.0
                            vals.append(vv)
                        return vals
                    except Exception:
                        pass

                # chunked flat list: multiple of 96
                if len(file_data) % 96 == 0:
                    try:
                        for i in range(0, len(file_data), 96):
                            chunk = file_data[i:i+96]
                            for v in chunk:
                                try:
                                    vv = float(v) if v not in [None, '', ' '] else 0.0
                                except Exception:
                                    vv = 0.0
                                vals.append(vv)
                        return vals
                    except Exception:
                        pass

            return vals

        # MAIN loop
        result = {}

        for d in all_dates:
            d_str = d.strftime("%Y-%m-%d")
            # print(f"\n--- DATE: {d_str} ---")
            row_status = {col: "wrong" for col in forecast_columns}

            if data_type in ["dayahead", "intraday"]:
                if state_id is None:
                    print(f"INFO: no state selected — skipping date {d_str}")
                    result[d_str] = row_status
                    continue

                file_data = fetch_latest_for_date(table_name, state_id, d)
                if not file_data:
                    print(f"INFO: no file_data found for state={state_id} date={d_str} → all columns WRONG")
                    result[d_str] = row_status
                    continue

                # print summary of file_data type/length
                try:
                    sample_type = type(file_data[0]).__name__
                    length = len(file_data)
                except Exception:
                    sample_type = type(file_data).__name__
                    try:
                        length = len(file_data)
                    except Exception:
                        length = "unknown"
                # print(f"DEBUG: file_data type sample='{sample_type}', length={length}")

                for idx, col in enumerate(forecast_columns):
                    col_index_in_row = start_index + idx
                    values = extract_column_values(file_data, col_index_in_row, start_index)
                    non_zero_count = sum(1 for v in values if v != 0)
                    total_count = len(values)

                    # decision logic
                    if values and non_zero_count > 0:
                        row_status[col] = "tick"
                        decision = "tick (has non-zero)"
                    else:
                        row_status[col] = "wrong"
                        if not values:
                            decision = "wrong (no values extracted)"
                        else:
                            decision = "wrong (all zero or blank)"

                    # print concise trace for this column
                    # print(f"COL: '{col}' | extracted_count={total_count} | non_zero_count={non_zero_count} | sample={values[:6]} | decision={decision}")

                # SPECIAL: if you want to apply combined rule across DAY_AHEAD_EXTRA_COLUMNS (all zero => all wrong),
                # the per-column logic above already marks each extra column wrong if its values were all zero.
                result[d_str] = row_status

            else:
                # Week/Month ahead - find covering record(s)
                # Query for revisions covering this date (state required)
                if state_id is None:
                    # print(f"INFO: no state selected for week/month processing at {d_str}")
                    result[d_str] = row_status
                    continue

                cursor.execute(f'''
                    WITH MaxRevisions AS (
                        SELECT state_id, from_date, to_date, MAX(revision_no) AS max_revision
                        FROM {table_name}
                        WHERE (from_date <= %s::date AND to_date >= %s::date)
                        AND state_id = %s
                        GROUP BY state_id, from_date, to_date
                    )
                    SELECT t.from_date, t.to_date, t.file_data
                    FROM {table_name} t
                    INNER JOIN MaxRevisions mr
                      ON t.state_id = mr.state_id
                     AND t.from_date = mr.from_date
                     AND t.to_date = mr.to_date
                     AND t.revision_no = mr.max_revision
                    WHERE t.state_id = %s
                    ORDER BY t.from_date;
                ''', (d, d, state_id, state_id))
                rows = cursor.fetchall()

                if not rows:
                    # print(f"INFO: no week/month record covering date {d_str} for state {state_id}")
                    result[d_str] = row_status
                    continue

                rec = rows[0]
                from_d, to_d, rec_file_data = rec
                # print(f"DEBUG: Found covering record from {from_d} to {to_d} (state {state_id}), file_data present: {'yes' if rec_file_data else 'no'}")

                if not rec_file_data:
                    # print("INFO: record file_data empty → all wrong")
                    result[d_str] = row_status
                    continue

                # For each forecast column extract values and decide
                for idx, col in enumerate(forecast_columns):
                    col_index_in_row = start_index + idx
                    values = extract_column_values(rec_file_data, col_index_in_row, start_index)
                    non_zero_count = sum(1 for v in values if v != 0)
                    total_count = len(values)

                    if values and non_zero_count > 0:
                        row_status[col] = "tick"
                        decision = "tick (has non-zero)"
                    else:
                        row_status[col] = "wrong"
                        if not values:
                            decision = "wrong (no values extracted)"
                        else:
                            decision = "wrong (all zero or blank)"

                    # print(f"COL: '{col}' | extracted_count={total_count} | non_zero_count={non_zero_count} | sample={values[:6]} | decision={decision}")

                result[d_str] = row_status

        cursor.close()
        conn.close()

        # ensure ordering
        sorted_dates = sorted(result.keys())
        sorted_result = {d: {col: result[d].get(col, "wrong") for col in forecast_columns} for d in sorted_dates}

        response = {"status": "success", "columns": forecast_columns, "data": sorted_result}
        print("DEBUG: finished processing getDataStatus")
        return jsonify(response), HTTPStatus.OK

    except Exception as e:
        import traceback
        traceback.print_exc()
        # print("ERROR in getDataStatus:", e)
        return jsonify({"error": str(e)}), HTTPStatus.INTERNAL_SERVER_ERROR




@app.route('/api/getstatewisestatus', methods=['POST'])
@jwt_required(locations=["cookies"])
@token_required
# @role_required(allowed_roles=["user", "admin"], enforce_state=True)
@session_token_required
def getStateWiseStatus():
    try:
        import psycopg2.extras
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        params_outer = request.get_json().get("params", {})
        params = params_outer.get("params", params_outer)

        data_type = params.get("dataType")
        from_date = params.get("fromDate")
        to_date = params.get("toDate")

        if not all([data_type, from_date, to_date]):
            return jsonify({"error": "Missing required parameters"}), HTTPStatus.BAD_REQUEST

        # choose table + start_index + columns
        if data_type == "dayahead":
            table_name = "file_uploads"
            start_index = 2
            forecast_columns = BASE_FORECAST_COLUMNS + DAY_AHEAD_EXTRA_COLUMNS
            uses_from_to = False
        elif data_type == "weekahead":
            table_name = "week_ahead_file_uploads"
            start_index = 3
            forecast_columns = BASE_FORECAST_COLUMNS
            uses_from_to = True
        elif data_type == "monthahead":
            table_name = "month_ahead_file_uploads"
            start_index = 3
            forecast_columns = BASE_FORECAST_COLUMNS
            uses_from_to = True
        elif data_type == "intraday":
            table_name = "intraday_file_uploads"
            start_index = 2
            forecast_columns = BASE_FORECAST_COLUMNS
            uses_from_to = False
        else:
            return jsonify({"error": f"Invalid dataType {data_type}"}), HTTPStatus.BAD_REQUEST

        # parse request date range into Python date objects
        start = datetime.strptime(from_date, "%d/%m/%Y").date()
        end = datetime.strptime(to_date, "%d/%m/%Y").date()
        if end < start:
            return jsonify({"error": "toDate must be >= fromDate"}), HTTPStatus.BAD_REQUEST

        requested_dates = [(start + timedelta(days=i)) for i in range((end - start).days + 1)]
        # print(f"[DEBUG] Requested dates count: {len(requested_dates)} from {start} to {end}")

        # get distinct state ids present in the chosen table
        cursor.execute(f"SELECT DISTINCT state_id FROM {table_name}")
        state_rows = cursor.fetchall()
        state_ids = [r[0] for r in state_rows if r and r[0] is not None]
        # print(f"[DEBUG] Found state_ids in table {table_name}: {state_ids}")

        # build state_id -> state_name map if possible
        state_map = {}
        if state_ids:
            try:
                cursor.execute("SELECT state_id, state_name FROM states WHERE state_id = ANY(%s)", (state_ids,))
                for r in cursor.fetchall():
                    sid = r.get('state_id') if 'state_id' in r else r[0]
                    sname = r.get('state_name') or r.get('name') or (r[1] if len(r) > 1 else str(sid))
                    state_map[sid] = sname
            except Exception:
                cursor.execute("SELECT state_id, name FROM states WHERE state_id = ANY(%s)", (state_ids,))
                for r in cursor.fetchall():
                    sid = r.get('state_id') if 'state_id' in r else r[0]
                    sname = r.get('name') if 'name' in r else (r[1] if len(r) > 1 else str(sid))
                    state_map[sid] = sname
        # print(f"[DEBUG] state_map: {state_map}")

        # helpers
        def to_date_safe(val):
            if val is None:
                return None
            if isinstance(val, date):
                return val
            if isinstance(val, datetime):
                return val.date()
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(str(val), fmt).date()
                except Exception:
                    pass
            try:
                from dateutil.parser import parse
                return parse(str(val)).date()
            except Exception:
                return None

        def classify_raw(raw):
            if raw is None:
                return True, False, 0.0
            s = str(raw).strip()
            if s == "":
                return True, False, 0.0
            try:
                num = float(s)
                return False, True, num
            except Exception:
                return False, False, 0.0

        result = {}

        for sid in state_ids:
            sname = state_map.get(sid, str(sid))
            state_result = {col: "wrong" for col in forecast_columns}

            for col in forecast_columns:
                col_idx = forecast_columns.index(col)
                total_samples = 0
                blank_count = 0
                numeric_count = 0
                numeric_nonzero_count = 0
                non_numeric_count = 0

                if uses_from_to:
                    # WEEK / MONTH: query rows whose date ranges overlap the requested range
                    cursor.execute(f'''
                        SELECT t.from_date, t.to_date, t.file_data
                        FROM {table_name} t
                        WHERE t.state_id = %s
                          AND t.from_date <= %s
                          AND t.to_date >= %s
                        ORDER BY t.from_date;
                    ''', (sid, end, start))   # pass end as left-side bound, start as right-side bound
                    rows = cursor.fetchall()
                    # print(f"[DEBUG] state {sname} ({sid}) rows found: {len(rows)} for column '{col}'")

                    for rec in rows:
                        # rec may be DictRow or tuple
                        rec_from_raw = rec.get('from_date') if isinstance(rec, dict) else rec[0]
                        rec_to_raw = rec.get('to_date') if isinstance(rec, dict) else rec[1]
                        rec_from = to_date_safe(rec_from_raw)
                        rec_to = to_date_safe(rec_to_raw)
                        file_data = rec.get('file_data') if isinstance(rec, dict) else rec[2]

                        if rec_from is None or rec_to is None or not file_data:
                            # print(f"[DEBUG] skipping rec for state {sid} due to missing from/to/file_data: {rec_from},{rec_to},{bool(file_data)}")
                            continue

                        overlap_start = max(rec_from, start)
                        overlap_end = min(rec_to, end)
                        if overlap_start > overlap_end:
                            continue

                        # for each overlapping day compute offset relative to rec_from
                        for curr_day_offset in range((overlap_start - rec_from).days, (overlap_end - rec_from).days + 1):
                            forecast_index = curr_day_offset * 96
                            try:
                                day_block = file_data[forecast_index: forecast_index + 96]
                            except Exception as ex:
                                # print(f"[DEBUG] slicing error state {sid} offset {curr_day_offset}: {ex}")
                                continue

                            if not day_block:
                                continue

                            first_el = day_block[0]
                            if isinstance(first_el, (list, tuple, dict)):
                                for row_item in day_block:
                                    raw_val = None
                                    try:
                                        if isinstance(row_item, dict):
                                            vals = list(row_item.values())
                                            raw_val = vals[start_index + col_idx] if len(vals) > start_index + col_idx else None
                                        else:
                                            raw_val = row_item[start_index + col_idx] if len(row_item) > start_index + col_idx else None
                                    except Exception:
                                        raw_val = None
                                    is_blank, is_numeric, numeric_val = classify_raw(raw_val)
                                    total_samples += 1
                                    if is_blank:
                                        blank_count += 1
                                    elif is_numeric:
                                        numeric_count += 1
                                        if numeric_val != 0.0:
                                            numeric_nonzero_count += 1
                                    else:
                                        non_numeric_count += 1
                            else:
                                for raw_val in day_block:
                                    is_blank, is_numeric, numeric_val = classify_raw(raw_val)
                                    total_samples += 1
                                    if is_blank:
                                        blank_count += 1
                                    elif is_numeric:
                                        numeric_count += 1
                                        if numeric_val != 0.0:
                                            numeric_nonzero_count += 1
                                    else:
                                        non_numeric_count += 1

                else:
                    # DAYAHEAD / INTRADAY: fetch each upload_date row
                    for d in requested_dates:
                        cursor.execute(f'''
                            SELECT file_data
                            FROM {table_name}
                            WHERE state_id=%s AND upload_date=%s
                            ORDER BY revision_no DESC
                            LIMIT 1;
                        ''', (sid, d))
                        row = cursor.fetchone()
                        if not row or not row[0]:
                            continue
                        file_data = row[0]
                        if not file_data:
                            continue

                        for row_item in file_data:
                            raw_val = None
                            try:
                                if isinstance(row_item, dict):
                                    vals = list(row_item.values())
                                    raw_val = vals[start_index + col_idx] if len(vals) > start_index + col_idx else None
                                else:
                                    raw_val = row_item[start_index + col_idx] if len(row_item) > start_index + col_idx else None
                            except Exception:
                                raw_val = None

                            is_blank, is_numeric, numeric_val = classify_raw(raw_val)
                            total_samples += 1
                            if is_blank:
                                blank_count += 1
                            elif is_numeric:
                                numeric_count += 1
                                if numeric_val != 0.0:
                                    numeric_nonzero_count += 1
                            else:
                                non_numeric_count += 1

                # decide tick/wrong
                has_nonzero = numeric_nonzero_count > 0
                state_result[col] = "tick" if has_nonzero else "wrong"

                # print(f"[DEBUG] State '{sname}' (id={sid}) col='{col}': samples={total_samples}, blanks={blank_count}, numeric={numeric_count}, nonzero={numeric_nonzero_count}, non_numeric={non_numeric_count} => {'tick' if has_nonzero else 'wrong'}")

            result[sname] = state_result

        cursor.close()
        conn.close()

        sorted_result = {k: result[k] for k in sorted(result.keys())}

        return jsonify({
            "status": "success",
            "columns": ["STATE"] + forecast_columns,
            "data": sorted_result
        }), HTTPStatus.OK

    except Exception as e:
        try:
            log_error("statewisestatus", e)
        except Exception:
            pass
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), HTTPStatus.INTERNAL_SERVER_ERROR


@app.route('/api/breakupparameters', methods=['POST'])
@jwt_required(locations=["cookies"])
@token_required
@role_required(allowed_roles=["user", "admin"], enforce_state=True)
@session_token_required
def breakupParamters():
    try:
        import psycopg2.extras
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        params = request.get_json()
        state_id = params['params']['state']
        type_id = int(params['params']['indexNumber'])
        from_date = datetime.strptime(params['params']['fromDate'], "%d/%m/%Y")
        to_date = datetime.strptime(params['params']['toDate'], "%d/%m/%Y")

        #  **Generate all timestamps for Forecast & Actual (15-min interval)**
        all_timestamps = []
        current_time = from_date
        while current_time <= to_date.replace(hour=23, minute=45):
            all_timestamps.append(current_time.strftime("%Y-%m-%d %H:%M"))
            current_time += timedelta(minutes=15)

        #  **Generate all daily timestamps for MAPE**
        all_dates = []
        current_date = from_date.date()
        while current_date <= to_date.date():
            all_dates.append(datetime.combine(current_date, datetime.min.time()).strftime("%Y-%m-%d"))
            current_date += timedelta(days=1)

        #  **Fetching Forecast & Actual Data**
        query_forecast = '''
            SELECT timestamp, forecast_type, forecasted_demand, actual_demand
            FROM calculated_mape
            WHERE state_id = %s AND type_id = %s AND timestamp BETWEEN %s AND %s
            ORDER BY timestamp;
        '''
        cursor.execute(query_forecast, (state_id, type_id, from_date, to_date.replace(hour=23, minute=45)))
        forecast_results = cursor.fetchall()

        #  **Fetching MAPE Data**
        query_mape = '''
            SELECT timestamp::date AS date, forecast_type, mape
            FROM calculated_mape
            WHERE state_id = %s AND type_id = %s AND timestamp BETWEEN %s AND %s
            ORDER BY date;
        '''
        cursor.execute(query_mape, (state_id, type_id, from_date, to_date))
        mape_results = cursor.fetchall()

        #  **Initialize Forecast & Actual Data**
        full_data_dict = {
            ts: {"actual": None, "day_ahead": None, "week_ahead": None, "month_ahead": None} for ts in all_timestamps
        }

        #  **Process Forecast & Actual Data**
        for row in forecast_results:
            timestamp = row["timestamp"].strftime("%Y-%m-%d %H:%M")
            if row["forecast_type"] == "day_ahead":
                full_data_dict[timestamp]["day_ahead"] = int(row["forecasted_demand"]) if row["forecasted_demand"] is not None else None
            elif row["forecast_type"] == "week_ahead":
                full_data_dict[timestamp]["week_ahead"] = int(row["forecasted_demand"]) if row["forecasted_demand"] is not None else None
            elif row["forecast_type"] == "month_ahead":
                full_data_dict[timestamp]["month_ahead"] = int(row["forecasted_demand"]) if row["forecasted_demand"] is not None else None
            if row["actual_demand"] is not None:
                full_data_dict[timestamp]["actual"] = int(row["actual_demand"])

        #  **Initialize MAPE Data with None**
        mape_data = {
            "day_ahead_mape": {date: None for date in all_dates},
            "week_ahead_mape": {date: None for date in all_dates},
            "month_ahead_mape": {date: None for date in all_dates}
        }

        #  **Process MAPE Data**
        for row in mape_results:
            timestamp = row["date"].strftime("%Y-%m-%d")
            if row["forecast_type"] == "day_ahead":
                mape_data["day_ahead_mape"][timestamp] = float(row["mape"]) if row["mape"] is not None else None
            elif row["forecast_type"] == "week_ahead":
                mape_data["week_ahead_mape"][timestamp] = float(row["mape"]) if row["mape"] is not None else None
            elif row["forecast_type"] == "month_ahead":
                mape_data["month_ahead_mape"][timestamp] = float(row["mape"]) if row["mape"] is not None else None

        cursor.close()
        conn.close()

        #  **Format Response to Match Required Structure**
        response_data = {
            "status": "success",
            "title": f"MAPE and Comparison between Actual, Intraday, Day Ahead, Week Ahead, and Month Ahead data from {from_date.strftime('%d/%m/%Y')} to {to_date.strftime('%d/%m/%Y')} for SR",
            
            "data": [
                {"type": "line", "yAxis": 0,"name": "Day Ahead Forecast", "data": [{"x": ts, "y": full_data_dict[ts]["day_ahead"]} for ts in all_timestamps]},
                {"type": "line", "yAxis": 0,"name": "Week Ahead Forecast", "data": [{"x": ts, "y": full_data_dict[ts]["week_ahead"]} for ts in all_timestamps]},
                {"type": "line", "yAxis": 0,"name": "Month Ahead Forecast", "data": [{"x": ts, "y": full_data_dict[ts]["month_ahead"]} for ts in all_timestamps]},
                {"type": "line", "yAxis": 0,"name": "Actual", "data": [{"x": ts, "y": full_data_dict[ts]["actual"]} for ts in all_timestamps]},
                { "type": "column","name": "Day Ahead MAPE", "yAxis": 1, "data": [{"x": date, "y": mape_data["day_ahead_mape"][date]} for date in all_dates]},
                { "type": "column","name": "Week Ahead MAPE",  "yAxis": 1,"data": [{"x": date, "y": mape_data["week_ahead_mape"][date]} for date in all_dates]},
                { "type": "column","name": "Month Ahead MAPE", "yAxis": 1, "data": [{"x": date, "y": mape_data["month_ahead_mape"][date]} for date in all_dates]}
                
            ],
                
            
            
        }

        return jsonify(response_data)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify(status="failure", message=str(e))

    

@app.route('/api/dayrangestatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_day_range_status")
def dayRangeStatus():
    try:
        # Retrieve date range from request
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        print(params)

        jwt_data = get_jwt()

        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y')
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y')

        # SQL query to fetch data within the specified date range
        sql_query = """
                    WITH min_revision AS (
                    SELECT 
                        state_id, 
                        upload_date, 
                        MIN(revision_no) AS min_revision_no
                    FROM 
                        file_uploads
                    GROUP BY 
                        state_id, 
                        upload_date
                )
                SELECT 
                    states.state_name,
                    COALESCE(file_uploads.upload_date, %s) AS upload_date,
                    COALESCE(min_uploads.upload_time, NULL) AS upload_time,
                    CASE
                        WHEN min_uploads.upload_time IS NULL THEN 2  -- Not Uploaded
                        WHEN min_uploads.upload_time < (file_uploads.upload_date - INTERVAL '1 day' + INTERVAL '10 hours') THEN 1  -- Proper Upload before 10 AM on the previous day
                        ELSE 0  -- Late Upload
                    END AS upload_status_code,
                    COUNT(file_uploads.state_id) AS upload_count
                FROM 
                    states
                LEFT JOIN 
                    file_uploads ON states.state_id = file_uploads.state_id
                    AND file_uploads.upload_date BETWEEN %s AND %s
                LEFT JOIN (
                    SELECT 
                        file_uploads.state_id, 
                        file_uploads.upload_date, 
                        file_uploads.upload_time
                    FROM 
                        file_uploads
                    INNER JOIN min_revision ON 
                        file_uploads.state_id = min_revision.state_id
                        AND file_uploads.upload_date = min_revision.upload_date
                        AND file_uploads.revision_no = min_revision.min_revision_no
                ) AS min_uploads ON 
                    file_uploads.state_id = min_uploads.state_id 
                    AND file_uploads.upload_date = min_uploads.upload_date
                WHERE 
                    states.state_id IN (1,2,3,4,5,7)
                GROUP BY 
                    states.state_name, 
                    file_uploads.upload_date,
                    min_uploads.upload_time  
                ORDER BY
                    states.state_name,
                    upload_date DESC;
        """
        cursor.execute(sql_query, (end_date, start_date, end_date))
        results = cursor.fetchall()

        day_data = []
        date_range = [end_date - timedelta(days=i) for i in range((end_date - start_date).days + 1)]
        state_names = set(result[0] for result in results)

        for state_name in state_names:
            state_data = {"name": state_name, "data": []}
            for date in date_range:
                date_str = date.strftime('%Y-%m-%d')
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1].strftime("%Y-%m-%d") == date_str:
                        upload_count = result[4]
                        upload_time = result[2].strftime("%Y-%m-%d %H:%M:%S") if result[2] is not None else None
                        upload_status_code = result[3]
                        found_result = True
                        break
                if not found_result:
                    upload_count = 0
                    upload_time = None
                    upload_status_code = 2  # Not Uploaded
                state_data["data"].append({
                    'x': date_str, 
                    'y': upload_status_code, 
                    'upload_time': upload_time, 
                    'upload_count': upload_count
                })
            day_data.append(state_data)

        day_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }

        return jsonify(day=day_data, status="success")

    except Exception as e:
        log_error("dayrangestatus", e)
        cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact SRLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact SRLDC IT!", status="failure")





@app.route('/api/weekrangestatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_week_range_status")
def weekRangeStatus():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        jwt_data = get_jwt()

        # Retrieve date range from request
        params = request.get_json()
        custom_start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y').date()
        custom_end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y').date()

        # Adjust start date to the start of the week (Monday)
        if custom_start_date.weekday() != 0:  # If the start date is not a Monday
            custom_start_date = custom_start_date - timedelta(days=custom_start_date.weekday())  # Shift to the previous Monday

        # Adjust the end date to include the entire last week if it falls mid-week
        if custom_end_date.weekday() != 6:  # If the end date is not a Sunday
            custom_end_date = custom_end_date + timedelta(days=(6 - custom_end_date.weekday()))  # Extend to the next Sunday

        # Generate the week ranges (from Monday to Sunday)
        week_range = []
        week_start = custom_start_date
        while week_start <= custom_end_date:
            week_range.append(week_start)
            week_start += timedelta(weeks=1)

        # SQL query for fetching data across weeks
        sql_query = """
            WITH min_revision AS (
                SELECT 
                    state_id, 
                    from_date, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    week_ahead_file_uploads
                WHERE
                    from_date BETWEEN %s AND %s
                GROUP BY 
                    state_id, 
                    from_date
            ), min_uploads AS (
                SELECT 
                    fu.state_id, 
                    fu.from_date, 
                    fu.upload_time,
                    fu.revision_no
                FROM 
                    week_ahead_file_uploads fu
                INNER JOIN min_revision mr ON
                    fu.state_id = mr.state_id
                    AND fu.from_date = mr.from_date
                    AND fu.revision_no = mr.min_revision_no
            )
            SELECT 
                states.state_name,
                mu.from_date AS week_start_date,
                mu.upload_time,
                CASE
                    WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                    WHEN mu.upload_time < DATE_TRUNC('week', mu.from_date - INTERVAL '1 week')  + INTERVAL '1 day' THEN 1  -- Uploaded before the Monday of the previous week
                    ELSE 0  -- Late Upload
                END AS upload_status_code
            FROM 
                states
            LEFT JOIN 
                min_uploads mu ON states.state_id = mu.state_id
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                mu.from_date,
                mu.upload_time  
            ORDER BY
                states.state_name,
                mu.from_date DESC;
        """
        
        cursor.execute(sql_query, (custom_start_date.strftime('%Y-%m-%d'), custom_end_date.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Fetch state names
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Initialize data structure for states
        week_data = [{"name": state_name, "data": []} for state_name in all_states]

        # Process each state and week range
        for state in week_data:
            for week_start in week_range:
                week_end = week_start + timedelta(days=6)  # End of the week
                # Filter the results for the current state and week
                upload_info = next((item for item in results if item[0] == state["name"] and item[1] is not None and week_start <= item[1] <= week_end), None)

                # Determine upload status for the week
                if upload_info:
                    upload_status_code = upload_info[3]
                    upload_time = upload_info[2].strftime('%Y-%m-%d %H:%M:%S') if upload_info[2] else None
                    upload_count = 1  # At least one upload exists
                else:
                    upload_status_code = 2  # Not Uploaded
                    upload_time = None
                    upload_count = 0

                # Append data for the week
                state["data"].append({
                    "x": f"{week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')}",
                    "y": upload_status_code,
                    "upload_time": upload_time,
                    "upload_count": upload_count
                })

            # Sort the state's data in descending order of week (latest first)
            # print(state["data"])
            # print(jwt_data)
            state["data"] = sorted(state["data"], key=lambda x: x["x"], reverse=True)
        return jsonify(week=week_data, status="success")

    except Exception as e:
        log_error("weekrangestatus", e)
        cursor.close()
        return jsonify(message="There is a problem, please contact SRLDC IT!", status="failure")




@app.route('/api/monthrangestatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_month_range_status")
def monthRangeStatus():
    try:
        # Get the current date
        today = datetime.now()

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Retrieve date range from request

        # Retrieve date range from request
        params = request.get_json()
        print(params)


        jwt_data = get_jwt()

        # Assuming 'params' is a dictionary that includes 'fromDate' and 'toDate'
        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y')
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y')

        # Calculate the start of the first month in the custom range
        start_of_first_month = start_date.replace(day=1)

        # Calculate the end of the last month in the custom range
        _, last_day = monthrange(end_date.year, end_date.month)
        end_of_last_month = end_date.replace(day=last_day)

        # Execute SQL Query
        sql_query = """
        WITH min_revision AS (
            SELECT 
                state_id, 
                DATE_TRUNC('month', from_date) AS from_month, 
                MIN(revision_no) AS min_revision_no
            FROM 
                month_ahead_file_uploads
            WHERE
                from_date BETWEEN %s AND %s
            GROUP BY 
                state_id, 
                from_month
        ), min_uploads AS (
            SELECT 
                fu.state_id, 
                DATE_TRUNC('month', fu.from_date) AS from_month, 
                fu.upload_time,
                fu.revision_no
            FROM 
                month_ahead_file_uploads fu
            INNER JOIN min_revision mr ON
                fu.state_id = mr.state_id
                AND DATE_TRUNC('month', fu.from_date) = mr.from_month
                AND fu.revision_no = mr.min_revision_no
        )
        SELECT 
            states.state_name,
            COALESCE(mu.from_month, %s) AS month_start_date,
            COALESCE(mu.upload_time, NULL) AS upload_time,
            CASE
                WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                WHEN mu.upload_time < DATE_TRUNC('month', mu.from_month) - INTERVAL '1 month' + INTERVAL '5 day' THEN 1  -- Uploaded on time
                ELSE 0  -- Late Upload
            END AS upload_status_code,
            COUNT(mu.state_id) AS upload_count
        FROM 
            states
        LEFT JOIN 
            min_uploads mu ON states.state_id = mu.state_id
        WHERE 
            states.state_id IN (1,2,3,4,5,7)
        GROUP BY 
            states.state_name, 
            mu.from_month,
            mu.upload_time  
        ORDER BY
            states.state_name,
            mu.from_month DESC;
        """
        cursor.execute(sql_query, (start_of_first_month.strftime('%Y-%m-%d'), end_of_last_month.strftime('%Y-%m-%d'), end_of_last_month.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Process the results
        month_range = [start_of_first_month + timedelta(days=32 * i) for i in range((end_of_last_month.year - start_of_first_month.year) * 12 + end_of_last_month.month - start_of_first_month.month + 1)]
        month_range = [date.replace(day=1) for date in month_range]
        all_states = set(result[0] for result in results)

        month_data = []
        # Process results and sort by latest month first
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for month_start in sorted(month_range, reverse=True):
                _, last_day_of_month = monthrange(month_start.year, month_start.month)
                month_end = month_start.replace(day=last_day_of_month)
                month_range_str = f"{month_start.strftime('%Y-%m-%d')} to {month_end.strftime('%Y-%m-%d')}"

                upload_count = 0
                upload_status_code = 2  # Default to Not Uploaded
                upload_time = None
                found_result = False

                # Check each result for the current state and month
                for result in results:
                    result_state, result_month_start, result_upload_time, result_status, result_count = result
                    if result_state == state_name and result_month_start.strftime('%Y-%m-%d') == month_start.strftime('%Y-%m-%d'):
                        upload_count = result_count
                        upload_time = result_upload_time.strftime("%Y-%m-%d %H:%M:%S") if result_upload_time else None
                        upload_status_code = result_status
                        found_result = True
                        break

                if not found_result:
                    upload_count = 0
                    upload_status_code = 2  # Not Uploaded

                state_data["data"].append({
                    'x': month_range_str,
                    'y': upload_status_code,
                    'upload_time': upload_time,
                    'upload_count': upload_count
                })

            month_data.append(state_data)

        # # Prepare JSON for frontend
        # month_dates = {
        #     "start_date": start_of_first_month.strftime('%Y-%m-%d'),
        #     "end_date": end_of_last_month.strftime('%Y-%m-%d')
        # }

        month_data
        return jsonify(month=month_data, status="success")

    except Exception as e:
        log_error("monthrangestatus", e)
        cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact SRLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact SRLDC IT!")


@app.route('/api/yearrangestatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_year_range_status")
def yearRangeStatus():
    try:
        today = datetime.now()

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Retrieve date range from request
        params = request.get_json()
        jwt_data = get_jwt()

        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y')
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y')

        # Determine the current and next financial years based on start_date
        if start_date.month < 4:
            current_financial_year_start = datetime(start_date.year - 1, 4, 1)
        else:
            current_financial_year_start = datetime(start_date.year, 4, 1)

        current_financial_year_end = datetime(current_financial_year_start.year + 1, 3, 31)
        next_financial_year_start = datetime(current_financial_year_start.year + 1, 4, 1)
        next_financial_year_end = datetime(next_financial_year_start.year + 1, 3, 31)

        # Fetch data (keep SQL simple)
        sql_query = """
        SELECT 
            states.state_name,
            yafu.from_date,
            yafu.to_date,
            yafu.upload_time,
            COUNT(yafu.state_id) AS upload_count
        FROM 
            states
        LEFT JOIN 
            year_ahead_file_uploads yafu ON states.state_id = yafu.state_id
        WHERE 
            states.state_id IN (1,2,3,4,5,7)
            AND (yafu.from_date BETWEEN %s AND %s OR yafu.from_date IS NULL)
        GROUP BY 
            states.state_name, yafu.from_date, yafu.to_date, yafu.upload_time
        ORDER BY
            states.state_name, yafu.from_date DESC;
        """

        cursor.execute(
            sql_query,
            (current_financial_year_start.strftime('%Y-%m-%d'),
             next_financial_year_end.strftime('%Y-%m-%d'))
        )
        results = cursor.fetchall()

        # Fetch state names to ensure coverage of all required states
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Organize the fetched data by state and financial year
        year_data = []
        financial_years = [current_financial_year_start, next_financial_year_start]

        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for financial_year in financial_years:
                found = False
                for result in results:
                    result_state, from_date, to_date, upload_time, upload_count = result

                    # normalize dates
                    if isinstance(from_date, date) and not isinstance(from_date, datetime):
                        from_date = datetime.combine(from_date, datetime.min.time())
                    if isinstance(to_date, date) and not isinstance(to_date, datetime):
                        to_date = datetime.combine(to_date, datetime.min.time())

                    if result_state == state_name and from_date and from_date.date() == financial_year.date():
                        found = True

                        # === Apply business rules ===
                        if not upload_time:
                            upload_status = 2  # Not uploaded
                        else:
                            fy_start = from_date
                            sept30_preceding = datetime(fy_start.year - 1, 9, 30)

                            if upload_time > sept30_preceding:
                                upload_status = 0  # Late (after Sept 30 of preceding year)
                            elif upload_time <= from_date + relativedelta(months=+5, days=-1):
                                upload_status = 1  # On time (within 5 months of from_date)
                            else:
                                upload_status = 0  # Late
                        upload_time_formatted = upload_time.strftime("%Y-%m-%d %H:%M:%S") if upload_time else None
                        state_data["data"].append({
                            'x': f"{from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')}",
                            'y': upload_status,
                            'upload_time': upload_time_formatted,
                            'upload_count': upload_count
                        })
                        break
                if not found:
                    state_data["data"].append({
                        'x': f"{financial_year.strftime('%Y-%m-%d')} to {financial_year.replace(year=financial_year.year + 1, month=3, day=31).strftime('%Y-%m-%d')}",
                        'y': 2,  # Not Uploaded
                        'upload_time': None,
                        'upload_count': 0
                    })
            year_data.append(state_data)

        # Prepare JSON data for the frontend
        year_dates = {
            "start_date": current_financial_year_start.strftime('%Y-%m-%d'),
            "end_date": next_financial_year_end.strftime('%Y-%m-%d')
        }

        return jsonify(year=year_data, status="success")

    except Exception as e:
        log_error("yearrangestatus", e)
        cursor.close()
        return jsonify(message="There is a problem, please contact SRLDC IT!")


@app.route('/api/mailstatusrangestatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_mail_status")
def mailStatusRangeStatus():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()

        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y').date()
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y').date()

        # Query to get all sent entries within range
        sql_query = """
            SELECT forecast_type, from_date, to_date
            FROM forecast_mail_status
            WHERE forecast_type IN ('day', 'week', 'month')
              AND from_date <= %s
              AND to_date >= %s
        """
        cursor.execute(sql_query, (end_date, start_date))
        results = cursor.fetchall()  # [(forecast_type, from_date, to_date), ...]

        # Create date sets for each type
        sent_dates = {
            'day': set(),
            'week': set(),
            'month': set()
        }

        for f_type, from_date, to_date in results:
            for d in range((to_date - from_date).days + 1):
                sent_dates[f_type].add(from_date + timedelta(days=d))

        # Generate date range
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        mail_status_data = []
        for f_type in ['day', 'week', 'month']:
            type_data = {"name": f_type, "data": []}
            for date in date_range:
                status = 1 if date in sent_dates[f_type] else 0
                type_data["data"].append({
                    "x": date.strftime('%Y-%m-%d'),
                    "y": status
                })
            mail_status_data.append(type_data)

        return jsonify(mail_status=mail_status_data, status="success")

    except Exception as e:
        log_error("mailstatusrangestatus", e)
        cursor.close()
        return jsonify(message="There is a problem, please contact SRLDC IT!", status="failure")





@app.route('/api/intradayrangestatus', methods=['POST'])
@session_token_required
@jwt_required(locations=["cookies"])
@token_required
@rbac_required("view_intraday_range_status")
def intradayRangeStatus():
    try:
        # Retrieve date range from request
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        print(params)

        jwt_data = get_jwt()

        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y')
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y')




        sql_query = """
                                WITH min_revision AS (
                SELECT 
                    state_id, 
                    upload_date, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    intraday_file_uploads
                GROUP BY 
                    state_id, 
                    upload_date
            ),
            file_count AS (
                SELECT 
                    state_id, 
                    upload_date, 
                    COUNT(revision_no) AS revision_count,
                    MAX(CASE WHEN file_type = 'D' THEN 1 ELSE 0 END) AS has_file_type_D
                FROM 
                    intraday_file_uploads
                GROUP BY 
                    state_id, 
                    upload_date
            )
            SELECT 
                states.state_name,
                COALESCE(intraday_file_uploads.upload_date, %s) AS upload_date,
                CASE
                    WHEN file_count.revision_count = 1 AND file_count.has_file_type_D = 1 THEN 0  -- Single revision and file_type is 'D' → Not Uploaded
                    WHEN MIN(intraday_file_uploads.revision_no) IS NULL THEN 0  -- No Uploads → Not Uploaded
                    ELSE 1  -- Uploaded
                END AS upload_status_code
            FROM 
                states
            LEFT JOIN 
                intraday_file_uploads ON states.state_id = intraday_file_uploads.state_id
                AND intraday_file_uploads.upload_date BETWEEN %s AND %s
            LEFT JOIN 
                file_count ON states.state_id = file_count.state_id 
                AND intraday_file_uploads.upload_date = file_count.upload_date
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                intraday_file_uploads.upload_date, 
                file_count.revision_count, 
                file_count.has_file_type_D
            ORDER BY
                states.state_name,
                upload_date DESC;

        """
        cursor.execute(sql_query, (end_date, start_date, end_date))
        results = cursor.fetchall()

        intraday_data = []
        date_range = [end_date - timedelta(days=i) for i in range(30)]
        state_names = set(result[0] for result in results)

        for state_name in state_names:
            state_data = {"name": state_name, "data": []}
            for date in date_range:
                date_str = date.strftime('%Y-%m-%d')
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1].strftime("%Y-%m-%d") == date_str:
                        upload_status_code = result[2]
                        found_result = True
                        break
                if not found_result:
                    upload_status_code = 0  # Not Uploaded
                state_data["data"].append({
                    'x': date_str, 
                    'y': upload_status_code
                })
            intraday_data.append(state_data)

        intraday_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }


        

        return jsonify(intraday=intraday_data, status="success")


    except Exception as e:
        log_error("intradayrangestatus", e)
        cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact SRLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact SRLDC IT!", status="failure")






    
# Special /data endpoint only accessible by script token
@app.route('/api/data', methods=['POST'])
@rbac_required("view_forecast_data")
@jwt_required(locations=["cookies"])
# @role_required(allowed_roles=["user", "admin"], enforce_state=True)
@token_required
def get_data():
    data = request.get_json()
    state_id = data.get('state')
    input_from_date = data.get('from_date')
    input_to_date = data.get('to_date')

    jwt_data = get_jwt()

    conn = psycopg2.connect(
    database="demand_forecast_states", user='prasadbabu', 
    password='BabuPrasad#123', host='10.0.100.79', port='5432'
    )

    cursor = conn.cursor()

    print(f"State ID: {state_id}, From Date: {input_from_date}, To Date: {input_to_date}")

    if not (state_id and input_from_date and input_to_date):
        return jsonify({'error': 'Missing data for state_id, from_date, or to_date'}), 400

    try:


        # Updated SQL query to filter by state_id, from_date, and to_date
        query = """
        SELECT
          state_id,
          from_date,
          to_date,
          revision_no,
          upload_time,
          file_data
        FROM (
          SELECT
            state_id,
            from_date,
            to_date,
            revision_no,
            upload_time,
            file_data,
            ROW_NUMBER() OVER (PARTITION BY state_id, from_date, to_date ORDER BY upload_time DESC) as rn
          FROM
            week_ahead_file_uploads
          WHERE state_id = %s AND from_date >= %s AND to_date <= %s
        ) sub
        WHERE sub.rn = 1;
        """

        cursor.execute(query, (state_id, input_from_date, input_to_date))
        records = cursor.fetchall()

        df_list = []

        results = []
        columns = ['Date', 'Block', 'Period', 'Forcasted Demand_MW (A)', 'From its own Sources Excl. Renewable_THERMAL_MW', 'From its own Sources Excl. Renewable_GAS_MW','From its own Sources Excl. Renewable_HYDRO_MW', 'From its own Sources Excl. Renewable_TOTAL (B)_MW', 'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 'From Renewable Sources_Other RES (biomass)_MW', 'From Renewable Sources_TOTAL (C)_MW', 'From ISGS & Other LTA & MTOA (D)_MW', 'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 'Total Availability  (F)= (B+C+D+E)_MW', 'Gap between Demand & Availability (G) = (A)-(F)  Surplus(-) / Deficit (+)_MW' , 'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 'Proposed Procurement_Through Power Exchange (I)_MW', 'Shortages after day ahead procurement from market (J) =(G)-(H+I)  Surplus(-) / Deficit (+)_MW', 'Relief through planned restrictions/ rostering/ power cuts (K)_MW','Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 'Reactive Power Forecast_MVar' ]

        for record in records:
            from_date = record[1]  # From Date
            to_date = record[2]    # To Date
            
            # Generate the Date column based on from_date and to_date
            date_range = generate_date_range(from_date, to_date)

            # Ensure the generated date list has the correct length to match the rows
            if len(date_range) != len(record[5]):
                raise ValueError(f"Generated date range length {len(date_range)} does not match file data length {len(record[5])}")

            # Replace "Invalid Date" with the actual date range
            file_data = record[5]
            for i, row in enumerate(file_data):
                row[0] = date_range[i]  # Assign the correct date to the first column
            
            result = {
                "State ID": record[0],
                "From Date": from_date.isoformat(),
                "To Date": to_date.isoformat(),
                "Revision No": record[3],
                "Upload Time": record[4].isoformat(),
                "File Data": file_data
            }

            # Create DataFrame from the updated File Data
            df_file_data = pd.DataFrame(result["File Data"], columns=columns)
            df_list.append(df_file_data)

        if not df_list:
            return jsonify({'error': 'No data present for this period'}), 500


        final_df = pd.concat(df_list, ignore_index=True)
        final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst = True)
        df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format

        cursor.close()
        conn.close()

        return jsonify(df_json)

    except Exception as e:
        print(f"Database connection or SQL execution error: {e}")
        log_error("data query", e)
        return jsonify({'error': 'Database query failed'}), 500



################ CONSOLIDATED DATA

@app.route("/api/fetchforecastdata", methods=['POST'])
@rbac_required("view_forecast_data")
@jwt_required(locations=["cookies"])
# @role_required(allowed_roles=["user", "admin"], enforce_state=True)
@token_required
def get_forecast_data():
    try:
        data = request.get_json()
        state_id = data.get('state')
        input_from_date = data.get('from_date')
        input_to_date = data.get('to_date')
        data_type = data.get('data_type')

        jwt_data = get_jwt()

        conn = psycopg2.connect(    
        database="demand_forecast_states", user='prasadbabu', 
        password='BabuPrasad#123', host='10.0.100.79', port='5432'
        )

        cursor = conn.cursor()

        if not (state_id and input_from_date and input_to_date and data_type):
            return jsonify({'error': 'Missing data for state_id, from_date, or to_date or type'}), 400


        if data_type == 'day':
            # SQL query to filter by state_id and upload_date range
            query = """
                SELECT
                state_id,
                upload_date,
                revision_no,
                upload_time,
                file_data
                FROM (
                SELECT
                    state_id,
                    upload_date,
                    revision_no,
                    upload_time,
                    file_data,
                    ROW_NUMBER() OVER (PARTITION BY state_id, upload_date ORDER BY upload_time DESC) as rn
                FROM
                    file_uploads  -- Table without from_date and to_date, using upload_date
                WHERE state_id = %s AND upload_date BETWEEN %s AND %s
                ) sub
                WHERE sub.rn = 1;
            """

            # Execute the query with state_id, input_from_date, and input_to_date
            cursor.execute(query, (state_id, input_from_date, input_to_date))
            records = cursor.fetchall()

            df_list = []

            results = []
            columns = ['Date', 'Block', 'Period', 'Forecasted Demand_MW (A)', 
                    'From its own Sources Excl. Renewable_THERMAL_MW', 
                    'From its own Sources Excl. Renewable_GAS_MW',
                    'From its own Sources Excl. Renewable_HYDRO_MW', 
                    'From its own Sources Excl. Renewable_TOTAL (B)_MW', 
                    'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 
                    'From Renewable Sources_Other RES (biomass)_MW', 
                    'From Renewable Sources_TOTAL (C)_MW', 
                    'From ISGS & Other LTA & MTOA (D)_MW', 
                    'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 
                    'Total Availability  (F)= (B+C+D+E)_MW', 
                    'Gap between Demand & Availability (G) = (A)-(F) Surplus(-) / Deficit (+)_MW', 
                    'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 
                    'Proposed Procurement_Through Power Exchange (I)_MW', 
                    'Shortages after day ahead procurement from market (J) =(G)-(H+I) Surplus(-) / Deficit (+)_MW', 
                    'Relief through planned restrictions/ rostering/ power cuts (K)_MW',
                    'Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 
                    'Reactive Power Forecast_MVar']

            for record in records:
                upload_date = record[1]  # Upload Date

                # Generate a list of the same date to fill the 'Date' column for all 96 rows
                date_range = [upload_date] * 96  # Assuming 96 rows per date as per your data

                # Get the file_data from the record
                file_data = record[4]

                # Ensure the correct date is added to each row's first element
                for i, row in enumerate(file_data):
                    row.insert(0, date_range[i])  # Insert the date at the beginning of each row

                result = {
                    "State ID": record[0],
                    "Upload Date": upload_date.isoformat(),
                    "Revision No": record[2],
                    "Upload Time": record[3].isoformat(),
                    "File Data": file_data
                }

                # Now that date has been added to file_data, create the DataFrame
                df_file_data = pd.DataFrame(result["File Data"], columns=columns)
                df_list.append(df_file_data)

            if not df_list:
                return jsonify({'error': 'No data present for this period'}), 500

            # Concatenate all dataframes into one final dataframe
            final_df = pd.concat(df_list, ignore_index=True)
            
            # Convert 'Date' column to datetime format
            final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst=True)

            # Convert the final DataFrame to JSON
            df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format

            cursor.close()
            conn.close()


            
            return jsonify(df_json)

        elif data_type == 'week':
            # Updated SQL query to filter by state_id, from_date, and to_date
            query = """
            SELECT
            state_id,
            from_date,
            to_date,
            revision_no,
            upload_time,
            file_data
            FROM (
            SELECT
                state_id,
                from_date,
                to_date,
                revision_no,
                upload_time,
                file_data,
                ROW_NUMBER() OVER (PARTITION BY state_id, from_date, to_date ORDER BY upload_time DESC) as rn
            FROM
                week_ahead_file_uploads
            WHERE state_id = %s AND from_date >= %s AND to_date <= %s
            ) sub
            WHERE sub.rn = 1;
            """

            cursor.execute(query, (state_id, input_from_date, input_to_date))
            records = cursor.fetchall()

            df_list = []

            results = []
            columns = ['Date', 'Block', 'Period', 'Forcasted Demand_MW (A)', 'From its own Sources Excl. Renewable_THERMAL_MW', 'From its own Sources Excl. Renewable_GAS_MW','From its own Sources Excl. Renewable_HYDRO_MW', 'From its own Sources Excl. Renewable_TOTAL (B)_MW', 'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 'From Renewable Sources_Other RES (biomass)_MW', 'From Renewable Sources_TOTAL (C)_MW', 'From ISGS & Other LTA & MTOA (D)_MW', 'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 'Total Availability  (F)= (B+C+D+E)_MW', 'Gap between Demand & Availability (G) = (A)-(F)  Surplus(-) / Deficit (+)_MW' , 'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 'Proposed Procurement_Through Power Exchange (I)_MW', 'Shortages after day ahead procurement from market (J) =(G)-(H+I)  Surplus(-) / Deficit (+)_MW', 'Relief through planned restrictions/ rostering/ power cuts (K)_MW','Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 'Reactive Power Forecast_MVar' ]

            for record in records:
                from_date = record[1]  # From Date
                to_date = record[2]    # To Date
                
                # Generate the Date column based on from_date and to_date
                date_range = generate_date_range(from_date, to_date)

                # Ensure the generated date list has the correct length to match the rows
                if len(date_range) != len(record[5]):
                    raise ValueError(f"Generated date range length {len(date_range)} does not match file data length {len(record[5])}")

                # Replace "Invalid Date" with the actual date range
                file_data = record[5]
                for i, row in enumerate(file_data):
                    row[0] = date_range[i]  # Assign the correct date to the first column
                
                result = {
                    "State ID": record[0],
                    "From Date": from_date.isoformat(),
                    "To Date": to_date.isoformat(),
                    "Revision No": record[3],
                    "Upload Time": record[4].isoformat(),
                    "File Data": file_data
                }

                # Create DataFrame from the updated File Data
                df_file_data = pd.DataFrame(result["File Data"], columns=columns)
                df_list.append(df_file_data)

            if not df_list:
                return jsonify({'error': 'No data present for this period'}), 500


            final_df = pd.concat(df_list, ignore_index=True)
            final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst = True)
            df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format

            cursor.close()
            conn.close()


            return jsonify(df_json)

        elif data_type == 'month':
            # Updated SQL query to filter by state_id, from_date, and to_date
            query = """
            SELECT
            state_id,
            from_date,
            to_date,
            revision_no,
            upload_time,
            file_data
            FROM (
            SELECT
                state_id,
                from_date,
                to_date,
                revision_no,
                upload_time,
                file_data,
                ROW_NUMBER() OVER (PARTITION BY state_id, from_date, to_date ORDER BY upload_time DESC) as rn
            FROM
                month_ahead_file_uploads
            WHERE state_id = %s AND from_date >= %s AND to_date <= %s
            ) sub
            WHERE sub.rn = 1;
            """

            cursor.execute(query, (state_id, input_from_date, input_to_date))
            records = cursor.fetchall()

            df_list = []

            results = []
            columns = ['Date', 'Block', 'Period', 'Forcasted Demand_MW (A)', 'From its own Sources Excl. Renewable_THERMAL_MW', 'From its own Sources Excl. Renewable_GAS_MW','From its own Sources Excl. Renewable_HYDRO_MW', 'From its own Sources Excl. Renewable_TOTAL (B)_MW', 'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 'From Renewable Sources_Other RES (biomass)_MW', 'From Renewable Sources_TOTAL (C)_MW', 'From ISGS & Other LTA & MTOA (D)_MW', 'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 'Total Availability  (F)= (B+C+D+E)_MW', 'Gap between Demand & Availability (G) = (A)-(F)  Surplus(-) / Deficit (+)_MW' , 'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 'Proposed Procurement_Through Power Exchange (I)_MW', 'Shortages after day ahead procurement from market (J) =(G)-(H+I)  Surplus(-) / Deficit (+)_MW', 'Relief through planned restrictions/ rostering/ power cuts (K)_MW','Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 'Reactive Power Forecast_MVar' ]

            for record in records:
                from_date = record[1]  # From Date
                to_date = record[2]    # To Date
                
                # Generate the Date column based on from_date and to_date
                date_range = generate_date_range(from_date, to_date)

                # Ensure the generated date list has the correct length to match the rows
                if len(date_range) != len(record[5]):
                    raise ValueError(f"Generated date range length {len(date_range)} does not match file data length {len(record[5])}")

                # Replace "Invalid Date" with the actual date range
                file_data = record[5]
                for i, row in enumerate(file_data):
                    row[0] = date_range[i]  # Assign the correct date to the first column
                
                result = {
                    "State ID": record[0],
                    "From Date": from_date.isoformat(),
                    "To Date": to_date.isoformat(),
                    "Revision No": record[3],
                    "Upload Time": record[4].isoformat(),
                    "File Data": file_data
                }

                # Create DataFrame from the updated File Data
                df_file_data = pd.DataFrame(result["File Data"], columns=columns)
                df_list.append(df_file_data)

            if not df_list:
                return jsonify({'error': 'No data present for this period'}), 500


            final_df = pd.concat(df_list, ignore_index=True)
            final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst = True)
            df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format
            
            return jsonify(df_json)

    except Exception as e:
        print(f"Some Problem with the API: {e}")
        log_error("forecast data fetch", e)
        return jsonify({'error': 'Database query failed'}), 500




    

@app.route("/api/forecastuploaddate", methods=['GET'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_forecast_upload_date")
def get_forecast_date():
    try:
        today = datetime.now().date()

        # Intraday: Today
        intraday_date = today.strftime("%d %b, %Y")

        # Day Ahead: Tomorrow
        day_ahead_date = (today + timedelta(days=1)).strftime("%d %b, %Y")

        # Week Ahead: Next Monday to Sunday
        next_monday = today + timedelta(days=(7 - today.weekday()))  # 0 is Monday
        next_sunday = next_monday + timedelta(days=6)
        week_ahead_range = f"{next_monday.strftime('%d %b, %Y')} to {next_sunday.strftime('%d %b, %Y')}"
        week_ahead_range = {"from": next_monday.strftime('%d %b, %Y'), "to": next_sunday.strftime('%d %b, %Y')}

        # Month Ahead: 1st to last of next month
        if today.month == 12:
            next_month = 1
            next_month_year = today.year + 1
        else:
            next_month = today.month + 1
            next_month_year = today.year
        month_start = datetime(next_month_year, next_month, 1).date()
        last_day = monthrange(next_month_year, next_month)[1]
        month_end = datetime(next_month_year, next_month, last_day).date()
        month_ahead_range = f"{month_start.strftime('%d %b, %Y')} to {month_end.strftime('%d %b, %Y')}"
        month_ahead_range = {"from": month_start.strftime('%d %b, %Y'), "to": month_end.strftime('%d %b, %Y')}

        # Year Ahead: 1 Apr of next year to 31 Mar of year after that
        next_year = today.year + 1
        year_start = datetime(next_year, 4, 1).date()
        year_end = datetime(next_year + 1, 3, 31).date()
        year_ahead_range = f"{year_start.strftime('%d %b, %Y')} to {year_end.strftime('%d %b, %Y')}"
        year_ahead_range = {"from": year_start.strftime('%d %b, %Y'), "to": year_end.strftime('%d %b, %Y')}

        # Return all
        return jsonify(
            intraday=intraday_date,
            day_ahead=day_ahead_date,
            week_ahead=week_ahead_range,
            month_ahead=month_ahead_range,
            year_ahead=year_ahead_range,
            status="success"
        )

    except Exception as e:
        log_error("forecastuploaddate", e)
        return jsonify(message="There is a problem, please contact SRLDC IT!", status="failure")

##################### LINEFLOWS REPORT



@app.route("/api/reports/lineflows", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_lineflows")
def displayLineFlows():
    try:
        conn2 = psycopg2.connect(
        database="MDPNew", user = 'prasadbabu',
        password = 'BabuPrasad#123', host = '10.0.100.219', port = '5432' 
        )

        jwt_data = get_jwt()

        cur = conn2.cursor()
        input_dat = request.get_json()
        start_date = datetime.strptime(input_dat['params']['fromDate'].replace('"', ''), '%d/%m/%Y').date()
        end_date = datetime.strptime(input_dat['params']['toDate'].replace('"', ''), '%d/%m/%Y').date()



        # Get columns data
        cur.execute("""
            SELECT col_name, out_mwh_columns.formula as formula, col_seq_no, item, col_type, out_mwh_columns.startdatetime as startdatetime, out_mwh_columns.enddatetime as enddatetime
            FROM out_mwh_config join out_mwh_columns on out_mwh_config.id = out_mwh_columns.item_fk_id
            WHERE item = 'LFL_MWH' AND out_mwh_columns.startdatetime <= %s;
        """, (end_date,))
        columns_data = cur.fetchall()
        columns_df = pd.DataFrame(columns_data, columns=['col_name', 'formula', 'col_seq_no', 'item', 'col_type', 'startdatetime', 'enddatetime'])

        # Regular expression pattern
        pattern = r"[-+]?[A-Za-z0-9]+-[A-Za-z0-9]+\s*"
        for i, col in columns_df.iterrows():
            if col['formula']:
                matches = re.findall(pattern, col['formula'])
                if col['col_type'] in ['SINGLE', 'FREQ'] and len(matches) >= 1:
                    columns_df.at[i, 'neg_factor'] = -1 if col['formula'][0] == '-' else 1
                    columns_df.at[i, 'formula'] = matches[0]
                else:
                    columns_df.at[i, 'neg_factor'] = 1

        columns_df = columns_df.sort_values('col_seq_no').copy()

        # print(columns_df)

        # Get fict computed list
        cur.execute("""
            SELECT DISTINCT short_location, startdate, time, mwh
            FROM fict_computation
            WHERE startdate BETWEEN %s AND %s;
        """, (start_date, end_date))
        fict_computed_list = pd.DataFrame(cur.fetchall(), columns=['short_location', 'startdate', 'time', 'mwh'])

        # print(fict_computed_list.head())

        # Get NPC meter data
        cur.execute("""
            SELECT DISTINCT startdate, time, meterno, mwh
            FROM npcmeterdata
            WHERE startdate BETWEEN %s AND %s;
        """, (start_date, end_date))
        queryset_npc = pd.DataFrame(cur.fetchall(), columns=['startdate', 'time', 'meterno', 'mwh'])

        # Get replaced meter data
        cur.execute("""
            SELECT DISTINCT find_meterno, startdate, time, mwh
            FROM replaced_meter_data
            WHERE startdate BETWEEN %s AND %s;
        """, (start_date, end_date))
        replaced_query_set = pd.DataFrame(cur.fetchall(), columns=['find_meterno', 'startdate', 'time', 'mwh'])

        # Get fict meters locations
        cur.execute("""
            SELECT short_location
            FROM create_fictmeter
            WHERE end_date_time >= CURRENT_TIMESTAMP OR end_date_time IS NULL;
        """)
        fict_meters_locations = [loc[0] for loc in cur.fetchall()]

        # Get real meter data
        cur.execute("""
            SELECT meter_number, short_location, start_date_time, end_date_time
            FROM create_meter;
        """)
        realmeter_obj = pd.DataFrame(cur.fetchall(), columns=['meter_number', 'short_location', 'start_date_time', 'end_date_time'])        

        # print(realmeter_obj.head())

            # folder creation
        # directory=create_folder(start_date,end_date)
        entity_df=generate_dataframe_forrange(start_date,end_date)

        for _,col in columns_df.iterrows():
            # below is to check whether column is Single fict like (AP-NEW)
            if col['col_type'] in ['SINGLE'] :
                # this column is there but no fict formula so just keep the column in output csv
                if col['formula']!= None and col['formula']!='' and len(col['formula']) > 0 :
                    # checking fict meter first
                    full_data_df = fict_computed_list[(fict_computed_list["short_location"] == col["formula"]) & (fict_computed_list["startdate"].between(start_date, end_date, inclusive="both"))][["startdate", "time", "mwh"]]
                    full_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': 'mwh1'}, inplace=True)

                    if not full_data_df.empty:
                        #replace None with --
                        full_data_df.fillna('--',inplace=True)
                        entity_df=pd.merge(entity_df, full_data_df, on=['DATE', 'TIME'], how='left')
                        entity_df.rename(columns={'mwh1':col['col_name']} , inplace=True)
                    elif col['formula'] in fict_meters_locations:
                        # fict meter but not computed
                        entity_df=_column(entity_df , col['col_name'] )
                    else:
                        # if it is real meter but replaced for particular week
                        # replaced_meter_data_df=pd.DataFrame(replaced_queryset.filter(find_meterno=col['outmwhcolumns__formula']).values(**{'DATE':F('startdate'),'TIME':F('time'),'mwh1':F('mwh') }),columns=['DATE','TIME','mwh1'] )
                        replaced_meter_data_df = replaced_query_set[(replaced_query_set["find_meterno"] == col["formula"]) & (replaced_query_set["startdate"].between(start_date, end_date, inclusive="both"))][["startdate", "time", "mwh"]]
                        replaced_meter_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': 'mwh1'}, inplace=True)

                        # first get the meterno corresponding short location
                        # meterno_obj=realmeter_obj.filter( Q(short_location=col['outmwhcolumns__formula']) ,( Q(end_date_time__range=[start_date,end_date] )| Q(end_date_time__isnull=True) ) )
                        # realmeter_obj['end_date_time'] = realmeter_obj['end_date_time'].dt.date
                        
                        meterno_obj = realmeter_obj[(realmeter_obj["short_location"] == col['formula']) & ((realmeter_obj["end_date_time"].between(pd.to_datetime(start_date).tz_localize('UTC'), pd.to_datetime(end_date).tz_localize('UTC'), inclusive="both")) | (realmeter_obj["end_date_time"].isnull()) )]

                        real_meter_data_df=pd.DataFrame([],columns=['DATE','TIME'])


                        for m, mtr in meterno_obj.iterrows():
                            # temp_multi_real_df=pd.DataFrame( queryset_npc.filter(meterno=mtr.meter_number).values(**{'DATE':F('startdate'),'TIME':F('time'),'mwh1':F('mwh') }) ,columns=['DATE','TIME','mwh1'])
                            
                            temp_multi_real_df = queryset_npc[queryset_npc["meterno"] == mtr["meter_number"]][["startdate", "time", "mwh"]]
                            temp_multi_real_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': 'mwh1'}, inplace=True)

                            temp_multi_real_df=check_start_enddate_col(temp_multi_real_df ,mtr["start_date_time"] , mtr["end_date_time"],'mwh1')

                            real_meter_data_df=pd.concat([real_meter_data_df,temp_multi_real_df])
                        
                        replaced_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)
                        real_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)

                        if not real_meter_data_df.empty:
                            # Update values in real_meter_data_df with corresponding values from replaced_meter_data_df based on the index (DATE, TIME)
                            real_meter_data_df.update(replaced_meter_data_df)
                            # Reset the index to columns
                            real_meter_data_df.reset_index(inplace=True)
                            
                            # print("ENTITY_DF")
                            # print(entity_df.head())
                            # print("REAL_METER_DF")
                            # print(real_meter_data_df.head())

                            entity_df=pd.merge(entity_df, real_meter_data_df[["DATE", "TIME", "mwh1"]], on=['DATE', 'TIME'], how='left')
                            entity_df.rename(columns={'mwh1':col['col_name']} , inplace=True)
                        elif not replaced_meter_data_df.empty:
                            # if real meter data not present but replaced meter data presents then updat entity_df with only replaced duration
                            # Reset the index to columns
                            replaced_meter_data_df.reset_index(inplace=True)

                            # print("ENTITY_DF")
                            # print(entity_df.head())
                            # print("REPLACED_METER_DF")
                            # print(replaced_meter_df.head())

                            entity_df=pd.merge(entity_df, replaced_meter_data_df, on=['DATE', 'TIME'], how='left')
                            entity_df.rename(columns={'mwh1':col['col_name']} , inplace=True)
                        else:
                            entity_df=_column(entity_df , col['col_name'] )
                    
                    # finally multiplying with negative factor
                    entity_df[col['col_name']]=(entity_df[col['col_name']]) * int(col['neg_factor']) 

                else: 
                    # this else no fict formula but to keep the column in output file to match the srpc format
                    entity_df=_column(entity_df , col['col_name'] )

            elif col['col_type'] in ['MULTI'] :
                short_locations=set(split_string(col['formula']))
                multi_df=generate_dataframe_forrange(start_date,end_date)


                for loct in short_locations:
                    real_meter_data_df=generate_dataframe_forrange(start_date,end_date)
                    real_meter_data_df[loct]=None

                    modified_loc=loct.replace('(','').replace(')','')
                    # fist check if it is fict meter else real meter
                    # multi_fict_data_df=pd.DataFrame( fict_computed_list.filter(short_location= modified_loc , startdate__range=[start_date,end_date]).values(**{'DATE':F('startdate'),'TIME':F('time'), loct:F('mwh') } ) , columns=['DATE','TIME',loct])
                    multi_fict_data_df = fict_computed_list[(fict_computed_list["short_location"] == modified_loc) & (fict_computed_list["startdate"].between(start_date, end_date, inclusive="both"))]
                    multi_fict_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': loct}, inplace=True)

                    if not multi_fict_data_df.empty:
                        multi_fict_data_df.fillna('--',inplace=True)
                        multi_df=pd.merge(multi_df, multi_fict_data_df, on=['DATE', 'TIME'], how='inner')

                    elif modified_loc in fict_meters_locations:
                        multi_df=_column(multi_df , loct )

                    else:
                        # if it is real meter but replaced for particular week
                        # replaced_meter_data_df=pd.DataFrame(replaced_queryset.filter(find_meterno=modified_loc).values(**{'DATE':F('startdate'),'TIME':F('time'),loct:F('mwh') }),columns=['DATE','TIME',loct] )

                        replaced_meter_data_df = replaced_query_set[(replaced_query_set["find_meterno"] == modified_loc) ][["startdate", "time", "mwh"]]
                        replaced_meter_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': loct}, inplace=True)

                        # meterno_obj=realmeter_obj.filter( Q(short_location=modified_loc) , ( Q(end_date_time__range=[start_date,end_date] )| Q(end_date_time__isnull=True) ) )

                        meterno_obj = realmeter_obj[(realmeter_obj["short_location"] == modified_loc) & ((realmeter_obj["end_date_time"].between(pd.to_datetime(start_date).tz_localize('UTC'), pd.to_datetime(end_date).tz_localize('UTC'), inclusive="both")) | (realmeter_obj["end_date_time"].isnull()) )]


                        temp_multi_df1=pd.DataFrame([],columns=['DATE','TIME'])

                        for m, mtr in meterno_obj.iterrows():
                            # temp_multi_real_df=pd.DataFrame( queryset_npc.filter(meterno=mtr.meter_number).values(**{'DATE':F('startdate'),'TIME':F('time'),'mwh1':F('mwh') }) ,columns=['DATE','TIME','mwh1'])
                            
                            temp_multi_real_df = queryset_npc[queryset_npc["meterno"] == mtr["meter_number"]][["startdate", "time", "mwh"]]
                            temp_multi_real_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': loct}, inplace=True)

                            temp_multi_real_df=check_start_enddate_col(temp_multi_real_df ,mtr["start_date_time"] , mtr["end_date_time"],loct)

                            temp_multi_df1=pd.concat([temp_multi_df1,temp_multi_real_df])
                        
                        real_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)
                        temp_multi_df1.set_index(['DATE', 'TIME'], inplace=True)
                        # Must set_index before using update method
                        real_meter_data_df.update(temp_multi_df1)
                        
                        replaced_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)
                        

                        if not real_meter_data_df.empty:
                            # Update values in real_meter_data_df with corresponding values from replaced_meter_data_df based on the index (DATE, TIME)
                            real_meter_data_df.update(replaced_meter_data_df)
                            # Reset the index to columns
                            real_meter_data_df.reset_index(inplace=True)

                            multi_df=pd.merge(multi_df, real_meter_data_df[["DATE", "TIME", loct]], on=['DATE', 'TIME'], how='left')
                        elif not replaced_meter_data_df.empty:
                            # if real meter data not present but replaced meter data presents then updat entity_df with only replaced duration
                            # Reset the index to columns
                            replaced_meter_data_df.reset_index(inplace=True)

                            multi_df=pd.merge(multi_df, replaced_meter_data_df, on=['DATE', 'TIME'], how='left')   
                        else:  
                            multi_df=_column(multi_df , loct )

                if not multi_df.empty:
                    multi_df[col['col_name']] = multi_df.apply(calculate_formula,args=(short_locations, col['formula']), axis=1)
                else: 
                    multi_df=_column(multi_df , col['col_name'] )

                multi_df.drop(columns=short_locations,inplace=True)
                entity_df=pd.merge(entity_df, multi_df, on=['DATE', 'TIME'], how='left')


            else:
                pass
            
            #skip if col_type is DATE,TIME and TOTAL
            if col['col_type'] not in ['DATE','TIME','TOTAL']:
                # if startdatetime is greater than current date then putting -- as mwh1 else original data 
                new_frame_df=entity_df.copy()
                new_frame_df['DATETIME'] = pd.to_datetime(new_frame_df['DATE'].astype(str) + ' ' + new_frame_df['TIME'].astype(str) ,  errors='coerce')
                new_frame_df['DATETIME'] = new_frame_df['DATETIME'].apply(lambda x: x.tz_localize(None))

                temp_timestamp=pd.to_datetime(col['startdatetime']).tz_convert(None)
                temp_endtimestamp=pd.to_datetime(col['enddatetime'],errors='coerce')
                
                # temp_timestamp = temp_timestamp.apply(lambda x: x.tz_localize(None))
                # temp_endtimestamp = temp_endtimestamp.apply(lambda x: x.tz_localize(None))
            
                # print(new_frame_df.dtypes)

                if pd.isnull(temp_endtimestamp):
                    # Handle the invalid timestamps by replacing them with a default value
                    default_endtimestamp = pd.to_datetime('2050-01-01')
                else:
                    default_endtimestamp = temp_endtimestamp.tz_convert(None)

                record_dt_time_condition=new_frame_df['DATETIME'].apply(lambda dt: dt >= temp_timestamp and dt <= default_endtimestamp )

                new_frame_df.loc[~(record_dt_time_condition), col['col_name']] = '--'
                new_frame_df.drop(columns=['DATETIME'],inplace=True)
                entity_df=new_frame_df.copy()

        entity_df=entity_df.fillna('--')
        new_frame_df=entity_df.copy()

        if not new_frame_df.empty:
            new_frame_df.set_index(['DATE', 'TIME'], inplace=True)
            # Specify columns to include in calculations
            columns_to_include = new_frame_df.columns.difference(['DATE', 'TIME'])

            new_frame_df[new_frame_df.columns[2:]] = new_frame_df[new_frame_df.columns[2:]].apply(pd.to_numeric, errors='coerce')

            # Apply the function column-wise on the selected columns
            result_df = new_frame_df[columns_to_include].apply(calculate_column_stats, axis=0)
            try:
                # to rearrange the column order
                result_df = result_df[entity_df.columns[2:]]
            except Exception as e: 
                extractdb_errormsg(e)
            # Transpose the result DataFrame
            result_df = result_df.T
            result_df.reset_index(inplace=True)
            result_df.rename( columns={'Pos_Sum':'Export (MU)','Pos_Max':'Maximum Power Flow (MW)','Neg_Sum':'Import (MU)','Neg_Max':'Maximum Power Flow (MW)(I)' ,'index':'NAME OF THE LINE'},inplace=True ) 
        else:
            result_df=pd.DataFrame(['No data found , Please check'])

        filename='line_flows_'+start_date.strftime('%d-%m-%y')+'.xlsx'
        
        entity_df.set_index('DATE', inplace=True)

        df = entity_df
        df.columns = [col[:-2] if col.endswith('.1') else col for col in df.columns]
        df.drop(columns=['TIME'], inplace=True)
        df = df.applymap(lambda x: 0 if isinstance(x, str) else x)
        def calculate_positive_negative_sum(df, column_name):
            positive_sum = df[df[column_name] > 0][column_name].groupby(level=0).sum()
            negative_sum = df[df[column_name] < 0][column_name].groupby(level=0).sum()
            return positive_sum, negative_sum
        df2=pd.DataFrame()
        positive_sums = {}
        negative_sums = {}
        all_dates = df.index.unique()

        for column_name in df.columns:
            positive_sum, negative_sum = calculate_positive_negative_sum(df, column_name)
            positive_sum = positive_sum.reindex(all_dates, fill_value=0)
            negative_sum = negative_sum.reindex(all_dates, fill_value=0)
            df2[column_name + ' imp'] = positive_sum
            df2[column_name + ' exp'] = negative_sum
        df2= df2.div(1000)
        interreg=["RGDM-CHNDPR imp","RGDM-CHNDPR exp","765KV RCR-SLPR I&II imp","765KV RCR-SLPR I&II exp","HVDC GAJUWAKA imp","HVDC GAJUWAKA exp","HVDC TAL-KLR imp","HVDC TAL-KLR exp","765KV KDG-KLPR I&II imp","765KV KDG-KLPR I&II exp","ANG-SKLM imp","ANG-SKLM exp","WAR-NZB imp","WAR-NZB exp","RGH-PUG imp","RGH-PUG exp","220KV AMBWDI-XLDM imp","220KV AMBWDI-XLDM exp","220KV AMBWDI-PONDA imp","220KV AMBWDI-PONDA exp","765KV WAR-WRNGL I&II.1 imp","765KV WAR-WRNGL I&II.1 exp"]
        column_map ={"BHADRAVATHI IMP":"RGDM-CHNDPR imp","BHADRAVATHI EXP":"RGDM-CHNDPR exp","RCR-SOLPR IMP":"765KV RCR-SLPR I&II imp","RCR-SLPR EXP":"765KV RCR-SLPR I&II exp","GAZUWAKA IMP":"HVDC GAJUWAKA imp","GAZUWAKA EXP":"HVDC GAJUWAKA exp","TAL-KOLAR":"HVDC TAL-KLR imp","TAL-KOLAR EXP":"HVDC TAL-KLR exp","KUDGI-KOLAPUR IMP":"765KV KDG-KLPR I&II imp","KUDGI-KOLAPUR EXP":"765KV KDG-KLPR I&II exp","Angul-Srikakulam IMP":"ANG-SKLM imp","Angul-Srikakulam EXP":"ANG-SKLM exp","765kV Wardha-Nizamabad 1 & 2 IMP":"WAR-NZB imp","765kV Wardha-Nizamabad 1 & 2 EXP":"WAR-NZB exp","Raighar-Pugalur HVDC IMP":"RGH-PUG imp","Raighar-Pugalur HVDC EXP":"RGH-PUG exp","Ambewadi-Xeldom IMP":"220KV AMBWDI-XLDM imp","Ambewadi-Xeldom EXP":"220KV AMBWDI-XLDM exp","Ambewadi-Ponda IMP":"220KV AMBWDI-PONDA imp","Ambewadi-Ponda EXP":"220KV AMBWDI-PONDA exp","765KV WAR-WRNGL I&II IMP":"765KV WAR-WRNGL I&II imp","765KV WAR-WRNGL I&II EXP":"765KV WAR-WRNGL I&II exp"}

        # Load the workbook and select the desired sheet
        wb = openpyxl.load_workbook(os.path.join(shared_drive_path,"Lineflows Report work","LineFlows.xlsx"))
        # os.path.join(shared_drive_path
        sheet = wb['SUMMARY']

        # Clear the contents of the range A2:W32
        for row in sheet.iter_rows(min_row=2, max_row=32, min_col=1, max_col=23):  # A2:W32 is 2nd row to 32nd, and columns A to W (23 columns)
            for cell in row:
                cell.value = None

        # Assuming df2 is your DataFrame, converting the index to datetime
        index_datetime = pd.to_datetime(df2.index)

        # Writing dates to column A starting from row 2
        for row_num, date in enumerate(index_datetime.strftime('%d-%m-%Y'), start=2):
            sheet.cell(row=row_num, column=1, value=date)

        # Prepare JSON output
        json_output = []

        # Assuming column_map is a dictionary that maps your output columns to DataFrame columns
        for output_column, df_column in column_map.items():
            try:
                # Find the column index for the output column in the first row
                output_column_index = None
                for col_num, cell in enumerate(sheet[1], start=1):  # Sheet row 1
                    if cell.value == output_column:
                        output_column_index = col_num
                        break
                
                if output_column_index is None:
                    continue  # Skip if the column wasn't found
                
                # Writing data from the DataFrame to the corresponding column
                data = df2[df_column].values.tolist()
                for row_num, value in enumerate(data, start=2):
                    sheet.cell(row=row_num, column=output_column_index, value=value if pd.notna(value) else None)

                # Prepare the JSON output data
                data_entries = [{'x': date.strftime('%Y-%m-%d'), 'y': round(value,4) if pd.notna(value) else ''} for date, value in zip(df2.index, df2[df_column])]
                json_output.append({
                    'name': df_column,
                    'data': data_entries
                })
            except Exception as e:
                print(e)
                continue

        # Save the workbook with a timestamped name
        file_name = os.path.join(shared_drive_path,"Lineflows Report work","lineflows_report_{0}.xlsx".format(datetime.now().strftime('%Y%m%d_%H%M%S')))
        
        wb.save(file_name)
        wb.close()

        # Return the response in the same format as the original
        return jsonify({
            'status': 'success',
            "message": "Report generated successfully!",
            "file_link": file_name,
            "data": json_output,
            "title": "Lineflows for the period between {0} and {1}".format(input_dat['params']['fromDate'], input_dat['params']['toDate'])
        }), 200



    except Exception as e:
        log_error("displaylineflows", e)
        cursor.close()
        return jsonify(message="There is some problem in Fetching the data! Please contact SRLDC IT", status="failure")
        # return jsonify({"status": "failure", "error": str(e), "file": fname, "line": exc_tb.tb_lineno}), 500




        

@app.route("/api/reports/downloadlineflows", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_lineflows")
def downloadLineFlows():
    try:
        # Generate a unique filename with a timestamp
        download_link_json = request.get_json()
        download_link = download_link_json['downloadLink']

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'lineflows_report_{timestamp}.xlsx'
        file_path = download_link

        jwt_data = get_jwt()

        # Example logic for saving the Excel file
        # wb.save(file_path)
        # wb.close()

        # Directly send the file in the response
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloadlineflows", e)
        cursor.close()
        return jsonify(message="There is some problem in downloading the file! Please contact SRLDC IT", status="failure")
        # return jsonify({'status': 'failure', 'error': str(e)}), 500


@app.route("/api/reports/fetchmdpdescription", methods=["GET"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_mdp_description")
def fetchMdpDescription():
    try:
        # Establish database connection
        conn = psycopg2.connect(
            **mdp_db_params
        )
        cur = conn.cursor()

        # Execute the query to fetch distinct descriptions
        query = "SELECT DISTINCT description FROM create_fictmeter;"
        cur.execute(query)

        jwt_data = get_jwt()

        # Fetch all results
        result = cur.fetchall()

        # Close the cursor and connection
        cur.close()
        conn.close()

        # Format the results as a list of descriptions
        descriptions = [row[0] for row in result]


        return jsonify(data=descriptions, message="Successfully fetched descriptions!", status="success")

    except Exception as e:
        # Log the error and return a failure message
        log_error("fetchmdpdescription", e)
        return jsonify(message="There is some problem in fetching the names! Please contact SRLDC IT", status="failure")




@app.route("/api/reports/mdpdescriptiondata", methods=["POST"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_mdp_description_data")
def MdpDescriptionData():
    try:
        # Parse request data
        data = request.get_json()['params']
        element_list = data.get('elementList')  # List of elements
        input_from_date = data.get('fromDate')  # From date in DD/MM/YYYY format
        input_to_date = data.get('toDate')  # To date in DD/MM/YYYY format

        jwt_data = get_jwt()

        # Convert input dates to YYYY-MM-DD format
        from_date = datetime.strptime(input_from_date, "%d/%m/%Y").strftime("%Y-%m-%d")
        to_date = datetime.strptime(input_to_date, "%d/%m/%Y").strftime("%Y-%m-%d")

        # Establish database connection
        conn = psycopg2.connect(**mdp_db_params)
        cur = conn.cursor()

        # Prepare results for each element
        results = []
        for element in element_list:
            query = """
                SELECT
                    TO_CHAR(fict_computation."startdate" + fict_computation."time"::interval, 'YYYY-MM-DD HH24:MI') AS timestamp,
                    fict_computation."mw"
                FROM
                    public.fict_computation
                WHERE
                    fict_computation."short_location" = (
                        SELECT DISTINCT short_location FROM public.create_fictmeter WHERE description = %s
                    )
                    AND fict_computation."startdate" BETWEEN %s AND %s
                ORDER BY fict_computation."startdate", fict_computation."time"
            """
            cur.execute(query, (element, from_date, to_date))
            data_records = cur.fetchall()

            # Format the data
            formatted_data = [{'x': record[0], 'y': round(record[1], 2)} for record in data_records]
            results.append({'name': element, 'data': formatted_data})

        # Close the cursor and connection
        cur.close()
        conn.close()

        # Create a comma-separated list of elements for the title
        comma_separated_list = ', '.join(element_list)

        # Return the formatted results
        return jsonify(
            results=results,
            message="API Fetch Successful!",
            title="Showing Data for {} for the period between {} and {}".format(
                comma_separated_list, input_from_date, input_to_date
            ),
            status="success"
        )

    except Exception as e:
        # Log the error and return a failure message
        print(f"Error in mdpdescriptiondata: {e}")
        return jsonify(
            message="There is some problem in fetching the data! Please contact SRLDC IT",
            status="failure"
        )




@app.route("/api/reports/sendmergeddemandforecastmail", methods=["GET"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("send_merged_forecast_mail")
def sendMergedDemandForecastMail():
    try:
        # Use subprocess to execute the script
        result = subprocess.run(
            ["python", "D:\\Demand Forecasting Page\\custom scripts\\demand_forecast_combined_and_mail.py"],  # Command to execute the script
            capture_output=True,         # Capture the output
            text=True                    # Decode output as text
        )

        # Check the return code to ensure the script ran successfully
        if result.returncode == 0:
            print(f"Script output: {result.stdout}")
            return jsonify(
                message="Script successfully executed and Mail sent!",
                status="success"
            )
        else:
            print(f"Script error: {result.stderr}")
            return jsonify(
                message="Script executed but encountered some issues.",
                status="partial_success"
            )
    except Exception as e:
        print(f"Error executing script: {e}")
        return jsonify(
            message="There is some problem in fetching the data! Please contact SRLDC IT",
            status="failure"
        )
    

# Updated API for forecast mail status
@app.route('/api/reports/forecastmailstatus', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_mail_status")
def api_forecast_mail_status():
    from forecast_mail_status_utils import forecast_mail_status
    data = request.get_json(force=True)
    forecast_type = data.get('type')
    from_date = data.get('from_date')
    to_date = data.get('to_date')
    print(data)
    # Validate input
    if forecast_type not in ['day','week', 'month', 'year']:
        return jsonify({'status': 'failure', 'message': 'Invalid forecast type'}), 400
    if not from_date or not to_date:
        return jsonify({'status': 'failure', 'message': 'from_date and to_date required'}), 400
    result = forecast_mail_status(forecast_type, from_date, to_date)
    return jsonify(result)






@app.route('/api/reports/downloadconsolidatedfile', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("download_consolidated_file")
def download_consolidated_file():
    
    """
    Generate consolidated forecast file (day/week/month/year)
    """
    from forecast_file_utils import generate_consolidated_forecast_file
    try:
        data = request.get_json(force=True)
        forecast_type = data.get("type")
        from_date = data.get("from_date")
        to_date = data.get("to_date")

        if not forecast_type or not from_date or not to_date:
            return jsonify({"status": "failure", "message": "Missing required parameters"}), 400

        result = generate_consolidated_forecast_file(forecast_type, from_date, to_date)

        if result["status"] == "failure":
            return jsonify(result), 400

        return send_file(result["file_path"], as_attachment=True)

    except Exception as e:
        return jsonify({
            "status": "failure",
            "message": f"Internal error: {str(e)}",
            "trace": traceback.format_exc()
        }), 500




from psycopg2.extras import execute_values

@app.route('/api/share-allocation/upload', methods=['POST'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("upload_share_allocation")
def upload_share_allocation():

    allocation_date = request.form.get('allocation_date')
    json_data = json.loads(request.files['dataFile'].read())

    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    metric_map = {
        'Allocated %': 'Allocated_Percent',
        'Unallocated %': 'Unallocated_Percent',
        'Fp %': 'Fp_Percent',
        'Reserved MW': 'Reserved_MW',
        'Offbar %': 'Offbar_Percent'
    }

    records = []

    for row in json_data:
        metric_type = metric_map[row[0]]
        seller = row[1]
        buyer = row[2]

        for block_no, value in enumerate(row[3:], start=1):
            records.append((
                allocation_date,
                seller,
                buyer,
                metric_type,
                block_no,
                value
            ))

    sql = """
        INSERT INTO ra_share_allocation
        (
            allocation_date,
            seller_acronym,
            buyer_acronym,
            metric_type,
            block_no,
            value
        )
        VALUES %s
        ON CONFLICT (allocation_date, seller_acronym, buyer_acronym, metric_type, block_no)
        DO UPDATE SET
            value = EXCLUDED.value,
            created_at = now()
    """

    execute_values(
        cur,
        sql,
        records,
        page_size=5000
    )

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "status": "success",
        "message": f"Share allocation uploaded successfully ({len(records)} blocks)"
    })


@app.route('/api/share-allocation/current', methods=['GET'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_share_allocation")
def get_current_share_allocation():

    try:
        conn = psycopg2.connect(**db_params)
        cur = conn.cursor()

        # Latest allocation date
        cur.execute("""
            SELECT allocation_date, MAX(created_at)
            FROM ra_share_allocation
            GROUP BY allocation_date
            ORDER BY allocation_date DESC
            LIMIT 1
        """)
        row = cur.fetchone()
        if not row:
            return jsonify({"status": "failure", "message": "No data"}), 404

        allocation_date, uploaded_at = row

        cur.execute("""
            SELECT seller_acronym,
                   buyer_acronym,
                   metric_type,
                   block_no,
                   ABS(value)
            FROM ra_share_allocation
            WHERE allocation_date = %s
            ORDER BY seller_acronym, buyer_acronym, metric_type, block_no
        """, (allocation_date,))
        records = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    # ======================================================
    # HEADERS (JSpreadsheet)
    # ======================================================
    time_blocks = generate_15min_time_blocks()

    headers = ["Type", "Seller Acronym", "Buyer Acronym"] + time_blocks

    nested_headers = [
        [{"title": "", "colspan": 3}] +
        [{"title": str(i), "colspan": 1} for i in range(1, 97)]
    ]

    columns = (
        [{"width": 160, "readOnly": True}] * 3 +
        [{
            "width": 90,
            "readOnly": False,
            "type": "numeric",
            "mask": "0.0000"
        }] * 96
    )

    metric_map = {
        "Allocated_Percent": "Allocated %",
        "Unallocated_Percent": "Unallocated %",
        "Fp_Percent": "Fp %",
        "Reserved_MW": "Reserved MW",
        "Offbar_Percent": "Offbar %"
    }

    # ======================================================
    # SHARE ALLOCATION MATRIX
    # ======================================================
    data_map = {}

    for seller, buyer, metric, block_no, value in records:
        label = metric_map.get(metric, metric)
        key = (label, seller, buyer)

        if key not in data_map:
            data_map[key] = [label, seller, buyer] + [0.0] * 96

        data_map[key][block_no + 2] = float(value)

    # ======================================================
    # ISGS SHARE CALCULATION (NEW)
    # ======================================================
    STATE_DIC = {
        "AP": ["APTRANSCO"],
        "TG": ["TELANGANA"],
        "KA": ["KPTCL"],
        "TN": ["TNEB", "TN_STATENETWORK"],
        "KL": ["KSEB"],
        "PN": ["PONDY"]
    }

    # isgs_share[seller][state] = value
    isgs_share = {}

    for seller, buyer, metric, block_no, value in records:

        # ONLY Allocated %
        if metric != "Allocated_Percent":
            continue

        # ONLY first block (00:00–00:15)
        if block_no != 1:
            continue

        if seller not in isgs_share:
            isgs_share[seller] = {st: 0.0 for st in STATE_DIC}

        for state, buyers in STATE_DIC.items():
            if buyer in buyers:
                isgs_share[seller][state] += float(value)

    # ======================================================
    # RESPONSE
    # ======================================================
    return jsonify({
        "status": "success",
        "headers": headers,
        "nestedHeaders": nested_headers,
        "columns": columns,
        "rows": list(data_map.values()),
        "isgs_share": isgs_share,   #  NEW
        "applicable_from": str(allocation_date),
        "uploaded_at": uploaded_at.strftime("%d-%m-%Y %H:%M:%S")
    })


@app.route("/api/share-allocation/download-format", methods=["GET"])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_share_allocation")
def download_share_allocation_format():
    try:
        # Base directory where formats are stored

        file_path = os.path.join(
            shared_drive_path,
            "FORMATS",
            "SR_ShareAllocation_format.csv"   #  your format file
        )

        if not os.path.exists(file_path):
            return jsonify({
                "status": "failure",
                "message": "Format file not found on server"
            }), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name="SR_Share_Allocation_Format.csv",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        log_error("download_share_allocation_format", e)
        return jsonify({
            "status": "failure",
            "message": str(e)
        }), 500

@app.route('/api/share-allocation/download-file', methods=['GET'])
@jwt_required(locations=["cookies"])
@session_token_required
@token_required
@rbac_required("view_share_allocation")
def download_share_allocation():

    allocation_date = request.args.get('allocation_date')
    if not allocation_date:
        return jsonify({"error": "allocation_date is required"}), 400

    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            metric_type,
            seller_acronym,
            buyer_acronym,
            block_no,
            value
        FROM ra_share_allocation
        WHERE allocation_date = %s
        ORDER BY
            seller_acronym,
            buyer_acronym,
            metric_type,
            block_no
    """, (allocation_date,))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        return jsonify({"error": "No data found"}), 404

    # =====================================================
    # FIX 1: Metric ordering EXACTLY as Excel expects
    # =====================================================
    METRIC_ORDER = [
        "Allocated %",
        "Unallocated %",
        "Fp %",
        "Reserved MW",
        "Offbar %"
    ]

    # DB → Excel label mapping (IMPORTANT)
    METRIC_MAP = {
        "Allocated_Percent": "Allocated %",
        "Unallocated_Percent": "Unallocated %",
        "Fp_Percent": "Fp %",
        "Reserved_MW": "Reserved MW",
        "Offbar_Percent": "Offbar %"
    }

    from collections import defaultdict

    # matrix[(seller,buyer)][metric] = [96 values]
    matrix = defaultdict(lambda: defaultdict(lambda: [0.0] * 96))

    for metric, seller, buyer, block, value in rows:
        excel_metric = METRIC_MAP.get(metric, metric)
        matrix[(seller, buyer)][excel_metric][block - 1] = float(value)

    # =====================================================
    # FIX 2: Use YOUR time-block generator
    # =====================================================
    time_blocks = generate_15min_time_blocks()  # 96 labels

    # =====================================================
    # CSV CREATION
    # =====================================================
    import csv
    from io import StringIO
    from flask import make_response

    output = StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Type",
        "Seller Acronym",
        "Buyer Acronym",
        *time_blocks
    ])

    # Data rows
    for (seller, buyer), metric_data in matrix.items():
        for metric in METRIC_ORDER:
            writer.writerow([
                metric,
                seller,
                buyer,
                *metric_data.get(metric, [0.0] * 96)
            ])

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = (
        f"attachment; filename=SR_Share_Allocation_{allocation_date}.csv"
    )
    response.headers["Content-Type"] = "text/csv"

    return response


if __name__ == '__main__':
    app.run(host='0.0.0.0',debug=True)






